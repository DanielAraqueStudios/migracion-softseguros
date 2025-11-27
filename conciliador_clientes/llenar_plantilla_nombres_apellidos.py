
import pandas as pd
from openpyxl import load_workbook
import os

# Rutas de archivos
informe_path = os.path.join(os.path.dirname(__file__), 'data_celer', 'InformedePersonas CELER.xlsx')
plantilla_path = os.path.join(os.path.dirname(__file__), 'plantilla', 'PLANTILLA DE SOTSEGUROS.xlsx')
salida_path = os.path.join(os.path.dirname(__file__), 'plantilla', 'PLANTILLA_DE_SOTSEGUROS_NOMBRES_APELLIDOS.xlsx')

# Leer informe de personas usando la fila 3 como encabezado
informe_df = pd.read_excel(informe_path, header=3)

# Extraer columna de nombre completo
if 'Nombre' not in informe_df.columns:
    raise ValueError('No se encontró la columna "Nombre" en el informe.')

def separar_nombre_apellido(nombre_completo):
    partes = str(nombre_completo).strip().split()
    if len(partes) < 3:
        return ' '.join(partes[:-1]), partes[-1] if partes else ('', '')
    # Asume que los dos últimos son apellidos
    return ' '.join(partes[:-2]), ' '.join(partes[-2:])

nombres = []
apellidos = []
for nombre_completo in informe_df['Nombre']:
    nombre, apellido = separar_nombre_apellido(nombre_completo)
    nombres.append(nombre)
    apellidos.append(apellido)

# Cargar plantilla
wb = load_workbook(plantilla_path)
ws = wb.active

# Escribir nombres y apellidos en la plantilla
for idx, (nombre, apellido) in enumerate(zip(nombres, apellidos), start=2):
    ws.cell(row=idx, column=1, value=nombre)
    ws.cell(row=idx, column=2, value=apellido)

wb.save(salida_path)
print(f'Plantilla llenada y guardada en: {salida_path}')
