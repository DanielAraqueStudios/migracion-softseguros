"""
Asignación Automática de Género
=================================
Este script asigna automáticamente el género (M/F) basándose en el primer nombre
de la persona. Los NITs (empresas) quedan sin género.
"""

import pandas as pd
import re
from pathlib import Path
from datetime import datetime
import logging
from openpyxl.styles import Font, PatternFill, Alignment

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class AsignadorGenero:
    """Asigna género automáticamente basándose en nombres"""
    
    def __init__(self, archivo_entrada):
        self.archivo_entrada = archivo_entrada
        self.df = None
        self.estadisticas = {
            'masculinos': 0,
            'femeninos': 0,
            'nits': 0,
            'ambiguos': 0,
            'desconocidos': 0
        }
        
        # Base de datos de nombres colombianos más comunes
        self.nombres_masculinos = {
            'JUAN', 'CARLOS', 'JOSE', 'LUIS', 'MIGUEL', 'JORGE', 'PEDRO', 'FRANCISCO',
            'JESUS', 'ANTONIO', 'FERNANDO', 'MANUEL', 'RAFAEL', 'DAVID', 'DANIEL',
            'RICARDO', 'ROBERTO', 'EDUARDO', 'ANDRES', 'SERGIO', 'JAVIER', 'OSCAR',
            'ALBERTO', 'GUSTAVO', 'RAUL', 'CESAR', 'ARTURO', 'MARTIN', 'EDGAR',
            'GERMAN', 'ALVARO', 'RODRIGO', 'PABLO', 'ENRIQUE', 'HECTOR', 'RAMON',
            'MARIO', 'DIEGO', 'JAIME', 'VICTOR', 'ALEJANDRO', 'SANTIAGO', 'SEBASTIAN',
            'CAMILO', 'FELIPE', 'NICOLAS', 'MATEO', 'SAMUEL', 'JULIAN', 'ADRIAN',
            'LEONARDO', 'CRISTIAN', 'MAURICIO', 'FABIAN', 'GUILLERMO', 'JOAQUIN',
            'ALFREDO', 'ERNESTO', 'GUILLERMO', 'FERNANDO', 'RUBEN', 'NELSON',
            'WILSON', 'ANDERSON', 'JHON', 'ALEXANDER', 'JHONATAN', 'JONATHAN',
            'FREDY', 'FERNEY', 'JHONNY', 'EDISON', 'EDILSON', 'WILMAR', 'Elder',
            'YESID', 'YEISON', 'DIDIER', 'JAIR', 'JHOAN', 'BRAYAN', 'DUVAN',
            'ELDER', 'EDER', 'JEISON', 'JAIDER', 'YAIR', 'YEFERSON', 'JHONATHAN'
        }
        
        self.nombres_femeninos = {
            'MARIA', 'ANA', 'CARMEN', 'ROSA', 'MARTHA', 'LUCIA', 'PATRICIA', 'SANDRA',
            'GLORIA', 'CLAUDIA', 'DIANA', 'NANCY', 'ANGELA', 'ADRIANA', 'MONICA',
            'ELIZABETH', 'LILIANA', 'BEATRIZ', 'CLARA', 'HELENA', 'ISABEL', 'PAULA',
            'ANDREA', 'CAROLINA', 'NATALIA', 'CATALINA', 'LAURA', 'DANIELA', 'PAOLA',
            'MARCELA', 'ALEJANDRA', 'VALERIA', 'CAMILA', 'SOFIA', 'VALENTINA',
            'MARIANA', 'ISABELLA', 'GABRIELA', 'JESSICA', 'TATIANA', 'YOLANDA',
            'OLGA', 'PILAR', 'TERESA', 'AMPARO', 'CECILIA', 'CONSUELO', 'DORA',
            'ESPERANZA', 'GRACIELA', 'INES', 'LEONOR', 'LIGIA', 'LUZ', 'MAGNOLIA',
            'MERCEDES', 'MYRIAM', 'NELLY', 'NUBIA', 'ROCIO', 'SILVIA', 'SONIA',
            'STELLA', 'VICTORIA', 'VIVIANA', 'YANETH', 'YENNY', 'YUDY', 'ALBA',
            'BLANCA', 'DORIS', 'EDITH', 'ELIANA', 'FABIOLA', 'FLOR', 'FRANCIA',
            'JENNY', 'JOHANNA', 'LEIDY', 'LINA', 'LIZETH', 'MARIBEL', 'MILENA',
            'MIRIAM', 'SHIRLEY', 'XIOMARA', 'YASMIN', 'YULIANA', 'ZULAY'
        }
        
        # Nombres ambiguos (pueden ser M o F)
        self.nombres_ambiguos = {
            'GUADALUPE', 'TRINIDAD', 'REFUGIO', 'CONCEPCION', 'ROSARIO',
            'JESUS', 'ANGEL', 'GUADALUPE'
        }
    
    def cargar_datos(self):
        """Carga el archivo Excel"""
        logger.info(f"Cargando archivo: {self.archivo_entrada}")
        try:
            self.df = pd.read_excel(self.archivo_entrada)
            logger.info(f"✅ {len(self.df)} registros cargados")
            
            # Crear columna GÉNERO si no existe
            if 'GÉNERO' not in self.df.columns:
                self.df['GÉNERO'] = ''
            
            return True
        except Exception as e:
            logger.error(f"❌ Error cargando archivo: {e}")
            return False
    
    def extraer_primer_nombre(self, nombre_completo):
        """Extrae el primer nombre de un nombre completo"""
        if pd.isna(nombre_completo) or nombre_completo == '':
            return None
        
        # Normalizar: mayúsculas, sin espacios extras
        nombre_str = str(nombre_completo).strip().upper()
        
        # Extraer primera palabra
        partes = nombre_str.split()
        if len(partes) > 0:
            return partes[0]
        return None
    
    def determinar_genero(self, nombre, tipo_doc):
        """Determina el género basándose en el nombre y tipo de documento"""
        # Si es NIT, no asignar género (empresas)
        if tipo_doc == 'NIT':
            return ''
        
        # Si no hay nombre, devolver vacío
        if not nombre:
            return 'REVISAR'
        
        primer_nombre = self.extraer_primer_nombre(nombre)
        
        if not primer_nombre:
            return 'REVISAR'
        
        # Verificar si es ambiguo
        if primer_nombre in self.nombres_ambiguos:
            return 'REVISAR'
        
        # Verificar en base de datos de nombres
        if primer_nombre in self.nombres_masculinos:
            return 'M'
        elif primer_nombre in self.nombres_femeninos:
            return 'F'
        else:
            # Intentar identificar por terminaciones comunes
            if primer_nombre.endswith(('A', 'IS', 'IDAD', 'IEL', 'ETH')):
                # Excepciones masculinas que terminan en A
                excepciones_masc = {'JOSHUA', 'GARCIA', 'ELISHA', 'EZRA'}
                if primer_nombre in excepciones_masc:
                    return 'M'
                return 'F'
            elif primer_nombre.endswith(('O', 'N', 'R', 'S', 'L', 'D', 'X')):
                return 'M'
            else:
                return 'REVISAR'
    
    def asignar_generos(self):
        """Asigna género a todos los registros"""
        logger.info("\n=== ASIGNANDO GÉNEROS ===")
        
        casos_revisar = []
        
        for idx, row in self.df.iterrows():
            tipo_doc = row['TIPO DE DOCUMENTO']
            nombres = row['NOMBRES']
            
            genero = self.determinar_genero(nombres, tipo_doc)
            self.df.at[idx, 'GÉNERO'] = genero
            
            # Actualizar estadísticas
            if genero == 'M':
                self.estadisticas['masculinos'] += 1
            elif genero == 'F':
                self.estadisticas['femeninos'] += 1
            elif genero == '':
                self.estadisticas['nits'] += 1
            elif genero == 'REVISAR':
                self.estadisticas['desconocidos'] += 1
                primer_nombre = self.extraer_primer_nombre(nombres)
                casos_revisar.append({
                    'fila': idx + 2,
                    'documento': row['NÚMERO DE DOCUMENTO'],
                    'tipo_doc': tipo_doc,
                    'nombres': nombres,
                    'primer_nombre': primer_nombre
                })
        
        logger.info(f"\n📊 ESTADÍSTICAS:")
        logger.info(f"   Masculinos (M): {self.estadisticas['masculinos']}")
        logger.info(f"   Femeninos (F): {self.estadisticas['femeninos']}")
        logger.info(f"   NITs (vacío): {self.estadisticas['nits']}")
        logger.info(f"   Por revisar: {self.estadisticas['desconocidos']}")
        
        if casos_revisar:
            logger.info(f"\n⚠️  CASOS PARA REVISAR MANUALMENTE ({len(casos_revisar)}):")
            for caso in casos_revisar[:10]:  # Mostrar primeros 10
                logger.info(f"   Fila {caso['fila']}: {caso['primer_nombre']} - {caso['nombres']}")
            if len(casos_revisar) > 10:
                logger.info(f"   ... y {len(casos_revisar) - 10} casos más")
        
        return casos_revisar
    
    def generar_archivo_con_generos(self, ruta_salida='./data/output/04_actualizaciones'):
        """Genera el archivo Excel con géneros asignados"""
        logger.info("\n=== GENERANDO ARCHIVO CON GÉNEROS ===")
        
        Path(ruta_salida).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_salida = f"{ruta_salida}/CLIENTES_SOFTSEGUROS_CON_GENEROS_{timestamp}.xlsx"
        
        # Guardar Excel con formato
        with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
            self.df.to_excel(writer, sheet_name='CLIENTES', index=False)
            
            worksheet = writer.sheets['CLIENTES']
            self._aplicar_formato_header(worksheet)
            self._resaltar_generos(worksheet)
            self._ajustar_ancho_columnas(worksheet)
        
        logger.info(f"✅ Archivo generado: {archivo_salida}")
        return archivo_salida
    
    def generar_reporte_generos(self, casos_revisar, ruta_salida='./data/output/04_actualizaciones'):
        """Genera reporte detallado de asignación de géneros"""
        logger.info("\n=== GENERANDO REPORTE DE GÉNEROS ===")
        
        Path(ruta_salida).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_reporte = f"{ruta_salida}/REPORTE_GENEROS_{timestamp}.xlsx"
        
        with pd.ExcelWriter(archivo_reporte, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            df_resumen = pd.DataFrame([
                ['Total de registros', len(self.df)],
                ['Masculinos (M)', self.estadisticas['masculinos']],
                ['Femeninos (F)', self.estadisticas['femeninos']],
                ['NITs (vacío)', self.estadisticas['nits']],
                ['Por revisar', self.estadisticas['desconocidos']],
                ['', ''],
                ['% Asignados automáticamente', 
                 f"{((self.estadisticas['masculinos'] + self.estadisticas['femeninos']) / len(self.df) * 100):.1f}%"]
            ], columns=['Concepto', 'Cantidad'])
            
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja 2: Casos para revisar
            if casos_revisar:
                df_revisar = pd.DataFrame(casos_revisar)
                df_revisar.to_excel(writer, sheet_name='Para_Revisar', index=False)
            
            # Aplicar formato
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                self._aplicar_formato_header(worksheet)
                self._ajustar_ancho_columnas(worksheet)
        
        logger.info(f"✅ Reporte generado: {archivo_reporte}")
        return archivo_reporte
    
    def _aplicar_formato_header(self, worksheet):
        """Aplica formato a los encabezados"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _resaltar_generos(self, worksheet):
        """Resalta la columna de género con colores"""
        # Encontrar columna GÉNERO
        genero_col = None
        for idx, cell in enumerate(worksheet[1], 1):
            if cell.value == 'GÉNERO':
                genero_col = idx
                break
        
        if not genero_col:
            return
        
        # Colores para géneros
        azul_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        rosa_fill = PatternFill(start_color="F9E6F2", end_color="F9E6F2", fill_type="solid")
        amarillo_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=genero_col)
            if cell.value == 'M':
                cell.fill = azul_fill
            elif cell.value == 'F':
                cell.fill = rosa_fill
            elif cell.value == 'REVISAR':
                cell.fill = amarillo_fill
    
    def _ajustar_ancho_columnas(self, worksheet):
        """Ajusta el ancho de las columnas automáticamente"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Función principal"""
    # Archivo de entrada (el último actualizado)
    archivo_entrada = './data/output/CLIENTES_SOFTSEGUROS_ACTUALIZADO_20251107_155237.xlsx'
    
    logger.info("="*60)
    logger.info("ASIGNACIÓN AUTOMÁTICA DE GÉNERO")
    logger.info("="*60)
    
    # Crear asignador
    asignador = AsignadorGenero(archivo_entrada)
    
    # Cargar datos
    if not asignador.cargar_datos():
        return
    
    # Asignar géneros
    casos_revisar = asignador.asignar_generos()
    
    # Generar archivos de salida
    archivo_con_generos = asignador.generar_archivo_con_generos()
    archivo_reporte = asignador.generar_reporte_generos(casos_revisar)
    
    logger.info("\n" + "="*60)
    logger.info("✅ PROCESO COMPLETADO")
    logger.info("="*60)
    logger.info(f"📁 Archivo con géneros: {archivo_con_generos}")
    logger.info(f"📊 Reporte: {archivo_reporte}")


if __name__ == "__main__":
    main()
