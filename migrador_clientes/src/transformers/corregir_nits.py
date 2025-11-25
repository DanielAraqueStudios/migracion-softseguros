"""
Corrección de Formato de NITs - CLIENTES SOFTSEGUROS
====================================================
Este script corrige el formato de los NITs agregando el guión verificador
cuando falta, y genera un nuevo archivo con los datos corregidos.
"""

import pandas as pd
import re
from pathlib import Path
from datetime import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class CorreccionNITs:
    """Corrige el formato de NITs en el archivo de CLIENTES SOFTSEGUROS"""
    
    def __init__(self, archivo_entrada):
        self.archivo_entrada = archivo_entrada
        self.df = None
        self.correcciones = []
        
    def cargar_datos(self):
        """Carga el archivo Excel"""
        logger.info(f"Cargando archivo: {self.archivo_entrada}")
        try:
            self.df = pd.read_excel(self.archivo_entrada)
            logger.info(f"✅ {len(self.df)} registros cargados")
        except Exception as e:
            logger.error(f"❌ Error al cargar archivo: {e}")
            raise
    
    def calcular_digito_verificacion(self, nit_sin_dv):
        """
        Calcula el dígito de verificación de un NIT colombiano
        usando el algoritmo oficial de la DIAN
        """
        try:
            # Remover caracteres no numéricos
            nit_str = re.sub(r'\D', '', str(nit_sin_dv))
            
            if not nit_str:
                return None
            
            # Secuencia de multiplicadores
            multiplicadores = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
            
            # Tomar los multiplicadores necesarios (de derecha a izquierda)
            nit_len = len(nit_str)
            if nit_len > len(multiplicadores):
                return None
            
            multiplicadores_usar = multiplicadores[-nit_len:]
            
            # Calcular suma
            suma = 0
            for i, digito in enumerate(nit_str):
                suma += int(digito) * multiplicadores_usar[i]
            
            # Calcular dígito de verificación
            residuo = suma % 11
            
            if residuo == 0 or residuo == 1:
                dv = residuo
            else:
                dv = 11 - residuo
            
            return str(dv)
            
        except Exception as e:
            logger.warning(f"Error calculando DV para {nit_sin_dv}: {e}")
            return None
    
    def corregir_formato_nit(self, nit_original, tipo_doc):
        """
        Corrige el formato de un NIT agregando el guión entre el número y el dígito verificador.
        El formato correcto es: últimos 8 o 9 dígitos del NIT, guión, último dígito.
        Ejemplo: 900310074 → 90031007-4 (no 900310074-1)
        """
        if tipo_doc != 'NIT' or pd.isna(nit_original):
            return nit_original, False
        
        nit_str = str(nit_original).strip()
        
        # Si ya tiene el formato correcto (números-dígito), no hacer nada
        if re.match(r'^\d+-\d$', nit_str):
            return nit_str, False
        
        # Extraer solo números del NIT
        nit_numeros = re.sub(r'\D', '', nit_str)
        
        if not nit_numeros:
            return nit_original, False
        
        # El último dígito siempre es el dígito verificador
        # Separar: todos los dígitos menos el último - último dígito
        if len(nit_numeros) >= 2:
            nit_base = nit_numeros[:-1]  # Todos menos el último
            dv = nit_numeros[-1]         # El último dígito
            
            # Formato correcto: BASE-DV
            nit_corregido = f"{nit_base}-{dv}"
            return nit_corregido, True
        else:
            # NIT muy corto, no se puede formatear
            return nit_original, False
    
    def procesar_correcciones(self):
        """Procesa todas las correcciones de NITs"""
        logger.info("\n=== PROCESANDO CORRECCIONES DE NITs ===")
        
        col_id = 'NÚMERO DE DOCUMENTO'
        col_tipo = 'TIPO DE DOCUMENTO'
        col_nombre = 'NOMBRES'
        
        # Crear columna para NITs corregidos
        self.df['NIT_CORREGIDO'] = self.df[col_id]
        self.df['FUE_CORREGIDO'] = False
        
        nits_corregidos = 0
        nits_sin_correccion = 0
        
        for idx, row in self.df.iterrows():
            if row[col_tipo] == 'NIT':
                nit_corregido, fue_modificado = self.corregir_formato_nit(
                    row[col_id], 
                    row[col_tipo]
                )
                
                if fue_modificado:
                    self.df.at[idx, 'NIT_CORREGIDO'] = nit_corregido
                    self.df.at[idx, 'FUE_CORREGIDO'] = True
                    nits_corregidos += 1
                    
                    self.correcciones.append({
                        'fila': idx + 2,
                        'nombre': row[col_nombre],
                        'nit_original': row[col_id],
                        'nit_corregido': nit_corregido
                    })
                else:
                    nits_sin_correccion += 1
        
        logger.info(f"✅ NITs corregidos: {nits_corregidos}")
        logger.info(f"⚠️  NITs sin corrección: {nits_sin_correccion}")
        
        if nits_corregidos > 0:
            logger.info("\n📝 Ejemplos de correcciones:")
            for corr in self.correcciones[:5]:
                logger.info(f"  {corr['nombre'][:40]:40} | {corr['nit_original']:15} → {corr['nit_corregido']}")
        
        # Actualizar la columna original con los valores corregidos
        self.df[col_id] = self.df['NIT_CORREGIDO']
        
        return nits_corregidos
    
    def generar_archivo_corregido(self, ruta_salida='./data/output'):
        """Genera el archivo Excel corregido manteniendo el formato original"""
        logger.info("\n=== GENERANDO ARCHIVO CORREGIDO ===")
        
        Path(ruta_salida).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_salida = f"{ruta_salida}/CLIENTES_SOFTSEGUROS_CORREGIDO_{timestamp}.xlsx"
        
        # Eliminar columnas auxiliares antes de guardar
        df_final = self.df.drop(columns=['NIT_CORREGIDO', 'FUE_CORREGIDO'], errors='ignore')
        
        # Guardar Excel con formato
        with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name='CLIENTES', index=False)
            
            # Obtener worksheet para aplicar formato
            worksheet = writer.sheets['CLIENTES']
            
            # Aplicar formato a encabezados
            self._aplicar_formato_header(worksheet)
            
            # Auto-ajustar ancho de columnas
            self._ajustar_ancho_columnas(worksheet)
        
        logger.info(f"✅ Archivo generado: {archivo_salida}")
        return archivo_salida
    
    def generar_reporte_correcciones(self, ruta_salida='./data/output'):
        """Genera un reporte detallado de las correcciones realizadas"""
        if not self.correcciones:
            logger.info("No hay correcciones para reportar")
            return None
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_reporte = f"{ruta_salida}/REPORTE_CORRECCIONES_NITS_{timestamp}.xlsx"
        
        df_correcciones = pd.DataFrame(self.correcciones)
        
        with pd.ExcelWriter(archivo_reporte, engine='openpyxl') as writer:
            # Hoja de correcciones
            df_correcciones.to_excel(writer, sheet_name='Correcciones', index=False)
            
            # Hoja de resumen
            resumen = [
                ['REPORTE DE CORRECCIONES DE NITs', ''],
                ['Fecha', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['', ''],
                ['Total de NITs corregidos', len(self.correcciones)],
                ['Total de registros procesados', len(self.df)],
                ['Archivo original', self.archivo_entrada]
            ]
            
            df_resumen = pd.DataFrame(resumen, columns=['Descripción', 'Valor'])
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Aplicar formato
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                self._aplicar_formato_header(worksheet)
                self._ajustar_ancho_columnas(worksheet)
        
        logger.info(f"✅ Reporte de correcciones: {archivo_reporte}")
        return archivo_reporte
    
    def _aplicar_formato_header(self, worksheet):
        """Aplica formato profesional a los encabezados"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _ajustar_ancho_columnas(self, worksheet):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Función principal"""
    archivo_entrada = 'CLIENTES SOFTSEGUROSv2.xlsx'
    
    logger.info("="*60)
    logger.info("CORRECCIÓN DE FORMATO DE NITs")
    logger.info("="*60)
    
    # Crear corrector
    corrector = CorreccionNITs(archivo_entrada)
    
    # Cargar datos
    corrector.cargar_datos()
    
    # Procesar correcciones
    num_correcciones = corrector.procesar_correcciones()
    
    if num_correcciones > 0:
        # Generar archivo corregido
        archivo_salida = corrector.generar_archivo_corregido()
        
        # Generar reporte de correcciones
        archivo_reporte = corrector.generar_reporte_correcciones()
        
        logger.info("\n" + "="*60)
        logger.info("✅ PROCESO COMPLETADO")
        logger.info("="*60)
        logger.info(f"📁 Archivo corregido: {archivo_salida}")
        logger.info(f"📊 Reporte de correcciones: {archivo_reporte}")
    else:
        logger.info("\n" + "="*60)
        logger.info("ℹ️  No se encontraron NITs para corregir")
        logger.info("="*60)


if __name__ == "__main__":
    main()
