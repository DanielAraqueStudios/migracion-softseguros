"""
Calculador de Dígitos de Verificación - Interfaz Profesional
============================================================
GUI moderna con PyQt6 para calcular y agregar DVs a NITs
- Dark Mode profesional
- Selección de archivos con explorador
- Usa API DIAN
- Solo personas JURÍDICAS
"""

import sys
import subprocess
import os
import time
import re
import requests
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QFileDialog, QGroupBox,
    QProgressBar, QFrame, QMessageBox, QSplitter
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor


# ==================== ESTILOS DARK MODE ====================
DARK_STYLESHEET = """
QMainWindow {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                stop:0 #0f172a, stop:1 #1e293b);
}

QWidget {
    background-color: transparent;
    color: #e2e8f0;
    font-family: 'Segoe UI', Arial;
    font-size: 10pt;
}

QGroupBox {
    background-color: rgba(30, 41, 59, 0.8);
    border: 2px solid #334155;
    border-radius: 12px;
    margin-top: 15px;
    padding: 20px;
    font-weight: 600;
    color: #e2e8f0;
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 8px 20px;
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 #10b981, stop:1 #059669);
    border-radius: 6px;
    color: white;
    font-weight: 700;
}

QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #10b981, stop:1 #059669);
    color: white;
    border: none;
    border-radius: 10px;
    padding: 15px 25px;
    font-weight: 700;
    font-size: 10pt;
    min-height: 25px;
}

QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #34d399, stop:1 #10b981);
}

QPushButton:pressed {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #059669, stop:1 #047857);
}

QPushButton:disabled {
    background: #334155;
    color: #64748b;
}

QPushButton#btnPrimary {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #3b82f6, stop:1 #2563eb);
    min-height: 30px;
    font-size: 11pt;
}

QPushButton#btnPrimary:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #60a5fa, stop:1 #3b82f6);
}

QPushButton#btnDanger {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #ef4444, stop:1 #dc2626);
}

QPushButton#btnDanger:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #f87171, stop:1 #ef4444);
}

QLabel {
    color: #e2e8f0;
    background: transparent;
}

QLabel#titleLabel {
    font-size: 18pt;
    font-weight: 800;
    color: #10b981;
    text-align: center;
}

QLabel#subtitleLabel {
    font-size: 11pt;
    color: #94a3b8;
}

QLabel#statLabel {
    font-size: 32pt;
    font-weight: 800;
    color: #10b981;
}

QLabel#statusLabel {
    font-size: 10pt;
    color: #64748b;
    font-style: italic;
    padding: 8px;
}

QLabel#statusOk {
    color: #10b981;
    font-weight: 600;
}

QTextEdit {
    background-color: #020617;
    border: 2px solid #1e293b;
    border-radius: 10px;
    color: #e2e8f0;
    padding: 12px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 9pt;
    selection-background-color: #10b981;
}

QTextEdit:focus {
    border: 2px solid #10b981;
}

QProgressBar {
    background-color: #1e293b;
    border: 2px solid #334155;
    border-radius: 10px;
    text-align: center;
    color: white;
    font-weight: 700;
    height: 30px;
}

QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 #10b981, stop:0.5 #34d399, stop:1 #6ee7b7);
    border-radius: 8px;
}

QScrollBar:vertical {
    background: #0f172a;
    width: 16px;
    border-radius: 8px;
}

QScrollBar::handle:vertical {
    background: #334155;
    min-height: 30px;
    border-radius: 8px;
}

QScrollBar::handle:vertical:hover {
    background: #475569;
}
"""


class APIThread(QThread):
    """Thread para iniciar la API DIAN"""
    progreso = pyqtSignal(str, str)  # mensaje, tipo
    terminado = pyqtSignal(bool)
    
    def __init__(self):
        super().__init__()
    
    def run(self):
        try:
            # Verificar si ya está corriendo
            try:
                response = requests.get("http://localhost:8000/health", timeout=2)
                if response.status_code == 200:
                    self.progreso.emit("✅ API DIAN ya está corriendo", "success")
                    self.terminado.emit(True)
                    return
            except:
                pass
            
            self.progreso.emit("🚀 Iniciando servidor API DIAN...", "info")
            
            # Ruta al backend
            script_dir = Path(__file__).parent.parent
            backend_dir = script_dir / 'backend'
            
            if not backend_dir.exists():
                self.progreso.emit(f"❌ No se encuentra el directorio backend", "error")
                self.terminado.emit(False)
                return
            
            # Iniciar uvicorn
            subprocess.Popen(
                [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8000"],
                cwd=str(backend_dir),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            # Esperar a que inicie
            for i in range(10):
                time.sleep(1)
                try:
                    response = requests.get("http://localhost:8000/health", timeout=2)
                    if response.status_code == 200:
                        self.progreso.emit("✅ API DIAN iniciada correctamente", "success")
                        self.terminado.emit(True)
                        return
                except:
                    self.progreso.emit(f"   Esperando... ({i+1}/10)", "info")
            
            self.progreso.emit("❌ No se pudo iniciar la API", "error")
            self.terminado.emit(False)
            
        except Exception as e:
            self.progreso.emit(f"❌ Error: {str(e)}", "error")
            self.terminado.emit(False)


class CalculadorThread(QThread):
    """Thread para calcular DVs sin bloquear UI"""
    progreso = pyqtSignal(str, str)  # mensaje, tipo
    terminado = pyqtSignal(bool, dict)
    
    def __init__(self, archivo):
        super().__init__()
        self.archivo = archivo
    
    def run(self):
        try:
            self.progreso.emit("📂 Cargando archivo...", "info")
            
            from openpyxl import load_workbook
            wb = load_workbook(self.archivo)
            ws = wb.active
            
            total_filas = ws.max_row - 1
            self.progreso.emit(f"📊 Total de registros: {total_filas}", "info")
            
            # Columnas
            COL_POLIZA = 1  # A
            COL_DOCUMENTO = 28  # AB
            COL_TIPO_PERSONA = 29  # AC
            
            cambios = []
            errores = []
            naturales = 0
            ya_tienen_dv = 0
            
            for fila in range(2, ws.max_row + 1):
                poliza = str(ws.cell(row=fila, column=COL_POLIZA).value or "").strip()
                documento = str(ws.cell(row=fila, column=COL_DOCUMENTO).value or "").strip()
                tipo = str(ws.cell(row=fila, column=COL_TIPO_PERSONA).value or "").strip().upper()
                
                # Mostrar progreso cada 100 filas
                if (fila - 1) % 100 == 0:
                    self.progreso.emit(f"📝 Procesando fila {fila} de {ws.max_row}...", "info")
                
                # Solo JURÍDICAS
                if tipo != 'J':
                    if tipo == 'N':
                        naturales += 1
                    continue
                
                # Ya tiene DV
                if '-' in documento:
                    ya_tienen_dv += 1
                    self.progreso.emit(f"  ⏭️  Fila {fila} - Póliza {poliza}: {documento} (ya tiene DV)", "info")
                    continue
                
                # Validar número
                doc_limpio = re.sub(r'\D', '', documento)
                if not doc_limpio or len(doc_limpio) < 5:
                    errores.append({'poliza': poliza, 'fila': fila, 'documento': documento})
                    self.progreso.emit(f"  ⚠️  Fila {fila} - Póliza {poliza}: {documento} (NIT inválido)", "warning")
                    continue
                
                # Calcular DV
                dv = self._calcular_dv(doc_limpio)
                
                if dv is not None:
                    nit_completo = f"{doc_limpio}-{dv}"
                    ws.cell(row=fila, column=COL_DOCUMENTO).value = nit_completo
                    cambios.append({
                        'poliza': poliza,
                        'fila': fila,
                        'anterior': documento,
                        'nuevo': nit_completo
                    })
                    
                    # Mostrar CADA cambio
                    self.progreso.emit(f"  ✓ Fila {fila} - Póliza {poliza}: {documento} → {nit_completo}", "success")
                else:
                    errores.append({'poliza': poliza, 'fila': fila, 'documento': documento})
                    self.progreso.emit(f"  ❌ Fila {fila} - Póliza {poliza}: Error calculando DV para {documento}", "error")
            
            # Guardar si hay cambios
            if len(cambios) > 0:
                self.progreso.emit("💾 Guardando cambios...", "info")
                wb.save(self.archivo)
            
            wb.close()
            
            estadisticas = {
                'total': total_filas,
                'naturales': naturales,
                'ya_tienen_dv': ya_tienen_dv,
                'cambios': len(cambios),
                'errores': len(errores)
            }
            
            self.terminado.emit(True, estadisticas)
            
        except Exception as e:
            self.progreso.emit(f"❌ Error: {str(e)}", "error")
            self.terminado.emit(False, {})
    
    def _calcular_dv(self, nit, reintentos=3):
        """Calcula DV usando API DIAN"""
        for intento in range(reintentos):
            try:
                response = requests.post(
                    "http://localhost:8000/calcular",
                    json={"nit": nit},
                    timeout=10
                )
                if response.status_code == 200:
                    return str(response.json()["digito_verificacion"])
                else:
                    if intento < reintentos - 1:
                        time.sleep(0.5)
            except:
                if intento < reintentos - 1:
                    time.sleep(0.5)
        return None


class CalculadorDVGUI(QMainWindow):
    """Ventana principal del calculador de DVs"""
    
    def __init__(self):
        super().__init__()
        self.archivo_seleccionado = None
        self.api_iniciada = False
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Calculador de Dígitos de Verificación - API DIAN")
        self.setMinimumSize(1000, 800)
        
        # ========== WIDGET CENTRAL CON SPLITTER ==========
        central = QWidget()
        self.setCentralWidget(central)
        
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Splitter horizontal (izquierda-derecha)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(4)
        splitter.setStyleSheet("""
            QSplitter::handle {
                background: #10b981;
            }
            QSplitter::handle:hover {
                background: #059669;
            }
        """)
        
        # ========== PANEL IZQUIERDO (Controles y Estadísticas) ==========
        left_panel = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)
        
        # Header
        header = QWidget()
        header.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                        stop:0 #10b981, stop:0.5 #059669, stop:1 #047857);
            border-radius: 15px;
            padding: 25px;
        """)
        header_layout = QVBoxLayout()
        
        title = QLabel("🔢 CALCULADOR DE DVs")
        title.setObjectName("titleLabel")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: white; background: transparent;")
        
        subtitle = QLabel("API DIAN - Personas Jurídicas")
        subtitle.setObjectName("subtitleLabel")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle.setStyleSheet("color: rgba(255, 255, 255, 0.9); background: transparent;")
        
        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        header.setLayout(header_layout)
        
        layout.addWidget(header)
        
        # Grupo API
        grupo_api = QGroupBox("🌐 Estado de la API DIAN")
        layout_api = QVBoxLayout()
        
        self.lbl_api_status = QLabel("⏳ API no iniciada")
        self.lbl_api_status.setObjectName("statusLabel")
        self.lbl_api_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.btn_iniciar_api = QPushButton("🚀 INICIAR API DIAN")
        self.btn_iniciar_api.setObjectName("btnPrimary")
        self.btn_iniciar_api.clicked.connect(self.iniciar_api)
        self.btn_iniciar_api.setMinimumHeight(55)
        
        layout_api.addWidget(self.lbl_api_status)
        layout_api.addWidget(self.btn_iniciar_api)
        
        grupo_api.setLayout(layout_api)
        layout.addWidget(grupo_api)
        
        # Grupo Archivo
        grupo_archivo = QGroupBox("📁 Archivo MAVISO")
        layout_archivo = QVBoxLayout()
        
        self.lbl_archivo = QLabel("No se ha seleccionado ningún archivo")
        self.lbl_archivo.setObjectName("statusLabel")
        self.lbl_archivo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.btn_seleccionar = QPushButton("📂 SELECCIONAR ARCHIVO")
        self.btn_seleccionar.clicked.connect(self.seleccionar_archivo)
        self.btn_seleccionar.setEnabled(False)
        
        layout_archivo.addWidget(self.lbl_archivo)
        layout_archivo.addWidget(self.btn_seleccionar)
        
        grupo_archivo.setLayout(layout_archivo)
        layout.addWidget(grupo_archivo)
        
        # Botón procesar
        self.btn_procesar = QPushButton("⚡ CALCULAR Y AGREGAR DVs")
        self.btn_procesar.setObjectName("btnPrimary")
        self.btn_procesar.clicked.connect(self.procesar_archivo)
        self.btn_procesar.setEnabled(False)
        self.btn_procesar.setMinimumHeight(65)
        layout.addWidget(self.btn_procesar)
        
        # Progreso
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)
        
        # Estadísticas
        grupo_stats = QGroupBox("📊 Estadísticas")
        layout_stats = QHBoxLayout()
        layout_stats.setSpacing(10)
        
        stats_info = [
            ('total', 'Total', '#3b82f6'),
            ('cambios', 'Actualiz.', '#10b981'),
            ('naturales', 'Naturales', '#8b5cf6'),
            ('errores', 'Errores', '#ef4444')
        ]
        
        self.stats_widgets = {}
        for key, label, color in stats_info:
            widget = QWidget()
            widget_layout = QVBoxLayout()
            widget_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            widget_layout.setSpacing(5)
            
            lbl_value = QLabel("0")
            lbl_value.setObjectName("statLabel")
            lbl_value.setStyleSheet(f"color: {color}; background: transparent; font-size: 26pt;")
            lbl_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            lbl_name = QLabel(label)
            lbl_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_name.setStyleSheet("color: #64748b; font-size: 9pt; background: transparent;")
            
            widget_layout.addWidget(lbl_value)
            widget_layout.addWidget(lbl_name)
            widget.setLayout(widget_layout)
            
            layout_stats.addWidget(widget)
            self.stats_widgets[key] = lbl_value
        
        grupo_stats.setLayout(layout_stats)
        layout.addWidget(grupo_stats)
        
        # Botón abrir archivo
        self.btn_abrir = QPushButton("📂 ABRIR ARCHIVO")
        self.btn_abrir.setEnabled(False)
        self.btn_abrir.clicked.connect(self.abrir_archivo)
        layout.addWidget(self.btn_abrir)
        
        # Spacer para empujar contenido hacia arriba
        layout.addStretch()
        
        left_panel.setLayout(layout)
        
        # ========== PANEL DERECHO (Logs) ==========
        right_panel = QWidget()
        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(25, 25, 25, 25)
        right_layout.setSpacing(15)
        
        # Título del panel de logs
        log_header = QLabel("📋 REGISTRO DE ACTIVIDAD")
        log_header.setStyleSheet("""
            color: #10b981;
            font-size: 16pt;
            font-weight: bold;
            padding: 10px;
        """)
        log_header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        right_layout.addWidget(log_header)
        
        # Logs
        self.txt_logs = QTextEdit()
        self.txt_logs.setReadOnly(True)
        self.txt_logs.setMinimumWidth(350)
        self.txt_logs.setStyleSheet("""
            QTextEdit {
                background-color: #0f172a;
                color: #e2e8f0;
                border: 2px solid #10b981;
                border-radius: 12px;
                padding: 15px;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 11px;
            }
        """)
        right_layout.addWidget(self.txt_logs, 1)
        
        # Botón limpiar logs
        btn_limpiar = QPushButton("🗑️ LIMPIAR LOGS")
        btn_limpiar.clicked.connect(lambda: self.txt_logs.clear())
        btn_limpiar.setMinimumHeight(40)
        right_layout.addWidget(btn_limpiar)
        
        right_panel.setLayout(right_layout)
        
        # Agregar paneles al splitter
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        
        # Configurar tamaños iniciales (45% izquierda, 55% derecha)
        splitter.setSizes([450, 550])
        
        main_layout.addWidget(splitter)
        central.setLayout(main_layout)
        
        # Centrar ventana
        self.center()
        
        # Log inicial
        self.log("🎯 Sistema iniciado", "info")
        self.log("💡 Paso 1: Iniciar API DIAN", "info")
    
    def center(self):
        """Centra la ventana en la pantalla"""
        screen = QApplication.primaryScreen().geometry()
        size = self.geometry()
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )
    
    def log(self, mensaje, tipo="info"):
        """Agrega mensaje a los logs con color"""
        colores = {
            'info': '#3b82f6',
            'success': '#10b981',
            'warning': '#f59e0b',
            'error': '#ef4444'
        }
        color = colores.get(tipo, '#e2e8f0')
        timestamp = datetime.now().strftime('%H:%M:%S')
        html = f'<span style="color: #64748b;">[{timestamp}]</span> <span style="color: {color}; font-weight: 600;">{mensaje}</span>'
        self.txt_logs.append(html)
    
    def iniciar_api(self):
        """Inicia la API DIAN en background"""
        self.btn_iniciar_api.setEnabled(False)
        self.lbl_api_status.setText("⏳ Iniciando API...")
        self.log("", "info")
        self.log("═" * 50, "info")
        self.log("🚀 INICIANDO API DIAN", "info")
        self.log("═" * 50, "info")
        
        self.api_thread = APIThread()
        self.api_thread.progreso.connect(self.log)
        self.api_thread.terminado.connect(self.api_iniciada_callback)
        self.api_thread.start()
    
    def api_iniciada_callback(self, exito):
        """Callback cuando la API termina de iniciar"""
        if exito:
            self.api_iniciada = True
            self.lbl_api_status.setText("✅ API DIAN operativa")
            self.lbl_api_status.setObjectName("statusOk")
            self.lbl_api_status.setStyleSheet("color: #10b981; font-weight: 700; font-size: 12pt;")
            self.btn_iniciar_api.setText("✅ API INICIADA")
            self.btn_seleccionar.setEnabled(True)
            self.log("", "success")
            self.log("🎉 API lista para calcular dígitos de verificación", "success")
        else:
            self.btn_iniciar_api.setEnabled(True)
            self.lbl_api_status.setText("❌ Error al iniciar API")
            self.log("", "error")
            self.log("💡 Intente iniciar manualmente: cd backend; python -m uvicorn app:app --port 8000", "warning")
    
    def seleccionar_archivo(self):
        """Abre diálogo para seleccionar archivo"""
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo MAVISO",
            "",
            "Excel Files (*.xlsx *.xlsm)"
        )
        
        if archivo:
            self.archivo_seleccionado = archivo
            nombre = Path(archivo).name
            self.lbl_archivo.setText(f"✅ {nombre}")
            self.lbl_archivo.setObjectName("statusOk")
            self.lbl_archivo.setStyleSheet("color: #10b981; font-weight: 600; font-size: 11pt;")
            self.btn_procesar.setEnabled(True)
            self.btn_abrir.setEnabled(True)
            self.log(f"✅ Archivo seleccionado: {nombre}", "success")
    
    def procesar_archivo(self):
        """Procesa el archivo calculando DVs"""
        respuesta = QMessageBox.question(
            self,
            "Confirmar Operación",
            "¿Desea calcular y agregar el DV a los NITs de personas jurídicas?\n\n"
            "• Se procesarán solo personas JURÍDICAS (columna AC = 'J')\n"
            "• Se ignorarán personas NATURALES (columna AC = 'N')\n"
            "• Se usará la API DIAN para cálculos precisos\n"
            "⚠️ El archivo original será modificado manteniendo su formato.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        self.btn_procesar.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        
        self.log("", "info")
        self.log("═" * 50, "info")
        self.log("⚡ INICIANDO PROCESAMIENTO", "info")
        self.log("═" * 50, "info")
        
        self.calc_thread = CalculadorThread(self.archivo_seleccionado)
        self.calc_thread.progreso.connect(self.log)
        self.calc_thread.terminado.connect(self.procesamiento_terminado)
        self.calc_thread.start()
    
    def procesamiento_terminado(self, exito, estadisticas):
        """Callback cuando termina el procesamiento"""
        self.btn_procesar.setEnabled(True)
        self.progress.setVisible(False)
        
        if exito:
            # Actualizar estadísticas
            for key, value in estadisticas.items():
                if key in self.stats_widgets:
                    self.stats_widgets[key].setText(str(value))
            
            self.log("", "success")
            self.log("═" * 50, "success")
            self.log("✅ PROCESAMIENTO COMPLETADO", "success")
            self.log("═" * 50, "success")
            self.log(f"📊 Total de registros: {estadisticas.get('total', 0)}", "info")
            self.log(f"✅ NITs actualizados: {estadisticas.get('cambios', 0)}", "success")
            self.log(f"👤 Personas naturales ignoradas: {estadisticas.get('naturales', 0)}", "info")
            self.log(f"📝 Ya tenían DV: {estadisticas.get('ya_tienen_dv', 0)}", "info")
            
            if estadisticas.get('errores', 0) > 0:
                self.log(f"⚠️ Errores al calcular: {estadisticas.get('errores', 0)}", "warning")
            
            mensaje = f"✅ Proceso completado exitosamente\n\n"
            mensaje += f"NITs actualizados: {estadisticas.get('cambios', 0)}\n"
            mensaje += f"Personas naturales ignoradas: {estadisticas.get('naturales', 0)}\n"
            
            if estadisticas.get('errores', 0) > 0:
                mensaje += f"\n⚠️ {estadisticas.get('errores', 0)} NITs no pudieron ser calculados"
            
            QMessageBox.information(self, "Proceso Completado", mensaje)
        else:
            self.log("", "error")
            self.log("❌ El procesamiento falló", "error")
            QMessageBox.critical(self, "Error", "Ocurrió un error durante el procesamiento.")
    
    def abrir_archivo(self):
        """Abre el archivo procesado"""
        if self.archivo_seleccionado and os.name == 'nt':
            os.startfile(self.archivo_seleccionado)


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setStyleSheet(DARK_STYLESHEET)
    
    ventana = CalculadorDVGUI()
    ventana.show()
    
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
