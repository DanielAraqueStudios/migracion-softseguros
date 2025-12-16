"""
Comparador de Archivos - Interfaz Gráfica con Pestañas
======================================================
GUI para comparar archivos Excel:
- Pestaña 1: Maviso Manual vs Maviso Generado
- Pestaña 2: Maviso vs CELER (validar datos originales)
- Pestaña 3: Validar Mensuales sin Prima
"""

import sys
import subprocess
import os
import time
import re
import requests
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QFileDialog, QGroupBox,
    QProgressBar, QFrame, QMessageBox, QTabWidget, QSpinBox,
    QComboBox, QCheckBox, QScrollArea
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QPalette
import pandas as pd

from comparar_archivos import ComparadorMaviso
from comparar_maviso_celer import ComparadorMavisoCeler


# Estilos CSS para tema oscuro
DARK_STYLE = """
QMainWindow {
    background-color: #1e1e1e;
}
QWidget {
    background-color: #1e1e1e;
    color: #d4d4d4;
    font-family: 'Segoe UI', Arial, sans-serif;
}
QTabWidget::pane {
    border: 1px solid #3c3c3c;
    border-radius: 4px;
    background-color: #1e1e1e;
}
QTabBar::tab {
    background-color: #2d2d2d;
    color: #d4d4d4;
    padding: 10px 20px;
    margin-right: 2px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
}
QTabBar::tab:selected {
    background-color: #0e639c;
    color: white;
}
QTabBar::tab:hover:!selected {
    background-color: #3c3c3c;
}
QGroupBox {
    border: 1px solid #3c3c3c;
    border-radius: 8px;
    margin-top: 12px;
    padding-top: 10px;
    font-weight: bold;
    color: #569cd6;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 5px;
}
QPushButton {
    background-color: #0e639c;
    color: white;
    border: none;
    border-radius: 4px;
    padding: 8px 16px;
    font-weight: bold;
    min-width: 120px;
}
QPushButton:hover {
    background-color: #1177bb;
}
QPushButton:pressed {
    background-color: #0d5a8c;
}
QPushButton:disabled {
    background-color: #3c3c3c;
    color: #6c6c6c;
}
QPushButton#btnComparar {
    background-color: #16825d;
    font-size: 14px;
    padding: 12px 24px;
}
QPushButton#btnComparar:hover {
    background-color: #1a9e6f;
}
QPushButton#btnExportar {
    background-color: #c27c0e;
}
QPushButton#btnExportar:hover {
    background-color: #d98f1a;
}
QPushButton#btnAbrir {
    background-color: #6a9955;
}
QPushButton#btnAbrir:hover {
    background-color: #7cb668;
}
QTextEdit {
    background-color: #252526;
    border: 1px solid #3c3c3c;
    border-radius: 4px;
    padding: 8px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 11px;
    color: #d4d4d4;
}
QLabel {
    color: #d4d4d4;
}
QLabel#titulo {
    font-size: 18px;
    font-weight: bold;
    color: #569cd6;
    padding: 5px;
}
QLabel#subtitulo {
    font-size: 11px;
    color: #808080;
    padding-bottom: 5px;
}
QLabel#archivo {
    background-color: #252526;
    border: 1px solid #3c3c3c;
    border-radius: 4px;
    padding: 8px;
    color: #9cdcfe;
}
QLabel#estadistica {
    font-size: 24px;
    font-weight: bold;
    color: #4ec9b0;
}
QLabel#estadistica_label {
    font-size: 11px;
    color: #808080;
}
QProgressBar {
    border: 1px solid #3c3c3c;
    border-radius: 4px;
    text-align: center;
    background-color: #252526;
}
QProgressBar::chunk {
    background-color: #16825d;
    border-radius: 3px;
}
QFrame#separator {
    background-color: #3c3c3c;
    max-height: 1px;
}
"""


class ComparadorThread(QThread):
    """Thread para ejecutar la comparación sin bloquear la GUI"""
    
    log_signal = pyqtSignal(str, str)  # mensaje, tipo
    finished_signal = pyqtSignal(bool, str, dict)  # éxito, mensaje, estadísticas
    
    def __init__(self, comparador, tipo="maviso"):
        super().__init__()
        self.comparador = comparador
        self.tipo = tipo
    
    def run(self):
        try:
            self.log_signal.emit("=" * 50, "info")
            self.log_signal.emit("INICIANDO COMPARACIÓN", "info")
            self.log_signal.emit("=" * 50, "info")
            
            # Ejecutar comparación
            exito, mensaje, resultados = self.comparador.comparar()
            
            if exito:
                # Emitir logs de discrepancias
                for disc in self.comparador.obtener_discrepancias():
                    self.log_signal.emit(f"❌ DISCREPANCIA - Póliza: {disc['poliza']}", "error")
                    
                    if self.tipo == "maviso":
                        self.log_signal.emit(f"   Fila manual: {disc['fila_manual']} | Fila generado: {disc['fila_generado']}", "warning")
                        for err in disc['errores']:
                            self.log_signal.emit(f"   → {err['campo']}: Manual='{err['manual']}' vs Generado='{err['generado']}'", "warning")
                    else:  # celer
                        self.log_signal.emit(f"   Fila Maviso: {disc['fila_maviso']} | Fila CELER: {disc['fila_celer']}", "warning")
                        for err in disc['errores']:
                            self.log_signal.emit(f"   → {err['campo']}: Maviso='{err['maviso']}' vs CELER='{err['celer']}'", "warning")
                
                # Logs de solo en uno
                if self.tipo == "maviso":
                    solo_1 = resultados.get('solo_manual', [])
                    solo_2 = resultados.get('solo_generado', [])
                    label_1, label_2 = "ARCHIVO MANUAL", "ARCHIVO GENERADO"
                else:
                    solo_1 = resultados.get('solo_maviso', [])
                    solo_2 = resultados.get('solo_celer', [])
                    label_1, label_2 = "MAVISO", "CELER"
                
                if solo_1:
                    self.log_signal.emit("", "info")
                    self.log_signal.emit(f"⚠️ PÓLIZAS SOLO EN {label_1}:", "warning")
                    for item in solo_1[:20]:
                        self.log_signal.emit(f"   • {item['poliza']} (Fila {item['fila']})", "warning")
                    if len(solo_1) > 20:
                        self.log_signal.emit(f"   ... y {len(solo_1) - 20} más", "warning")
                
                if solo_2:
                    self.log_signal.emit("", "info")
                    self.log_signal.emit(f"⚠️ PÓLIZAS SOLO EN {label_2}:", "warning")
                    for item in solo_2[:20]:
                        self.log_signal.emit(f"   • {item['poliza']} (Fila {item['fila']})", "warning")
                    if len(solo_2) > 20:
                        self.log_signal.emit(f"   ... y {len(solo_2) - 20} más", "warning")
                
                self.finished_signal.emit(True, mensaje, self.comparador.obtener_estadisticas())
            else:
                self.finished_signal.emit(False, mensaje, {})
                
        except Exception as e:
            self.log_signal.emit(f"❌ Error: {str(e)}", "error")
            self.finished_signal.emit(False, str(e), {})


class TabMavisoVsMaviso(QWidget):
    """Pestaña: Comparar Maviso Manual vs Maviso Generado"""
    
    def __init__(self):
        super().__init__()
        self.comparador = ComparadorMaviso()
        self.ruta_manual = None
        self.ruta_generado = None
        self.ruta_ultimo_reporte = None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Título
        titulo = QLabel("📊 Maviso Manual vs Maviso Generado")
        titulo.setObjectName("titulo")
        layout.addWidget(titulo)
        
        subtitulo = QLabel("Compara el archivo corregido manualmente contra el generado por el script")
        subtitulo.setObjectName("subtitulo")
        layout.addWidget(subtitulo)
        
        # Grupo: Archivos
        grupo_archivos = QGroupBox("📁 Archivos a Comparar")
        layout_archivos = QVBoxLayout(grupo_archivos)
        
        # Archivo Manual
        layout_manual = QHBoxLayout()
        lbl_manual = QLabel("Archivo Manual:")
        lbl_manual.setMinimumWidth(150)
        self.lbl_ruta_manual = QLabel("No seleccionado")
        self.lbl_ruta_manual.setObjectName("archivo")
        btn_manual = QPushButton("📂 Seleccionar")
        btn_manual.clicked.connect(self.seleccionar_archivo_manual)
        layout_manual.addWidget(lbl_manual)
        layout_manual.addWidget(self.lbl_ruta_manual, 1)
        layout_manual.addWidget(btn_manual)
        layout_archivos.addLayout(layout_manual)
        
        # Archivo Generado
        layout_generado = QHBoxLayout()
        lbl_generado = QLabel("Archivo Generado:")
        lbl_generado.setMinimumWidth(150)
        self.lbl_ruta_generado = QLabel("No seleccionado")
        self.lbl_ruta_generado.setObjectName("archivo")
        btn_generado = QPushButton("📂 Seleccionar")
        btn_generado.clicked.connect(self.seleccionar_archivo_generado)
        layout_generado.addWidget(lbl_generado)
        layout_generado.addWidget(self.lbl_ruta_generado, 1)
        layout_generado.addWidget(btn_generado)
        layout_archivos.addLayout(layout_generado)
        
        layout.addWidget(grupo_archivos)
        
        # Botón Comparar
        layout_btn = QHBoxLayout()
        layout_btn.addStretch()
        self.btn_comparar = QPushButton("🔍 COMPARAR")
        self.btn_comparar.setObjectName("btnComparar")
        self.btn_comparar.clicked.connect(self.ejecutar_comparacion)
        self.btn_comparar.setEnabled(False)
        layout_btn.addWidget(self.btn_comparar)
        layout_btn.addStretch()
        layout.addLayout(layout_btn)
        
        # Grupo: Estadísticas
        grupo_stats = QGroupBox("📊 Resultados")
        layout_stats = QHBoxLayout(grupo_stats)
        
        # Crear widgets de estadísticas
        self.stats_widgets = {}
        stats_config = [
            ('total_manual', 'Total Manual', '#4ec9b0'),
            ('total_generado', 'Total Generado', '#4ec9b0'),
            ('coincidencias', 'Coincidencias', '#6a9955'),
            ('discrepancias', 'Discrepancias', '#f44747'),
            ('solo_manual', 'Solo Manual', '#dcdcaa'),
            ('solo_generado', 'Solo Generado', '#dcdcaa'),
        ]
        
        for key, label, color in stats_config:
            frame = QFrame()
            frame_layout = QVBoxLayout(frame)
            frame_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            lbl_valor = QLabel("--")
            lbl_valor.setObjectName("estadistica")
            lbl_valor.setStyleSheet(f"color: {color};")
            lbl_valor.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            lbl_nombre = QLabel(label)
            lbl_nombre.setObjectName("estadistica_label")
            lbl_nombre.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            frame_layout.addWidget(lbl_valor)
            frame_layout.addWidget(lbl_nombre)
            layout_stats.addWidget(frame)
            
            self.stats_widgets[key] = lbl_valor
        
        layout.addWidget(grupo_stats)
        
        # Grupo: Logs
        grupo_logs = QGroupBox("📋 Log de Comparación")
        layout_logs = QVBoxLayout(grupo_logs)
        
        self.txt_logs = QTextEdit()
        self.txt_logs.setReadOnly(True)
        self.txt_logs.setMinimumHeight(150)
        layout_logs.addWidget(self.txt_logs)
        
        layout_btns_log = QHBoxLayout()
        btn_limpiar = QPushButton("🗑️ Limpiar")
        btn_limpiar.clicked.connect(self.limpiar_logs)
        self.btn_exportar = QPushButton("📊 Exportar Reporte")
        self.btn_exportar.setObjectName("btnExportar")
        self.btn_exportar.clicked.connect(self.exportar_reporte)
        self.btn_exportar.setEnabled(False)
        layout_btns_log.addWidget(btn_limpiar)
        layout_btns_log.addStretch()
        layout_btns_log.addWidget(self.btn_exportar)
        layout_logs.addLayout(layout_btns_log)
        
        layout.addWidget(grupo_logs, 1)
    
    def seleccionar_archivo_manual(self):
        """Abre diálogo para seleccionar archivo manual"""
        # Buscar carpeta padre (produccion_a_un_año)
        carpeta_base = Path(__file__).parent.parent
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo manual (corregido)",
            str(carpeta_base),
            "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            self.ruta_manual = archivo
            self.lbl_ruta_manual.setText(Path(archivo).name)
            self.log(f"✅ Archivo manual: {Path(archivo).name}", "success")
            self.verificar_archivos()
    
    def seleccionar_archivo_generado(self):
        """Abre diálogo para seleccionar archivo generado"""
        # Buscar en carpeta output del padre
        carpeta_base = Path(__file__).parent.parent / 'output'
        if not carpeta_base.exists():
            carpeta_base = Path(__file__).parent.parent
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo generado (script)",
            str(carpeta_base),
            "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            self.ruta_generado = archivo
            self.lbl_ruta_generado.setText(Path(archivo).name)
            self.log(f"✅ Archivo generado: {Path(archivo).name}", "success")
            self.verificar_archivos()
    
    def verificar_archivos(self):
        """Verifica si ambos archivos están seleccionados"""
        self.btn_comparar.setEnabled(
            self.ruta_manual is not None and self.ruta_generado is not None
        )
    
    def ejecutar_comparacion(self):
        """Inicia la comparación de archivos"""
        self.btn_comparar.setEnabled(False)
        self.btn_exportar.setEnabled(False)
        
        # Resetear estadísticas
        for widget in self.stats_widgets.values():
            widget.setText("--")
        
        # Cargar archivos
        self.log("", "info")
        self.log("📂 Cargando archivos...", "info")
        
        exito, mensaje = self.comparador.cargar_archivo_manual(self.ruta_manual)
        if not exito:
            self.log(f"❌ {mensaje}", "error")
            self.btn_comparar.setEnabled(True)
            return
        self.log(f"✅ {mensaje}", "success")
        
        exito, mensaje = self.comparador.cargar_archivo_generado(self.ruta_generado)
        if not exito:
            self.log(f"❌ {mensaje}", "error")
            self.btn_comparar.setEnabled(True)
            return
        self.log(f"✅ {mensaje}", "success")
        
        # Ejecutar comparación en thread
        self.worker_thread = ComparadorThread(self.comparador, "maviso")
        self.worker_thread.log_signal.connect(self.log)
        self.worker_thread.finished_signal.connect(self.comparacion_terminada)
        self.worker_thread.start()
    
    def comparacion_terminada(self, exito: bool, mensaje: str, estadisticas: dict):
        """Callback cuando termina la comparación"""
        self.btn_comparar.setEnabled(True)
        
        if exito:
            # Actualizar estadísticas
            for key, widget in self.stats_widgets.items():
                valor = estadisticas.get(key, 0)
                widget.setText(str(valor))
            
            self.log("", "info")
            self.log("=" * 50, "success")
            self.log("✅ COMPARACIÓN COMPLETADA", "success")
            self.log("=" * 50, "success")
            
            self.btn_exportar.setEnabled(True)
        else:
            self.log(f"❌ Error: {mensaje}", "error")
    
    def exportar_reporte(self):
        """Exporta el reporte de comparación y lo abre automáticamente"""
        try:
            ruta = self.comparador.exportar_reporte()
            self.ruta_ultimo_reporte = ruta
            self.log(f"📊 Reporte exportado: {ruta}", "success")
            
            # Abrir automáticamente el archivo
            os.startfile(ruta)
            self.log(f"📂 Abriendo reporte automáticamente...", "info")
        except Exception as e:
            self.log(f"❌ Error exportando: {str(e)}", "error")
    
    def log(self, mensaje: str, tipo: str = "info"):
        """Agrega un mensaje al log con formato de color"""
        colores = {
            "info": "#d4d4d4",
            "success": "#6a9955",
            "warning": "#dcdcaa",
            "error": "#f44747"
        }
        color = colores.get(tipo, "#d4d4d4")
        self.txt_logs.append(f'<span style="color: {color};">{mensaje}</span>')
        
        # Auto-scroll al final
        scrollbar = self.txt_logs.verticalScrollBar()
        if scrollbar:
            scrollbar.setValue(scrollbar.maximum())
    
    def limpiar_logs(self):
        """Limpia el área de logs"""
        self.txt_logs.clear()


class TabMavisoVsCeler(QWidget):
    """Pestaña: Comparar Maviso vs CELER (datos originales)"""
    
    def __init__(self):
        super().__init__()
        self.comparador = ComparadorMavisoCeler()
        self.ruta_maviso = None
        self.ruta_celer = None
        self.ruta_ultimo_reporte = None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Título
        titulo = QLabel("🔄 Maviso vs CELER (Validar Datos Originales)")
        titulo.setObjectName("titulo")
        layout.addWidget(titulo)
        
        subtitulo = QLabel("Valida que Maviso tenga los mismos valores de Prima, Fechas y Modalidad que CELER")
        subtitulo.setObjectName("subtitulo")
        layout.addWidget(subtitulo)
        
        # Grupo: Archivos
        grupo_archivos = QGroupBox("📁 Archivos a Comparar")
        layout_archivos = QVBoxLayout(grupo_archivos)
        
        # Archivo Maviso
        layout_maviso = QHBoxLayout()
        lbl_maviso = QLabel("Archivo Maviso:")
        lbl_maviso.setMinimumWidth(150)
        self.lbl_ruta_maviso = QLabel("No seleccionado")
        self.lbl_ruta_maviso.setObjectName("archivo")
        btn_maviso = QPushButton("📂 Seleccionar")
        btn_maviso.clicked.connect(self.seleccionar_archivo_maviso)
        layout_maviso.addWidget(lbl_maviso)
        layout_maviso.addWidget(self.lbl_ruta_maviso, 1)
        layout_maviso.addWidget(btn_maviso)
        layout_archivos.addLayout(layout_maviso)
        
        # Archivo CELER
        layout_celer = QHBoxLayout()
        lbl_celer = QLabel("Archivo CELER:")
        lbl_celer.setMinimumWidth(150)
        self.lbl_ruta_celer = QLabel("No seleccionado")
        self.lbl_ruta_celer.setObjectName("archivo")
        btn_celer = QPushButton("📂 Seleccionar")
        btn_celer.clicked.connect(self.seleccionar_archivo_celer)
        layout_celer.addWidget(lbl_celer)
        layout_celer.addWidget(self.lbl_ruta_celer, 1)
        layout_celer.addWidget(btn_celer)
        layout_archivos.addLayout(layout_celer)
        
        layout.addWidget(grupo_archivos)
        
        # Botón Comparar
        layout_btn = QHBoxLayout()
        layout_btn.addStretch()
        self.btn_comparar = QPushButton("🔍 COMPARAR")
        self.btn_comparar.setObjectName("btnComparar")
        self.btn_comparar.clicked.connect(self.ejecutar_comparacion)
        self.btn_comparar.setEnabled(False)
        layout_btn.addWidget(self.btn_comparar)
        layout_btn.addStretch()
        layout.addLayout(layout_btn)
        
        # Grupo: Estadísticas
        grupo_stats = QGroupBox("📊 Resultados")
        layout_stats = QHBoxLayout(grupo_stats)
        
        self.stats_widgets = {}
        stats_config = [
            ('total_maviso', 'Total Maviso', '#4ec9b0'),
            ('total_celer', 'Total CELER', '#4ec9b0'),
            ('coincidencias', 'Coincidencias', '#6a9955'),
            ('discrepancias', 'Discrepancias', '#f44747'),
            ('solo_maviso', 'Solo Maviso', '#dcdcaa'),
            ('solo_celer', 'Solo CELER', '#dcdcaa'),
        ]
        
        for key, label, color in stats_config:
            frame = QFrame()
            frame_layout = QVBoxLayout(frame)
            frame_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            lbl_valor = QLabel("--")
            lbl_valor.setObjectName("estadistica")
            lbl_valor.setStyleSheet(f"color: {color};")
            lbl_valor.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            lbl_nombre = QLabel(label)
            lbl_nombre.setObjectName("estadistica_label")
            lbl_nombre.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            frame_layout.addWidget(lbl_valor)
            frame_layout.addWidget(lbl_nombre)
            layout_stats.addWidget(frame)
            
            self.stats_widgets[key] = lbl_valor
        
        layout.addWidget(grupo_stats)
        
        # Grupo: Logs
        grupo_logs = QGroupBox("📋 Log de Comparación")
        layout_logs = QVBoxLayout(grupo_logs)
        
        self.txt_logs = QTextEdit()
        self.txt_logs.setReadOnly(True)
        self.txt_logs.setMinimumHeight(150)
        layout_logs.addWidget(self.txt_logs)
        
        layout_btns_log = QHBoxLayout()
        btn_limpiar = QPushButton("🗑️ Limpiar")
        btn_limpiar.clicked.connect(self.limpiar_logs)
        
        self.btn_buscar_poliza = QPushButton("🔍 Buscar Póliza")
        self.btn_buscar_poliza.setObjectName("btnBuscar")
        self.btn_buscar_poliza.clicked.connect(self.buscar_poliza)
        self.btn_buscar_poliza.setEnabled(False)
        self.btn_buscar_poliza.setToolTip("Busca una póliza específica en MAVISO y CELER")
        
        self.btn_agregar_faltantes = QPushButton("📥 Agregar Faltantes desde CELER")
        self.btn_agregar_faltantes.setObjectName("btnAgregarFaltantes")
        self.btn_agregar_faltantes.clicked.connect(self.agregar_faltantes_desde_celer)
        self.btn_agregar_faltantes.setEnabled(False)
        self.btn_agregar_faltantes.setToolTip("Agrega al archivo MAVISO las pólizas que están solo en CELER")
        
        self.btn_corregir = QPushButton("🔧 Corregir Modalidades")
        self.btn_corregir.setObjectName("btnCorregir")
        self.btn_corregir.clicked.connect(self.corregir_modalidades)
        self.btn_corregir.setEnabled(False)
        self.btn_corregir.setToolTip("Actualiza las modalidades en MAVISO para que coincidan con CELER")
        
        self.btn_corregir_primas = QPushButton("💰 Corregir Primas")
        self.btn_corregir_primas.setObjectName("btnCorregirPrimas")
        self.btn_corregir_primas.clicked.connect(self.corregir_primas)
        self.btn_corregir_primas.setEnabled(False)
        self.btn_corregir_primas.setToolTip("Actualiza las primas en MAVISO para que coincidan con CELER")
        
        self.btn_primas_mensuales_cero = QPushButton("🔄 Primas Mensuales a Cero")
        self.btn_primas_mensuales_cero.setObjectName("btnPrimasMensualesCero")
        self.btn_primas_mensuales_cero.clicked.connect(self.primas_mensuales_a_cero)
        self.btn_primas_mensuales_cero.setEnabled(False)
        self.btn_primas_mensuales_cero.setToolTip("Pone las primas en 0 para todas las pólizas MENSUALES")
        
        self.btn_colocar_nits = QPushButton("🔧 Colocar NITs Completos")
        self.btn_colocar_nits.setObjectName("btnColocarNits")
        self.btn_colocar_nits.clicked.connect(self.colocar_nits_completos)
        self.btn_colocar_nits.setEnabled(False)
        self.btn_colocar_nits.setToolTip("Calcula y agrega DV a NITs de personas jurídicas usando API DIAN")
        
        self.btn_corregir_vigencias = QPushButton("📅 Corregir Vigencias")
        self.btn_corregir_vigencias.setObjectName("btnCorregirVigencias")
        self.btn_corregir_vigencias.clicked.connect(self.corregir_vigencias)
        self.btn_corregir_vigencias.setEnabled(False)
        self.btn_corregir_vigencias.setToolTip("Actualiza las fechas de inicio y fin en MAVISO con los datos de CELER")
        
        self.btn_llenar_riesgos = QPushButton("📝 Llenar Riesgos Vacíos")
        self.btn_llenar_riesgos.setObjectName("btnLlenarRiesgos")
        self.btn_llenar_riesgos.clicked.connect(self.llenar_riesgos_vacios)
        self.btn_llenar_riesgos.setEnabled(False)
        self.btn_llenar_riesgos.setToolTip("Llena la columna RIESGO con SUBRAMO cuando está vacía")
        
        self.btn_exportar = QPushButton("📊 Exportar Reporte")
        self.btn_exportar.setObjectName("btnExportar")
        self.btn_exportar.clicked.connect(self.exportar_reporte)
        self.btn_exportar.setEnabled(False)
        layout_btns_log.addWidget(btn_limpiar)
        layout_btns_log.addWidget(self.btn_buscar_poliza)
        layout_btns_log.addWidget(self.btn_agregar_faltantes)
        layout_btns_log.addWidget(self.btn_corregir)
        layout_btns_log.addWidget(self.btn_corregir_primas)
        layout_btns_log.addWidget(self.btn_primas_mensuales_cero)
        layout_btns_log.addWidget(self.btn_colocar_nits)
        layout_btns_log.addWidget(self.btn_corregir_vigencias)
        layout_btns_log.addWidget(self.btn_llenar_riesgos)
        layout_btns_log.addStretch()
        layout_btns_log.addWidget(self.btn_exportar)
        layout_logs.addLayout(layout_btns_log)
        
        layout.addWidget(grupo_logs, 1)
    
    def seleccionar_archivo_maviso(self):
        carpeta_base = Path(__file__).parent.parent
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo Maviso", str(carpeta_base), "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            self.ruta_maviso = archivo
            self.lbl_ruta_maviso.setText(Path(archivo).name)
            self.log(f"✅ Archivo Maviso: {Path(archivo).name}", "success")
            self.verificar_archivos()
    
    def seleccionar_archivo_celer(self):
        carpeta_base = Path(__file__).parent.parent
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo CELER", str(carpeta_base), "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            self.ruta_celer = archivo
            self.lbl_ruta_celer.setText(Path(archivo).name)
            self.log(f"✅ Archivo CELER: {Path(archivo).name}", "success")
            self.verificar_archivos()
    
    def verificar_archivos(self):
        self.btn_comparar.setEnabled(self.ruta_maviso is not None and self.ruta_celer is not None)
    
    def ejecutar_comparacion(self):
        self.btn_comparar.setEnabled(False)
        self.btn_exportar.setEnabled(False)
        
        for widget in self.stats_widgets.values():
            widget.setText("--")
        
        self.log("", "info")
        self.log("📂 Cargando archivos...", "info")
        
        exito, mensaje = self.comparador.cargar_maviso(self.ruta_maviso)
        if not exito:
            self.log(f"❌ {mensaje}", "error")
            self.btn_comparar.setEnabled(True)
            return
        self.log(f"✅ {mensaje}", "success")
        
        exito, mensaje = self.comparador.cargar_celer(self.ruta_celer)
        if not exito:
            self.log(f"❌ {mensaje}", "error")
            self.btn_comparar.setEnabled(True)
            return
        self.log(f"✅ {mensaje}", "success")
        
        self.worker_thread = ComparadorThread(self.comparador, "celer")
        self.worker_thread.log_signal.connect(self.log)
        self.worker_thread.finished_signal.connect(self.comparacion_terminada)
        self.worker_thread.start()
    
    def comparacion_terminada(self, exito: bool, mensaje: str, estadisticas: dict):
        self.btn_comparar.setEnabled(True)
        
        if exito:
            for key, widget in self.stats_widgets.items():
                valor = estadisticas.get(key, 0)
                widget.setText(str(valor))
            
            self.log("", "info")
            self.log("=" * 50, "success")
            self.log("✅ COMPARACIÓN COMPLETADA", "success")
            self.log("=" * 50, "success")
            
            self.btn_exportar.setEnabled(True)
            self.btn_buscar_poliza.setEnabled(True)
            self.btn_agregar_faltantes.setEnabled(True)
            self.btn_corregir.setEnabled(True)
            self.btn_corregir_primas.setEnabled(True)
            self.btn_primas_mensuales_cero.setEnabled(True)
            self.btn_colocar_nits.setEnabled(True)
            self.btn_corregir_vigencias.setEnabled(True)
            self.btn_llenar_riesgos.setEnabled(True)
        else:
            self.log(f"❌ Error: {mensaje}", "error")
    
    def exportar_reporte(self):
        try:
            ruta = self.comparador.exportar_reporte()
            self.ruta_ultimo_reporte = ruta
            self.log(f"📊 Reporte exportado: {ruta}", "success")
            os.startfile(ruta)
        except Exception as e:
            self.log(f"❌ Error exportando: {str(e)}", "error")
    
    def log(self, mensaje: str, tipo: str = "info"):
        colores = {
            "info": "#d4d4d4",
            "success": "#6a9955",
            "warning": "#dcdcaa",
            "error": "#f44747"
        }
        color = colores.get(tipo, "#d4d4d4")
        self.txt_logs.append(f'<span style="color: {color};">{mensaje}</span>')
        scrollbar = self.txt_logs.verticalScrollBar()
        if scrollbar:
            scrollbar.setValue(scrollbar.maximum())
    
    def limpiar_logs(self):
        self.txt_logs.clear()
    
    def buscar_poliza(self):
        """Busca una póliza específica en MAVISO y CELER"""
        from PyQt6.QtWidgets import QInputDialog
        
        # Solicitar número de póliza
        poliza, ok = QInputDialog.getText(
            self,
            "Buscar Póliza",
            "Ingrese el número de póliza a buscar:"
        )
        
        if not ok or not poliza.strip():
            return
        
        poliza = poliza.strip()
        
        self.log("", "info")
        self.log("=" * 60, "info")
        self.log(f"🔍 BUSCANDO PÓLIZA: {poliza}", "info")
        self.log("=" * 60, "info")
        
        try:
            # Buscar en MAVISO
            import pandas as pd
            df_maviso = pd.read_excel(self.ruta_maviso, header=0)
            
            # Normalizar números de póliza en MAVISO
            maviso_encontrada = False
            fila_maviso = None
            datos_maviso = None
            
            for idx, row in df_maviso.iterrows():
                pol_maviso = str(row.iloc[0]).strip().upper()  # Columna A
                if pol_maviso == poliza.upper():
                    maviso_encontrada = True
                    fila_maviso = idx + 2  # +2 por header y 0-index
                    datos_maviso = {
                        'poliza': pol_maviso,
                        'prima': row.iloc[14],  # Columna O
                        'fecha_inicio': row.iloc[10],  # Columna K
                        'fecha_fin': row.iloc[11],  # Columna L
                        'modalidad': row.iloc[21],  # Columna V
                    }
                    break
            
            # Buscar en CELER
            df_celer = pd.read_excel(self.ruta_celer, skiprows=3)
            
            celer_encontrada = False
            fila_celer = None
            datos_celer = None
            
            for idx, row in df_celer.iterrows():
                pol_celer = str(row.iloc[20]).strip().upper()  # Columna U
                if pol_celer == poliza.upper():
                    celer_encontrada = True
                    fila_celer = idx + 5  # +5 por skiprows=3 + header + 0-index
                    datos_celer = {
                        'poliza': pol_celer,
                        'prima': row.iloc[42],  # Columna AQ
                        'fecha_inicio': row.iloc[22],  # Columna W
                        'fecha_fin': row.iloc[23],  # Columna X
                        'modalidad': row.iloc[27],  # Columna AB
                    }
                    break
            
            # Mostrar resultados
            self.log("", "info")
            self.log("📋 RESULTADOS DE BÚSQUEDA:", "info")
            self.log("-" * 60, "info")
            
            # MAVISO
            if maviso_encontrada:
                self.log(f"✅ ENCONTRADA EN MAVISO (Fila {fila_maviso}):", "success")
                self.log(f"   • Póliza: {datos_maviso['poliza']}", "info")
                self.log(f"   • Prima: {datos_maviso['prima']}", "info")
                self.log(f"   • Fecha Inicio: {datos_maviso['fecha_inicio']}", "info")
                self.log(f"   • Fecha Fin: {datos_maviso['fecha_fin']}", "info")
                self.log(f"   • Modalidad: {datos_maviso['modalidad']}", "info")
            else:
                self.log(f"❌ NO ENCONTRADA EN MAVISO", "error")
            
            self.log("", "info")
            
            # CELER
            if celer_encontrada:
                self.log(f"✅ ENCONTRADA EN CELER (Fila {fila_celer}):", "success")
                self.log(f"   • Póliza: {datos_celer['poliza']}", "info")
                self.log(f"   • Prima: {datos_celer['prima']}", "info")
                self.log(f"   • Fecha Inicio: {datos_celer['fecha_inicio']}", "info")
                self.log(f"   • Fecha Fin: {datos_celer['fecha_fin']}", "info")
                self.log(f"   • Modalidad: {datos_celer['modalidad']}", "info")
            else:
                self.log(f"❌ NO ENCONTRADA EN CELER", "error")
            
            self.log("-" * 60, "info")
            
            # Resumen
            if not maviso_encontrada and celer_encontrada:
                self.log("⚠️ CONCLUSIÓN: Póliza existe en CELER pero NO en MAVISO", "warning")
                self.log("   Esta póliza aparecerá en 'Solo en CELER' del reporte", "warning")
            elif maviso_encontrada and not celer_encontrada:
                self.log("⚠️ CONCLUSIÓN: Póliza existe en MAVISO pero NO en CELER", "warning")
                self.log("   Esta póliza aparecerá en 'Solo en MAVISO' del reporte", "warning")
            elif maviso_encontrada and celer_encontrada:
                self.log("✅ CONCLUSIÓN: Póliza existe en ambos archivos", "success")
            else:
                self.log("❌ CONCLUSIÓN: Póliza NO existe en ningún archivo", "error")
            
            self.log("=" * 60, "info")
            
        except Exception as e:
            self.log(f"❌ Error al buscar: {str(e)}", "error")
    
    def agregar_faltantes_desde_celer(self):
        """Agrega al archivo MAVISO las pólizas que están solo en CELER"""
        try:
            if not self.comparador or not self.comparador.resultados:
                self.log("⚠️ No hay resultados de comparación. Ejecuta primero la comparación.", "warning")
                return
            
            solo_celer = self.comparador.resultados.get('solo_celer', [])
            
            if not solo_celer:
                self.log("✅ No hay pólizas faltantes en MAVISO. Todos los datos ya están sincronizados.", "success")
                return
            
            self.log("", "info")
            self.log("="*60, "info")
            self.log(f"📥 AGREGANDO {len(solo_celer)} PÓLIZAS FALTANTES DESDE CELER", "info")
            self.log("="*60, "info")
            
            # Leer archivos con openpyxl para mantener formato
            from openpyxl import load_workbook
            from datetime import datetime
            import pandas as pd
            
            # Cargar MAVISO con openpyxl
            wb_maviso = load_workbook(self.ruta_maviso)
            ws_maviso = wb_maviso.active
            
            # Cargar CELER con pandas (skiprows=3)
            df_celer = pd.read_excel(self.ruta_celer, engine='openpyxl', skiprows=3)
            
            # Normalizar columna de pólizas en CELER
            poliza_col_celer = df_celer.iloc[:, 20]  # Columna U
            poliza_col_celer = poliza_col_celer.astype(str).str.strip().str.upper()
            
            agregadas = 0
            errores = 0
            
            for item_poliza in solo_celer:
                try:
                    # Extraer número de póliza del diccionario
                    poliza = item_poliza['poliza']
                    
                    # Buscar la póliza en CELER
                    poliza_normalizada = str(poliza).strip().upper()
                    mascara = poliza_col_celer == poliza_normalizada
                    
                    if not mascara.any():
                        self.log(f"⚠️ Póliza {poliza} no encontrada en CELER (extraño, debería estar)", "warning")
                        errores += 1
                        continue
                    
                    # Obtener la fila de CELER (primera coincidencia)
                    idx_celer = mascara.idxmax()
                    fila_celer = df_celer.iloc[idx_celer]
                    
                    # Mapeo CELER → MAVISO según README
                    # Crear nueva fila para MAVISO
                    nueva_fila = []
                    
                    # MAPEO DE COLUMNAS (índices 0-based)
                    # A (0): NÚMERO DE PÓLIZA ← CELER U (20)
                    nueva_fila.append(fila_celer.iloc[20])
                    
                    # B (1): RIESGO ← CELER AE (30) - Placa
                    nueva_fila.append(fila_celer.iloc[30])
                    
                    # C (2): ASEGURADORA ← CELER R (17)
                    nueva_fila.append(fila_celer.iloc[17])
                    
                    # D (3): vacío
                    nueva_fila.append("")
                    
                    # E (4): SUBRAMO ← CELER S (18) - Ramo
                    nueva_fila.append(fila_celer.iloc[18])
                    
                    # F-I (5-8): vacíos por ahora
                    for _ in range(4):
                        nueva_fila.append("")
                    
                    # J (9): FECHA INICIO ← CELER W (22)
                    nueva_fila.append(fila_celer.iloc[22])
                    
                    # K (10): FECHA FIN ← CELER X (23)
                    nueva_fila.append(fila_celer.iloc[23])
                    
                    # L-N (11-13): vacíos
                    for _ in range(3):
                        nueva_fila.append("")
                    
                    # O (14): PRIMA NETA ← CELER AQ (42)
                    nueva_fila.append(fila_celer.iloc[42])
                    
                    # P-V (15-21): vacíos por ahora, excepto V
                    for _ in range(6):
                        nueva_fila.append("")
                    
                    # V (21): MODALIDAD ← CELER AB (27)
                    nueva_fila.append(fila_celer.iloc[27])
                    
                    # W (22): FORMA PAGO - lógica condicional
                    forma_pago_celer = str(fila_celer.iloc[27]).strip().upper()
                    if forma_pago_celer == "MENSUAL":
                        nueva_fila.append("Fraccionado")
                    elif forma_pago_celer == "ANUAL":
                        nueva_fila.append("Contado")
                    else:
                        nueva_fila.append("")
                    
                    # X (23): VALOR RIESGO ASEGURADO ← CELER AP (41)
                    nueva_fila.append(fila_celer.iloc[41])
                    
                    # Y-AA (24-26): vacíos
                    for _ in range(3):
                        nueva_fila.append("")
                    
                    # AB (27): DOCUMENTO DEL CLIENTE ← CELER C (2)
                    nueva_fila.append(fila_celer.iloc[2])
                    
                    # AC (28): TIPO PERSONA ← CELER A (0)
                    nueva_fila.append(fila_celer.iloc[0])
                    
                    # AD (29): NOMBRE DEL TOMADOR ← CELER B (1)
                    nueva_fila.append(fila_celer.iloc[1])
                    
                    # AE (30): DOCUMENTO DEL TOMADOR ← CELER C (2)
                    nueva_fila.append(fila_celer.iloc[2])
                    
                    # AF (31): NOMBRE DEL ASEGURADO ← CELER AS (44)
                    nueva_fila.append(fila_celer.iloc[44])
                    
                    # AG (32): DOCUMENTO DEL ASEGURADO ← CELER AT (45)
                    nueva_fila.append(fila_celer.iloc[45])
                    
                    # AH (33): NOMBRE DEL BENEFICIARIO ← CELER AW (48)
                    nueva_fila.append(fila_celer.iloc[48])
                    
                    # AI (34): DOCUMENTO DEL BENEFICIARIO ← CELER AX (49)
                    nueva_fila.append(fila_celer.iloc[49])
                    
                    # AJ-AM (35-38): vacíos
                    for _ in range(4):
                        nueva_fila.append("")
                    
                    # Agregar fila al worksheet de MAVISO
                    ws_maviso.append(nueva_fila)
                    
                    agregadas += 1
                    self.log(f"✅ Póliza {poliza} agregada desde CELER", "success")
                    
                except Exception as e:
                    self.log(f"❌ Error al procesar póliza {poliza}: {str(e)}", "error")
                    errores += 1
                    continue
            
            # Guardar archivo MAVISO actualizado
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            carpeta_output = os.path.dirname(self.ruta_maviso)
            nombre_salida = f"MAVISO_con_faltantes_{timestamp}.xlsx"
            ruta_salida = os.path.join(carpeta_output, nombre_salida)
            
            wb_maviso.save(ruta_salida)
            
            self.log("", "info")
            self.log("="*60, "success")
            self.log(f"✅ PROCESO COMPLETADO", "success")
            self.log(f"📊 Pólizas agregadas: {agregadas}", "success")
            if errores > 0:
                self.log(f"⚠️ Errores: {errores}", "warning")
            self.log(f"💾 Archivo guardado: {nombre_salida}", "success")
            self.log("="*60, "success")
            
            # Preguntar si quiere recomparar con el nuevo archivo
            from PyQt6.QtWidgets import QMessageBox
            respuesta = QMessageBox.question(
                self,
                "Recomparar",
                f"Se agregaron {agregadas} pólizas al archivo MAVISO.\n\n"
                f"¿Deseas recargar y recomparar con el nuevo archivo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if respuesta == QMessageBox.StandardButton.Yes:
                self.ruta_maviso = ruta_salida
                self.lbl_ruta_maviso.setText(ruta_salida)
                self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al agregar faltantes: {str(e)}", "error")
            import traceback
            self.log(f"Detalles: {traceback.format_exc()}", "error")
    
    def recomparar_automaticamente(self):
        """Re-ejecuta la comparación para actualizar estadísticas después de correcciones"""
        try:
            # Recargar archivos
            exito, mensaje = self.comparador.cargar_maviso(self.ruta_maviso)
            if not exito:
                self.log(f"⚠️ No se pudo recargar MAVISO: {mensaje}", "warning")
                return
            
            exito, mensaje = self.comparador.cargar_celer(self.ruta_celer)
            if not exito:
                self.log(f"⚠️ No se pudo recargar CELER: {mensaje}", "warning")
                return
            
            # Ejecutar comparación en background
            self.worker_thread = ComparadorThread(self.comparador, "celer")
            self.worker_thread.log_signal.connect(self.log)
            self.worker_thread.finished_signal.connect(self.comparacion_actualizada)
            self.worker_thread.start()
            
        except Exception as e:
            self.log(f"⚠️ Error al recomparar: {str(e)}", "warning")
    
    def comparacion_actualizada(self, exito: bool, mensaje: str, estadisticas: dict):
        """Callback cuando termina la recomparación automática"""
        if exito:
            # Actualizar widgets de estadísticas
            for key, widget in self.stats_widgets.items():
                valor = estadisticas.get(key, 0)
                widget.setText(str(valor))
            
            self.log("✅ Comparación actualizada - El reporte reflejará los cambios", "success")
        else:
            self.log(f"⚠️ No se pudo actualizar la comparación: {mensaje}", "warning")
    
    def corregir_modalidades(self):
        """Corrige las modalidades en MAVISO para que coincidan con CELER"""
        if not self.ruta_maviso or not self.comparador.resultados:
            QMessageBox.warning(self, "Advertencia", "Debe ejecutar la comparación primero")
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Corrección",
            "¿Desea actualizar las modalidades en MAVISO para que coincidan con CELER?\n\n"
            "CELER se usará como fuente correcta.\n"
            "⚠️ Se modificará el archivo original manteniendo formato y colores.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("🔧 Iniciando corrección de modalidades...", "info")
            
            # Cargar CELER con pandas para crear diccionario
            df_celer = pd.read_excel(self.ruta_celer, skiprows=3)
            self.log(f"📂 Archivo CELER cargado: {len(df_celer)} registros", "info")
            
            # Columnas
            MAVISO_COL_POLIZA = 0  # Columna A (índice 0)
            MAVISO_COL_MODALIDAD = 21  # Columna V (índice 21)
            CELER_COL_POLIZA = 20  # Columna U
            CELER_COL_MODALIDAD = 27  # Columna AB
            
            # Crear diccionario de CELER
            celer_dict = {}
            for idx, row in df_celer.iterrows():
                poliza = str(row.iloc[CELER_COL_POLIZA]).strip()
                modalidad = str(row.iloc[CELER_COL_MODALIDAD]).strip().upper()
                celer_dict[poliza] = modalidad
            
            self.log(f"📋 Diccionario CELER creado: {len(celer_dict)} pólizas", "info")
            
            # Cargar MAVISO con openpyxl para mantener formato
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            total_filas = ws.max_row
            self.log(f"📂 Archivo MAVISO cargado: {total_filas - 1} registros", "info")
            
            # Corregir MAVISO manteniendo formato
            cambios = []
            for fila in range(2, total_filas + 1):  # Empezar desde fila 2 (después del encabezado)
                # Columna A = 1, Columna V = 22 (en openpyxl las columnas empiezan en 1)
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_modalidad = ws.cell(row=fila, column=MAVISO_COL_MODALIDAD + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                modalidad_maviso = str(celda_modalidad.value).strip().upper() if celda_modalidad.value else ""
                
                if poliza in celer_dict:
                    modalidad_celer = celer_dict[poliza]
                    if modalidad_maviso != modalidad_celer and modalidad_celer in ['MENSUAL', 'ANUAL']:
                        # Actualizar solo el valor, manteniendo formato
                        celda_modalidad.value = modalidad_celer
                        cambios.append({
                            'poliza': poliza,
                            'fila': fila,
                            'anterior': modalidad_maviso,
                            'nuevo': modalidad_celer
                        })
            
            if len(cambios) == 0:
                self.log("✅ No se encontraron discrepancias de modalidad para corregir", "success")
                QMessageBox.information(self, "Información", "No hay cambios necesarios.")
                wb.close()
                return
            
            # Guardar archivo sobrescribiendo el original
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"\n✅ Corrección completada:", "success")
            self.log(f"   Cambios realizados: {len(cambios)}", "success")
            self.log(f"   Archivo actualizado: {Path(self.ruta_maviso).name}", "success")
            
            # Mostrar resumen
            self.log("\n📋 Resumen de cambios:", "info")
            for i, cambio in enumerate(cambios[:10], 1):
                self.log(
                    f"   {i}. Póliza {cambio['poliza']} (Fila {cambio['fila']}): "
                    f"{cambio['anterior']} → {cambio['nuevo']}",
                    "warning"
                )
            
            if len(cambios) > 10:
                self.log(f"   ... y {len(cambios) - 10} cambios más", "info")
            
            QMessageBox.information(
                self,
                "Corrección Exitosa",
                f"Se corrigieron {len(cambios)} modalidades.\n\n"
                f"El archivo original ha sido actualizado manteniendo su formato.\n\n"
                f"🔄 Se volverá a ejecutar la comparación para actualizar el reporte."
            )
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
            
            # Recomparar para actualizar estadísticas
            self.log("", "info")
            self.log("🔄 Actualizando comparación...", "info")
            self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al corregir: {str(e)}", "error")
            QMessageBox.critical(self, "Error", f"Error al corregir modalidades:\n{str(e)}")
    
    def corregir_primas(self):
        """Corrige las primas en MAVISO para que coincidan con CELER"""
        if not self.ruta_maviso or not self.comparador.resultados:
            QMessageBox.warning(self, "Advertencia", "Debe ejecutar la comparación primero")
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Corrección",
            "¿Desea actualizar las primas en MAVISO para que coincidan con CELER?\n\n"
            "CELER se usará como fuente correcta.\n"
            "⚠️ Se modificará el archivo original manteniendo formato y colores.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("💰 Iniciando corrección de primas...", "info")
            
            # Cargar CELER con pandas para crear diccionario
            df_celer = pd.read_excel(self.ruta_celer, skiprows=3)
            self.log(f"📂 Archivo CELER cargado: {len(df_celer)} registros", "info")
            
            # Columnas
            MAVISO_COL_POLIZA = 0  # Columna A (índice 0)
            MAVISO_COL_PRIMA = 14  # Columna O (índice 14)
            CELER_COL_POLIZA = 20  # Columna U
            CELER_COL_PRIMA = 42  # Columna AQ - Prima sin IVA
            
            # Crear diccionario de CELER
            celer_dict = {}
            for idx, row in df_celer.iterrows():
                poliza = str(row.iloc[CELER_COL_POLIZA]).strip()
                try:
                    prima = abs(round(float(row.iloc[CELER_COL_PRIMA]), 2))
                    celer_dict[poliza] = prima
                except (ValueError, TypeError):
                    continue
            
            self.log(f"📋 Diccionario CELER creado: {len(celer_dict)} pólizas", "info")
            
            # Cargar MAVISO con openpyxl para mantener formato
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            total_filas = ws.max_row
            self.log(f"📂 Archivo MAVISO cargado: {total_filas - 1} registros", "info")
            
            # Corregir MAVISO manteniendo formato
            cambios = []
            for fila in range(2, total_filas + 1):  # Empezar desde fila 2 (después del encabezado)
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_prima = ws.cell(row=fila, column=MAVISO_COL_PRIMA + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                
                try:
                    prima_maviso = abs(round(float(celda_prima.value), 2)) if celda_prima.value else 0
                except (ValueError, TypeError):
                    prima_maviso = 0
                
                if poliza in celer_dict:
                    prima_celer = celer_dict[poliza]
                    if prima_maviso != prima_celer:
                        # Actualizar solo el valor, manteniendo formato
                        celda_prima.value = prima_celer
                        cambios.append({
                            'poliza': poliza,
                            'fila': fila,
                            'anterior': prima_maviso,
                            'nuevo': prima_celer
                        })
            
            if len(cambios) == 0:
                self.log("✅ No se encontraron discrepancias de prima para corregir", "success")
                QMessageBox.information(self, "Información", "No hay cambios necesarios.")
                wb.close()
                return
            
            # Guardar archivo sobrescribiendo el original
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"\n✅ Corrección completada:", "success")
            self.log(f"   Cambios realizados: {len(cambios)}", "success")
            self.log(f"   Archivo actualizado: {Path(self.ruta_maviso).name}", "success")
            
            # Mostrar resumen
            self.log("\n📋 Resumen de cambios:", "info")
            for i, cambio in enumerate(cambios[:10], 1):
                self.log(
                    f"   {i}. Póliza {cambio['poliza']} (Fila {cambio['fila']}): "
                    f"{cambio['anterior']:,.2f} → {cambio['nuevo']:,.2f}",
                    "warning"
                )
            
            if len(cambios) > 10:
                self.log(f"   ... y {len(cambios) - 10} cambios más", "info")
            
            QMessageBox.information(
                self,
                "Corrección Exitosa",
                f"Se corrigieron {len(cambios)} primas.\n\n"
                f"El archivo original ha sido actualizado manteniendo su formato.\n\n"
                f"🔄 Se volverá a ejecutar la comparación para actualizar el reporte."
            )
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
            
            # Recomparar para actualizar estadísticas
            self.log("", "info")
            self.log("🔄 Actualizando comparación...", "info")
            self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al corregir primas: {str(e)}", "error")
            QMessageBox.critical(self, "Error", f"Error al corregir primas:\n{str(e)}")
    
    def corregir_vigencias(self):
        """Actualiza las fechas de inicio y fin en MAVISO con los datos de CELER"""
        if not self.ruta_maviso or not self.ruta_celer:
            QMessageBox.warning(self, "Advertencia", "Debe cargar ambos archivos primero")
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Operación",
            "¿Desea actualizar las VIGENCIAS (fechas inicio y fin) en MAVISO con los datos de CELER?\n\n"
            "Esta operación actualizará TODAS las pólizas que coincidan.\n"
            "⚠️ Se modificará el archivo original manteniendo formato y colores.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("📅 Iniciando corrección de vigencias...", "info")
            
            # Columnas según README
            MAVISO_COL_POLIZA = 0  # Columna A
            MAVISO_COL_FECHA_INICIO = 10  # Columna K
            MAVISO_COL_FECHA_FIN = 11  # Columna L
            
            CELER_COL_POLIZA = 20  # Columna U
            CELER_COL_FECHA_INICIO = 22  # Columna W
            CELER_COL_FECHA_FIN = 23  # Columna X
            
            # Leer archivos
            from openpyxl import load_workbook
            import pandas as pd
            
            wb_maviso = load_workbook(self.ruta_maviso)
            ws_maviso = wb_maviso.active
            
            df_celer = pd.read_excel(self.ruta_celer, engine='openpyxl', skiprows=3)
            
            # Crear diccionario de vigencias CELER: {poliza: (fecha_inicio, fecha_fin)}
            vigencias_celer = {}
            for idx, row in df_celer.iterrows():
                poliza = str(row.iloc[CELER_COL_POLIZA]).strip().upper()
                fecha_inicio = row.iloc[CELER_COL_FECHA_INICIO]
                fecha_fin = row.iloc[CELER_COL_FECHA_FIN]
                vigencias_celer[poliza] = (fecha_inicio, fecha_fin)
            
            self.log(f"📊 Vigencias CELER cargadas: {len(vigencias_celer)}", "info")
            
            # Actualizar MAVISO
            actualizadas = 0
            no_encontradas = 0
            
            for row_idx in range(2, ws_maviso.max_row + 1):
                poliza_maviso = ws_maviso.cell(row_idx, MAVISO_COL_POLIZA + 1).value
                
                if poliza_maviso is None:
                    continue
                
                poliza_normalizada = str(poliza_maviso).strip().upper()
                
                if poliza_normalizada in vigencias_celer:
                    fecha_inicio_celer, fecha_fin_celer = vigencias_celer[poliza_normalizada]
                    
                    # Actualizar fechas en MAVISO
                    ws_maviso.cell(row_idx, MAVISO_COL_FECHA_INICIO + 1).value = fecha_inicio_celer
                    ws_maviso.cell(row_idx, MAVISO_COL_FECHA_FIN + 1).value = fecha_fin_celer
                    
                    actualizadas += 1
                    if actualizadas <= 10:  # Mostrar solo las primeras 10
                        self.log(f"✅ Póliza {poliza_normalizada}: Vigencias actualizadas", "success")
                else:
                    no_encontradas += 1
            
            if actualizadas > 10:
                self.log(f"... y {actualizadas - 10} más actualizadas", "success")
            
            # Guardar cambios
            wb_maviso.save(self.ruta_maviso)
            
            self.log("", "info")
            self.log("="*60, "success")
            self.log(f"✅ Vigencias corregidas: {actualizadas}", "success")
            if no_encontradas > 0:
                self.log(f"⚠️ Pólizas no encontradas en CELER: {no_encontradas}", "warning")
            self.log(f"💾 Archivo actualizado: {os.path.basename(self.ruta_maviso)}", "success")
            self.log("="*60, "success")
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
            
            # Recomparar para actualizar estadísticas
            self.log("", "info")
            self.log("🔄 Actualizando comparación...", "info")
            self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al corregir vigencias: {str(e)}", "error")
            QMessageBox.critical(self, "Error", f"Error al corregir vigencias:\n{str(e)}")
    
    def llenar_riesgos_vacios(self):
        """Llena la columna RIESGO (B) con SUBRAMO (F) cuando está vacía"""
        if not self.ruta_maviso:
            QMessageBox.warning(self, "Advertencia", "Debe cargar el archivo MAVISO primero")
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Operación",
            "¿Desea llenar los RIESGOS vacíos con el valor de SUBRAMO?\n\n"
            "Si la columna B (RIESGO) está vacía, se copiará el valor de la columna F (SUBRAMO).\n"
            "⚠️ Se modificará el archivo original manteniendo formato y colores.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("📝 Llenando riesgos vacíos...", "info")
            
            # Columnas
            MAVISO_COL_POLIZA = 0  # Columna A
            MAVISO_COL_RIESGO = 1  # Columna B
            MAVISO_COL_SUBRAMO = 5  # Columna F
            
            # Leer archivo con openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            llenados = 0
            
            # Recorrer todas las filas (empezando desde fila 2, asumiendo encabezados en fila 1)
            for row_idx in range(2, ws.max_row + 1):
                riesgo = ws.cell(row_idx, MAVISO_COL_RIESGO + 1).value
                subramo = ws.cell(row_idx, MAVISO_COL_SUBRAMO + 1).value
                poliza = ws.cell(row_idx, MAVISO_COL_POLIZA + 1).value
                
                # Si RIESGO está vacío y SUBRAMO tiene valor
                if (riesgo is None or str(riesgo).strip() == "") and subramo:
                    # Copiar SUBRAMO a RIESGO
                    ws.cell(row_idx, MAVISO_COL_RIESGO + 1).value = subramo
                    llenados += 1
                    
                    if llenados <= 10:  # Mostrar solo los primeros 10
                        self.log(f"✅ Póliza {poliza}: RIESGO ← '{subramo}'", "success")
            
            if llenados > 10:
                self.log(f"... y {llenados - 10} más actualizados", "success")
            
            if llenados == 0:
                self.log("ℹ️ No se encontraron riesgos vacíos para llenar", "info")
            
            # Guardar cambios
            wb.save(self.ruta_maviso)
            
            self.log("", "info")
            self.log("="*60, "success")
            self.log(f"✅ Riesgos llenados: {llenados}", "success")
            self.log(f"💾 Archivo actualizado: {os.path.basename(self.ruta_maviso)}", "success")
            self.log("="*60, "success")
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
            
            # Recomparar para actualizar estadísticas (si es necesario)
            if llenados > 0:
                self.log("", "info")
                self.log("🔄 Actualizando comparación...", "info")
                self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al llenar riesgos: {str(e)}", "error")
            QMessageBox.critical(self, "Error", f"Error al llenar riesgos:\n{str(e)}")
    
    def primas_mensuales_a_cero(self):
        """Pone las primas en 0 para todas las pólizas MENSUALES"""
        if not self.ruta_maviso:
            QMessageBox.warning(self, "Advertencia", "Debe cargar el archivo MAVISO primero")
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Operación",
            "¿Desea poner en CERO las primas de todas las pólizas MENSUALES?\n\n"
            "Esta operación afectará todas las pólizas con modalidad MENSUAL.\n"
            "⚠️ Se modificará el archivo original manteniendo formato y colores.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("🔄 Iniciando ajuste de primas mensuales...", "info")
            
            # Columnas
            MAVISO_COL_POLIZA = 0  # Columna A (índice 0)
            MAVISO_COL_PRIMA = 14  # Columna O (índice 14)
            MAVISO_COL_MODALIDAD = 21  # Columna V (índice 21)
            
            # Cargar MAVISO con openpyxl para mantener formato
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            total_filas = ws.max_row
            self.log(f"📂 Archivo MAVISO cargado: {total_filas - 1} registros", "info")
            
            # Poner primas a cero para MENSUALES
            cambios = []
            for fila in range(2, total_filas + 1):  # Empezar desde fila 2 (después del encabezado)
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_prima = ws.cell(row=fila, column=MAVISO_COL_PRIMA + 1)
                celda_modalidad = ws.cell(row=fila, column=MAVISO_COL_MODALIDAD + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                modalidad = str(celda_modalidad.value).strip().upper() if celda_modalidad.value else ""
                
                try:
                    prima_actual = abs(round(float(celda_prima.value), 2)) if celda_prima.value else 0
                except (ValueError, TypeError):
                    prima_actual = 0
                
                # Si es MENSUAL y la prima no es cero
                if modalidad == "MENSUAL" and prima_actual != 0:
                    celda_prima.value = 0
                    cambios.append({
                        'poliza': poliza,
                        'fila': fila,
                        'prima_anterior': prima_actual
                    })
            
            if len(cambios) == 0:
                self.log("✅ No se encontraron pólizas mensuales con prima diferente de cero", "success")
                QMessageBox.information(self, "Información", "No hay cambios necesarios.")
                wb.close()
                return
            
            # Guardar archivo sobrescribiendo el original
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"\n✅ Operación completada:", "success")
            self.log(f"   Primas modificadas: {len(cambios)}", "success")
            self.log(f"   Archivo actualizado: {Path(self.ruta_maviso).name}", "success")
            
            # Mostrar resumen
            self.log("\n📋 Resumen de cambios:", "info")
            for i, cambio in enumerate(cambios[:10], 1):
                self.log(
                    f"   {i}. Póliza {cambio['poliza']} (Fila {cambio['fila']}): "
                    f"{cambio['prima_anterior']:,.2f} → 0.00",
                    "warning"
                )
            
            if len(cambios) > 10:
                self.log(f"   ... y {len(cambios) - 10} cambios más", "info")
            
            QMessageBox.information(
                self,
                "Operación Exitosa",
                f"Se pusieron en CERO las primas de {len(cambios)} pólizas MENSUALES.\n\n"
                f"El archivo original ha sido actualizado manteniendo su formato.\n\n"
                f"🔄 Se volverá a ejecutar la comparación para actualizar el reporte."
            )
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
            
            # Recomparar para actualizar estadísticas
            self.log("", "info")
            self.log("🔄 Actualizando comparación...", "info")
            self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al ajustar primas: {str(e)}", "error")
            QMessageBox.critical(self, "Error", f"Error al ajustar primas mensuales:\n{str(e)}")
    
    def _verificar_api_dian(self):
        """Verifica si la API DIAN está corriendo"""
        try:
            response = requests.get("http://localhost:8000/health", timeout=2)
            return response.status_code == 200
        except:
            return False
    
    def _iniciar_api_dian(self):
        """Inicia la API DIAN en segundo plano"""
        if self._verificar_api_dian():
            return True
        
        self.log("🚀 Iniciando API DIAN...", "info")
        
        # Ruta al directorio backend (dos niveles arriba desde comparador_archivos)
        script_dir = Path(__file__).parent.parent.parent
        backend_dir = script_dir / 'backend'
        
        if not backend_dir.exists():
            self.log(f"❌ No se encuentra el directorio backend: {backend_dir}", "error")
            return False
        
        try:
            # Iniciar uvicorn en segundo plano
            subprocess.Popen(
                [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8000"],
                cwd=str(backend_dir),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            # Esperar a que la API esté lista
            for i in range(10):
                time.sleep(1)
                if self._verificar_api_dian():
                    self.log("✅ API DIAN iniciada correctamente", "success")
                    return True
                self.log(f"   Esperando API... ({i+1}/10)", "info")
            
            self.log("❌ No se pudo iniciar la API DIAN", "error")
            return False
            
        except Exception as e:
            self.log(f"❌ Error iniciando API: {str(e)}", "error")
            return False
    
    def _calcular_dv_dian(self, nit, reintentos=3):
        """Calcula el dígito de verificación usando la API DIAN con reintentos"""
        nit_str = re.sub(r'\D', '', str(nit))
        
        if not nit_str or len(nit_str) > 15 or len(nit_str) < 5:
            return None
        
        for intento in range(reintentos):
            try:
                response = requests.post(
                    "http://localhost:8000/calcular",
                    json={"nit": nit_str},
                    timeout=10
                )
                
                if response.status_code == 200:
                    data = response.json()
                    return str(data["digito_verificacion"])
                else:
                    if intento < reintentos - 1:
                        time.sleep(0.5)
                        continue
                    return None
            
            except requests.exceptions.ConnectionError:
                if intento < reintentos - 1:
                    if not self._verificar_api_dian():
                        self._iniciar_api_dian()
                        time.sleep(2)
                    continue
                return None
            except requests.exceptions.Timeout:
                if intento < reintentos - 1:
                    time.sleep(1)
                    continue
                return None
            except Exception:
                if intento < reintentos - 1:
                    time.sleep(0.5)
                    continue
                return None
        
        return None
    
    def colocar_nits_completos(self):
        """Calcula y agrega DV a NITs de personas jurídicas usando API DIAN"""
        if not self.ruta_maviso:
            QMessageBox.warning(self, "Advertencia", "Debe cargar el archivo MAVISO primero")
            return
        
        respuesta = QMessageBox.question(
            self,
            "Confirmar Operación",
            "¿Desea calcular y agregar el DV a los NITs de personas jurídicas?\n\n"
            "Esta operación:\n"
            "- Usa la API DIAN para calcular el DV\n"
            "- Solo procesa personas JURÍDICAS (columna AC = 'J')\n"
            "- Ignora personas NATURALES (columna AC = 'N')\n"
            "⚠️ Se modificará el archivo original manteniendo formato y colores.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("🔧 Iniciando cálculo de NITs completos...", "info")
            
            # Verificar/Iniciar API DIAN
            if not self._verificar_api_dian():
                if not self._iniciar_api_dian():
                    QMessageBox.critical(
                        self,
                        "Error API",
                        "No se pudo iniciar la API DIAN.\n\n"
                        "Verifique que el directorio 'backend' existe y tiene los archivos necesarios."
                    )
                    return
            else:
                self.log("✅ API DIAN disponible", "success")
            
            # Columnas
            MAVISO_COL_POLIZA = 0  # Columna A
            MAVISO_COL_DOCUMENTO = 27  # Columna AB (índice 27)
            MAVISO_COL_TIPO_PERSONA = 28  # Columna AC (índice 28)
            
            # Cargar MAVISO con openpyxl para mantener formato
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            total_filas = ws.max_row
            self.log(f"📂 Archivo MAVISO cargado: {total_filas - 1} registros", "info")
            
            # Procesar NITs
            cambios = []
            errores = []
            naturales_ignoradas = 0
            
            for fila in range(2, total_filas + 1):
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_documento = ws.cell(row=fila, column=MAVISO_COL_DOCUMENTO + 1)
                celda_tipo = ws.cell(row=fila, column=MAVISO_COL_TIPO_PERSONA + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                documento = str(celda_documento.value).strip() if celda_documento.value else ""
                tipo_persona = str(celda_tipo.value).strip().upper() if celda_tipo.value else ""
                
                # Solo procesar personas JURÍDICAS
                if tipo_persona != 'J':
                    if tipo_persona == 'N':
                        naturales_ignoradas += 1
                    continue
                
                # Verificar si ya tiene DV
                if '-' in documento:
                    continue
                
                # Verificar que sea un número válido
                documento_limpio = re.sub(r'\D', '', documento)
                if not documento_limpio or len(documento_limpio) < 5:
                    continue
                
                # Calcular DV usando API DIAN
                dv = self._calcular_dv_dian(documento_limpio)
                
                if dv is not None:
                    nit_completo = f"{documento_limpio}-{dv}"
                    celda_documento.value = nit_completo
                    cambios.append({
                        'poliza': poliza,
                        'fila': fila,
                        'anterior': documento,
                        'nuevo': nit_completo
                    })
                else:
                    errores.append({
                        'poliza': poliza,
                        'fila': fila,
                        'documento': documento
                    })
                
                # Mostrar progreso cada 50 registros
                if len(cambios) % 50 == 0 and len(cambios) > 0:
                    self.log(f"   Procesados: {len(cambios)} NITs...", "info")
            
            self.log(f"\n📊 Estadísticas:", "info")
            self.log(f"   Personas naturales ignoradas: {naturales_ignoradas}", "info")
            self.log(f"   NITs procesados exitosamente: {len(cambios)}", "success" if len(cambios) > 0 else "info")
            
            if len(errores) > 0:
                self.log(f"   Errores al calcular DV: {len(errores)}", "warning")
            
            if len(cambios) == 0:
                self.log("\n✅ No se encontraron NITs de personas jurídicas sin DV", "success")
                QMessageBox.information(self, "Información", "No hay cambios necesarios.")
                wb.close()
                return
            
            # Guardar archivo
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"\n✅ Operación completada:", "success")
            self.log(f"   NITs actualizados: {len(cambios)}", "success")
            self.log(f"   Archivo actualizado: {Path(self.ruta_maviso).name}", "success")
            
            # Mostrar resumen
            self.log("\n📋 Resumen de cambios:", "info")
            for i, cambio in enumerate(cambios[:10], 1):
                self.log(
                    f"   {i}. Póliza {cambio['poliza']} (Fila {cambio['fila']}): "
                    f"{cambio['anterior']} → {cambio['nuevo']}",
                    "warning"
                )
            
            if len(cambios) > 10:
                self.log(f"   ... y {len(cambios) - 10} cambios más", "info")
            
            mensaje_resultado = f"Se actualizaron {len(cambios)} NITs con su dígito de verificación.\n\n"
            if len(errores) > 0:
                mensaje_resultado += f"⚠️ {len(errores)} NITs no pudieron ser calculados.\n\n"
            mensaje_resultado += "El archivo original ha sido actualizado manteniendo su formato.\n\n"
            mensaje_resultado += "🔄 Se volverá a ejecutar la comparación para actualizar el reporte."
            
            QMessageBox.information(
                self,
                "Operación Exitosa",
                mensaje_resultado
            )
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
            
            # Recomparar para actualizar estadísticas
            self.log("", "info")
            self.log("🔄 Actualizando comparación...", "info")
            self.recomparar_automaticamente()
            
        except Exception as e:
            self.log(f"❌ Error al colocar NITs: {str(e)}", "error")
            QMessageBox.critical(self, "Error", f"Error al colocar NITs completos:\n{str(e)}")


# ==================================================================================
# Thread para Copiar Datos
# ==================================================================================

class CopiadorDatosThread(QThread):
    """Thread para copiar datos entre archivos"""
    log_signal = pyqtSignal(str, str)
    finished_signal = pyqtSignal(bool, str, dict)
    
    def __init__(self, archivo_base, skiprows_base, col_poliza_base,
                 archivo_destino, skiprows_destino, col_poliza_destino,
                 mapeos, archivo_salida):
        super().__init__()
        self.archivo_base = archivo_base
        self.skiprows_base = skiprows_base
        self.col_poliza_base = col_poliza_base
        self.archivo_destino = archivo_destino
        self.skiprows_destino = skiprows_destino
        self.col_poliza_destino = col_poliza_destino
        self.mapeos = mapeos  # [(col_base, col_destino), ...]
        self.archivo_salida = archivo_salida
    
    def run(self):
        try:
            from openpyxl import load_workbook
            
            # Leer archivo BASE
            self.log_signal.emit("📂 Leyendo archivo BASE...", "info")
            df_base = pd.read_excel(self.archivo_base, skiprows=self.skiprows_base, dtype=str)
            self.log_signal.emit(f"✅ BASE: {len(df_base)} registros", "success")
            
            # Leer archivo DESTINO con openpyxl para preservar formato
            self.log_signal.emit("📂 Leyendo archivo DESTINO...", "info")
            wb_destino = load_workbook(self.archivo_destino)
            ws_destino = wb_destino.active
            
            # Leer también con pandas para búsqueda
            df_destino = pd.read_excel(self.archivo_destino, skiprows=self.skiprows_destino, dtype=str)
            self.log_signal.emit(f"✅ DESTINO: {len(df_destino)} registros", "success")
            
            # Crear diccionarios para búsqueda rápida
            polizas_base = df_base[self.col_poliza_base].astype(str).str.strip().str.upper()
            df_base['_poliza_norm'] = polizas_base
            
            copiados = 0
            no_encontrados = []
            
            self.log_signal.emit("", "info")
            self.log_signal.emit("🔄 Copiando datos...", "info")
            self.log_signal.emit("="*60, "info")
            
            # Recorrer archivo DESTINO
            for idx_destino in range(self.skiprows_destino + 1, ws_destino.max_row + 1):
                # Obtener número de póliza del DESTINO
                col_letra_destino = self.obtener_letra_columna(self.col_poliza_destino)
                poliza_destino = ws_destino[f'{col_letra_destino}{idx_destino}'].value
                
                if not poliza_destino:
                    continue
                
                poliza_norm = str(poliza_destino).strip().upper()
                
                # Buscar en BASE
                mascara = df_base['_poliza_norm'] == poliza_norm
                
                if not mascara.any():
                    no_encontrados.append(poliza_destino)
                    if len(no_encontrados) <= 10:
                        self.log_signal.emit(f"⚠️ Póliza {poliza_destino}: No encontrada en BASE", "warning")
                    continue
                
                # Obtener fila de BASE
                fila_base = df_base[mascara].iloc[0]
                
                # Copiar datos según mapeos
                for col_base_idx, col_destino_idx in self.mapeos:
                    valor = fila_base.iloc[col_base_idx]
                    col_letra = self.obtener_letra_columna(col_destino_idx)
                    ws_destino[f'{col_letra}{idx_destino}'].value = valor
                
                copiados += 1
                if copiados <= 10 or copiados % 50 == 0:
                    self.log_signal.emit(f"✅ [{copiados}] Póliza {poliza_destino}: Datos copiados", "success")
            
            # Guardar
            self.log_signal.emit("", "info")
            self.log_signal.emit("💾 Guardando archivo...", "info")
            wb_destino.save(self.archivo_salida)
            self.log_signal.emit(f"✅ Guardado: {os.path.basename(self.archivo_salida)}", "success")
            
            estadisticas = {
                'total': ws_destino.max_row - self.skiprows_destino,
                'copiados': copiados,
                'no_encontrados': len(no_encontrados)
            }
            
            if no_encontrados:
                if len(no_encontrados) > 10:
                    self.log_signal.emit(f"⚠️ ... y {len(no_encontrados) - 10} pólizas más no encontradas", "warning")
            
            self.finished_signal.emit(True, "Copia completada", estadisticas)
            
        except Exception as e:
            self.log_signal.emit(f"❌ Error: {str(e)}", "error")
            self.finished_signal.emit(False, str(e), {})
    
    def obtener_letra_columna(self, indice):
        """Convierte índice numérico a letra de columna Excel"""
        letra = ""
        indice += 1  # Excel es 1-based
        while indice > 0:
            indice -= 1
            letra = chr(indice % 26 + 65) + letra
            indice //= 26
        return letra


# ==================================================================================
# Pestaña: Copiar Datos Entre Archivos
# ==================================================================================

class TabCopiarDatos(QWidget):
    """Pestaña para copiar datos entre archivos Excel"""
    
    def __init__(self):
        super().__init__()
        self.archivo_base = None
        self.archivo_destino = None
        self.columnas_base = []
        self.columnas_destino = []
        self.mapeos_widgets = []  # [(checkbox, combo_destino), ...]
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        
        # Título
        titulo = QLabel("📋 Copiar Datos Entre Archivos Excel")
        titulo.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(titulo)
        
        # Grupo: Archivo BASE
        grupo_base = QGroupBox("📂 Archivo BASE (Fuente de datos)")
        layout_base = QVBoxLayout(grupo_base)
        
        layout_base_file = QHBoxLayout()
        self.lbl_base = QLabel("No seleccionado")
        self.lbl_base.setObjectName("archivo")
        btn_base = QPushButton("📁 Seleccionar Archivo BASE")
        btn_base.clicked.connect(self.seleccionar_base)
        layout_base_file.addWidget(self.lbl_base, 1)
        layout_base_file.addWidget(btn_base)
        layout_base.addLayout(layout_base_file)
        
        layout_base_config = QHBoxLayout()
        lbl_skiprows_base = QLabel("Filas a saltar:")
        self.spin_skiprows_base = QSpinBox()
        self.spin_skiprows_base.setRange(0, 100)
        self.spin_skiprows_base.setValue(3)
        
        lbl_poliza_base = QLabel("Columna Póliza:")
        self.combo_poliza_base = QComboBox()
        self.combo_poliza_base.setEnabled(False)
        
        layout_base_config.addWidget(lbl_skiprows_base)
        layout_base_config.addWidget(self.spin_skiprows_base)
        layout_base_config.addWidget(lbl_poliza_base)
        layout_base_config.addWidget(self.combo_poliza_base, 1)
        layout_base.addLayout(layout_base_config)
        
        layout.addWidget(grupo_base)
        
        # Grupo: Archivo DESTINO
        grupo_destino = QGroupBox("📂 Archivo DESTINO (Donde se copiarán los datos)")
        layout_destino = QVBoxLayout(grupo_destino)
        
        layout_destino_file = QHBoxLayout()
        self.lbl_destino = QLabel("No seleccionado")
        self.lbl_destino.setObjectName("archivo")
        btn_destino = QPushButton("📁 Seleccionar Archivo DESTINO")
        btn_destino.clicked.connect(self.seleccionar_destino)
        layout_destino_file.addWidget(self.lbl_destino, 1)
        layout_destino_file.addWidget(btn_destino)
        layout_destino.addLayout(layout_destino_file)
        
        layout_destino_config = QHBoxLayout()
        lbl_skiprows_destino = QLabel("Filas a saltar:")
        self.spin_skiprows_destino = QSpinBox()
        self.spin_skiprows_destino.setRange(0, 100)
        self.spin_skiprows_destino.setValue(0)
        
        lbl_poliza_destino = QLabel("Columna Póliza:")
        self.combo_poliza_destino = QComboBox()
        self.combo_poliza_destino.setEnabled(False)
        
        layout_destino_config.addWidget(lbl_skiprows_destino)
        layout_destino_config.addWidget(self.spin_skiprows_destino)
        layout_destino_config.addWidget(lbl_poliza_destino)
        layout_destino_config.addWidget(self.combo_poliza_destino, 1)
        layout_destino.addLayout(layout_destino_config)
        
        layout.addWidget(grupo_destino)
        
        # Grupo: Mapeo de Columnas
        grupo_mapeo = QGroupBox("🔗 Mapeo de Columnas (Selecciona qué copiar)")
        layout_mapeo = QVBoxLayout(grupo_mapeo)
        
        # Scroll area para mapeos
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(200)
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        scroll.setWidget(self.scroll_widget)
        layout_mapeo.addWidget(scroll)
        
        self.btn_generar_mapeos = QPushButton("⚙️ Generar Lista de Mapeos")
        self.btn_generar_mapeos.clicked.connect(self.generar_mapeos)
        self.btn_generar_mapeos.setEnabled(False)
        layout_mapeo.addWidget(self.btn_generar_mapeos)
        
        layout.addWidget(grupo_mapeo)
        
        # Botón Copiar
        layout_btn = QHBoxLayout()
        layout_btn.addStretch()
        self.btn_copiar = QPushButton("🚀 COPIAR DATOS")
        self.btn_copiar.setObjectName("btnComparar")
        self.btn_copiar.clicked.connect(self.copiar_datos)
        self.btn_copiar.setEnabled(False)
        self.btn_copiar.setMinimumHeight(45)
        layout_btn.addWidget(self.btn_copiar)
        layout_btn.addStretch()
        layout.addLayout(layout_btn)
        
        # Grupo: Estadísticas
        grupo_stats = QGroupBox("📊 Estadísticas")
        layout_stats = QHBoxLayout(grupo_stats)
        
        self.stats_widgets = {}
        stats_config = [
            ('total', 'Total', '#4ec9b0'),
            ('copiados', 'Copiados', '#6a9955'),
            ('no_encontrados', 'No Encontrados', '#f44747')
        ]
        
        for key, label, color in stats_config:
            container = QWidget()
            container_layout = QVBoxLayout(container)
            container_layout.setContentsMargins(10, 5, 10, 5)
            
            lbl = QLabel(label)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet(f"color: {color}; font-weight: bold;")
            
            valor = QLabel("0")
            valor.setAlignment(Qt.AlignmentFlag.AlignCenter)
            valor.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: bold;")
            self.stats_widgets[key] = valor
            
            container_layout.addWidget(lbl)
            container_layout.addWidget(valor)
            layout_stats.addWidget(container)
        
        layout.addWidget(grupo_stats)
        
        # Grupo: Logs
        grupo_logs = QGroupBox("📝 Logs")
        layout_logs = QVBoxLayout(grupo_logs)
        
        self.txt_logs = QTextEdit()
        self.txt_logs.setReadOnly(True)
        self.txt_logs.setMinimumHeight(150)
        layout_logs.addWidget(self.txt_logs)
        
        layout_btns_log = QHBoxLayout()
        btn_limpiar = QPushButton("🗑️ Limpiar")
        btn_limpiar.clicked.connect(self.limpiar_logs)
        
        self.btn_abrir = QPushButton("📂 Abrir Archivo")
        self.btn_abrir.setObjectName("btnAbrir")
        self.btn_abrir.clicked.connect(self.abrir_archivo)
        self.btn_abrir.setEnabled(False)
        
        layout_btns_log.addWidget(btn_limpiar)
        layout_btns_log.addStretch()
        layout_btns_log.addWidget(self.btn_abrir)
        layout_logs.addLayout(layout_btns_log)
        
        layout.addWidget(grupo_logs)
    
    def seleccionar_base(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Archivo BASE",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            try:
                self.archivo_base = archivo
                self.lbl_base.setText(archivo)
                
                # Leer columnas
                skiprows = self.spin_skiprows_base.value()
                df = pd.read_excel(archivo, skiprows=skiprows, nrows=0)
                self.columnas_base = list(df.columns)
                
                self.combo_poliza_base.clear()
                self.combo_poliza_base.addItems(self.columnas_base)
                self.combo_poliza_base.setEnabled(True)
                
                self.log(f"✅ BASE cargado: {len(self.columnas_base)} columnas", "success")
                self.verificar_archivos()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al leer archivo:\n{str(e)}")
                self.archivo_base = None
    
    def seleccionar_destino(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Archivo DESTINO",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            try:
                self.archivo_destino = archivo
                self.lbl_destino.setText(archivo)
                
                # Leer columnas
                skiprows = self.spin_skiprows_destino.value()
                df = pd.read_excel(archivo, skiprows=skiprows, nrows=0)
                self.columnas_destino = list(df.columns)
                
                self.combo_poliza_destino.clear()
                self.combo_poliza_destino.addItems(self.columnas_destino)
                self.combo_poliza_destino.setEnabled(True)
                
                self.log(f"✅ DESTINO cargado: {len(self.columnas_destino)} columnas", "success")
                self.verificar_archivos()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al leer archivo:\n{str(e)}")
                self.archivo_destino = None
    
    def verificar_archivos(self):
        if self.archivo_base and self.archivo_destino:
            self.btn_generar_mapeos.setEnabled(True)
    
    def generar_mapeos(self):
        """Genera la lista de checkboxes y dropdowns para mapeo"""
        # Limpiar mapeos anteriores
        for i in reversed(range(self.scroll_layout.count())):
            self.scroll_layout.itemAt(i).widget().setParent(None)
        
        self.mapeos_widgets = []
        
        # Crear un mapeo por cada columna BASE
        for idx, col_base in enumerate(self.columnas_base):
            layout_fila = QHBoxLayout()
            
            # Checkbox
            checkbox = QCheckBox(f"{col_base}")
            checkbox.setChecked(False)
            
            # Label flecha
            lbl_flecha = QLabel("→")
            lbl_flecha.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            # ComboBox destino
            combo_destino = QComboBox()
            combo_destino.addItems(self.columnas_destino)
            combo_destino.setEnabled(False)
            
            # Conectar checkbox
            checkbox.stateChanged.connect(lambda state, c=combo_destino: c.setEnabled(state == 2))
            
            layout_fila.addWidget(checkbox, 2)
            layout_fila.addWidget(lbl_flecha)
            layout_fila.addWidget(combo_destino, 2)
            
            widget_fila = QWidget()
            widget_fila.setLayout(layout_fila)
            self.scroll_layout.addWidget(widget_fila)
            
            self.mapeos_widgets.append((checkbox, combo_destino, idx))
        
        self.scroll_layout.addStretch()
        self.btn_copiar.setEnabled(True)
        self.log("✅ Lista de mapeos generada. Selecciona las columnas a copiar", "success")
    
    def copiar_datos(self):
        # Obtener mapeos seleccionados
        mapeos = []
        for checkbox, combo, idx_base in self.mapeos_widgets:
            if checkbox.isChecked():
                col_destino_nombre = combo.currentText()
                idx_destino = self.columnas_destino.index(col_destino_nombre)
                mapeos.append((idx_base, idx_destino))
        
        if not mapeos:
            QMessageBox.warning(self, "Advertencia", "Debes seleccionar al menos una columna para copiar")
            return
        
        # Generar archivo de salida
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        carpeta_salida = os.path.join(os.path.dirname(self.archivo_destino), "output")
        os.makedirs(carpeta_salida, exist_ok=True)
        archivo_salida = os.path.join(carpeta_salida, f"DESTINO_con_datos_copiados_{timestamp}.xlsx")
        
        self.btn_copiar.setEnabled(False)
        self.limpiar_logs()
        
        self.log("", "info")
        self.log("="*60, "info")
        self.log("🚀 INICIANDO COPIA DE DATOS", "info")
        self.log(f"📊 Columnas a copiar: {len(mapeos)}", "info")
        self.log("="*60, "info")
        
        # Iniciar thread
        self.worker = CopiadorDatosThread(
            self.archivo_base,
            self.spin_skiprows_base.value(),
            self.combo_poliza_base.currentText(),
            self.archivo_destino,
            self.spin_skiprows_destino.value(),
            self.combo_poliza_destino.currentText(),
            mapeos,
            archivo_salida
        )
        self.archivo_salida = archivo_salida
        self.worker.log_signal.connect(self.log)
        self.worker.finished_signal.connect(self.copia_completada)
        self.worker.start()
    
    def copia_completada(self, exito, mensaje, estadisticas):
        if exito:
            self.log("", "info")
            self.log("="*60, "success")
            self.log("✅ COPIA COMPLETADA", "success")
            self.log("="*60, "success")
            
            for key, valor in estadisticas.items():
                if key in self.stats_widgets:
                    self.stats_widgets[key].setText(str(valor))
            
            self.btn_abrir.setEnabled(True)
        else:
            self.log(f"❌ Error: {mensaje}", "error")
        
        self.btn_copiar.setEnabled(True)
    
    def abrir_archivo(self):
        if hasattr(self, 'archivo_salida') and os.path.exists(self.archivo_salida):
            if os.name == 'nt':
                os.startfile(self.archivo_salida)
            self.log(f"📂 Abriendo: {os.path.basename(self.archivo_salida)}", "info")
    
    def limpiar_logs(self):
        self.txt_logs.clear()
    
    def log(self, mensaje: str, tipo: str = "info"):
        colores = {
            "info": "#d4d4d4",
            "success": "#6a9955",
            "warning": "#dcdcaa",
            "error": "#f44747"
        }
        color = colores.get(tipo, "#d4d4d4")
        self.txt_logs.append(f'<span style="color: {color};">{mensaje}</span>')


class ComparadorGUI(QMainWindow):
    """Ventana principal con pestañas"""
    
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("🔍 Comparador de Archivos - SoftSeguros")
        self.setMinimumSize(950, 750)
        self.setStyleSheet(DARK_STYLE)
        
        # Widget central
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Crear pestañas
        self.tabs = QTabWidget()
        
        # Pestaña 1: Maviso vs Maviso
        self.tab_maviso = TabMavisoVsMaviso()
        self.tabs.addTab(self.tab_maviso, "📊 Maviso Manual vs Generado")
        
        # Pestaña 2: Maviso vs CELER
        self.tab_celer = TabMavisoVsCeler()
        self.tabs.addTab(self.tab_celer, "🔄 Maviso vs CELER")
        
        # Pestaña 3: Validar Mensuales sin Prima
        self.tab_validar = TabValidarMensuales()
        self.tabs.addTab(self.tab_validar, "⚠️ Validar Mensuales sin Prima")
        
        # Pestaña 4: Copiar Datos
        self.tab_copiar = TabCopiarDatos()
        self.tabs.addTab(self.tab_copiar, "📋 Copiar Datos")
        
        layout.addWidget(self.tabs)


# ==================================================================================
# Thread para Validación de Mensuales
# ==================================================================================

class ValidadorThread(QThread):
    """Thread para ejecutar validación sin bloquear la GUI"""
    log = pyqtSignal(str)
    terminado = pyqtSignal(bool, str, dict)
    
    def __init__(self, ruta_archivo):
        super().__init__()
        self.ruta_archivo = ruta_archivo
    
    def run(self):
        try:
            # Configuración de columnas
            COL_POLIZA = 0
            COL_PRIMA = 14
            COL_MODALIDAD = 21
            
            self.log.emit("📂 Cargando archivo...")
            df = pd.read_excel(self.ruta_archivo)
            self.log.emit(f"   Total registros: {len(df)}")
            
            # Filtrar MENSUALES
            self.log.emit("\n🔍 Filtrando pólizas MENSUALES (Col 21)...")
            mask_mensual = df.iloc[:, COL_MODALIDAD].astype(str).str.strip().str.upper() == 'MENSUAL'
            df_mensuales = df[mask_mensual]
            self.log.emit(f"   Pólizas MENSUALES: {len(df_mensuales)}")
            
            # Filtrar prima = 0
            self.log.emit("\n🔍 Filtrando pólizas con Prima = 0 (Col 14)...")
            mask_prima_cero = df_mensuales.iloc[:, COL_PRIMA].fillna(0) == 0
            df_problema = df_mensuales[mask_prima_cero]
            self.log.emit(f"   MENSUALES con Prima = 0: {len(df_problema)}")
            
            # Preparar estadísticas
            stats = {
                'total': len(df),
                'mensuales': len(df_mensuales),
                'problematicas': len(df_problema),
                'porcentaje': len(df_problema)/len(df_mensuales)*100 if len(df_mensuales) > 0 else 0,
                'df_problema': df_problema
            }
            
            self.log.emit("\n✅ Validación completada")
            self.terminado.emit(True, "Validación exitosa", stats)
            
        except Exception as e:
            self.log.emit(f"\n❌ Error: {str(e)}")
            self.terminado.emit(False, f"Error: {str(e)}", {})


# ==================================================================================
# Pestaña: Validar Mensuales sin Prima
# ==================================================================================

class TabValidarMensuales(QWidget):
    """Pestaña para validar pólizas mensuales sin prima"""
    
    def __init__(self):
        super().__init__()
        self.ruta_archivo = None
        self.validador_thread = None
        self.df_resultados = None
        self.inicializar_ui()
    
    def inicializar_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Título
        titulo = QLabel("⚠️ Validación de Pólizas MENSUALES sin Prima")
        titulo.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(titulo)
        
        # Grupo: Selección de archivo
        grupo_archivo = QGroupBox("📁 Archivo a Validar")
        grupo_archivo.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        layout_archivo = QVBoxLayout()
        
        layout_btn_archivo = QHBoxLayout()
        self.btn_seleccionar = QPushButton("📂 Seleccionar Archivo MAVISO")
        self.btn_seleccionar.clicked.connect(self.seleccionar_archivo)
        self.btn_seleccionar.setMinimumHeight(40)
        layout_btn_archivo.addWidget(self.btn_seleccionar)
        
        self.lbl_archivo = QLabel("Ningún archivo seleccionado")
        self.lbl_archivo.setWordWrap(True)
        self.lbl_archivo.setStyleSheet("padding: 8px; background-color: #2d2d2d; border-radius: 4px;")
        
        layout_archivo.addLayout(layout_btn_archivo)
        layout_archivo.addWidget(self.lbl_archivo)
        grupo_archivo.setLayout(layout_archivo)
        layout.addWidget(grupo_archivo)
        
        # Grupo: Estadísticas
        grupo_stats = QGroupBox("📊 Estadísticas")
        grupo_stats.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        layout_stats = QVBoxLayout()
        
        self.lbl_total = QLabel("Total pólizas: -")
        self.lbl_mensuales = QLabel("Pólizas MENSUALES: -")
        self.lbl_problematicas = QLabel("MENSUALES sin prima: -")
        self.lbl_porcentaje = QLabel("Porcentaje: -")
        
        for lbl in [self.lbl_total, self.lbl_mensuales, self.lbl_problematicas, self.lbl_porcentaje]:
            lbl.setFont(QFont("Segoe UI", 10))
            layout_stats.addWidget(lbl)
        
        grupo_stats.setLayout(layout_stats)
        layout.addWidget(grupo_stats)
        
        # Botones de acción
        layout_botones = QHBoxLayout()
        
        self.btn_validar = QPushButton("🔍 Validar")
        self.btn_validar.setEnabled(False)
        self.btn_validar.clicked.connect(self.validar)
        self.btn_validar.setMinimumHeight(45)
        self.btn_validar.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        
        self.btn_exportar = QPushButton("💾 Exportar Resultados")
        self.btn_exportar.setEnabled(False)
        self.btn_exportar.clicked.connect(self.exportar_resultados)
        self.btn_exportar.setMinimumHeight(45)
        
        layout_botones.addWidget(self.btn_validar, 2)
        layout_botones.addWidget(self.btn_exportar, 1)
        layout.addLayout(layout_botones)
        
        # Log
        grupo_log = QGroupBox("📋 Registro de Actividad")
        grupo_log.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        layout_log = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        self.log_text.setMinimumHeight(200)
        
        layout_log.addWidget(self.log_text)
        grupo_log.setLayout(layout_log)
        layout.addWidget(grupo_log)
        
        # Barra de progreso
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setTextVisible(False)
        layout.addWidget(self.progress)
    
    def seleccionar_archivo(self):
        """Selecciona el archivo a validar"""
        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo MAVISO",
            "../output",
            "Archivos Excel (*.xlsx *.xls)"
        )
        
        if ruta:
            self.ruta_archivo = ruta
            nombre = Path(ruta).name
            self.lbl_archivo.setText(f"📄 {nombre}")
            self.btn_validar.setEnabled(True)
            self.agregar_log(f"Archivo seleccionado: {nombre}")
    
    def validar(self):
        """Ejecuta la validación"""
        if not self.ruta_archivo:
            QMessageBox.warning(self, "Advertencia", "Debe seleccionar un archivo primero")
            return
        
        self.btn_validar.setEnabled(False)
        self.btn_exportar.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)  # Modo indeterminado
        self.log_text.clear()
        
        # Limpiar estadísticas
        self.lbl_total.setText("Total pólizas: Validando...")
        self.lbl_mensuales.setText("Pólizas MENSUALES: Validando...")
        self.lbl_problematicas.setText("MENSUALES sin prima: Validando...")
        self.lbl_porcentaje.setText("Porcentaje: Validando...")
        
        # Iniciar thread
        self.validador_thread = ValidadorThread(self.ruta_archivo)
        self.validador_thread.log.connect(self.agregar_log)
        self.validador_thread.terminado.connect(self.validacion_terminada)
        self.validador_thread.start()
    
    def validacion_terminada(self, exito, mensaje, stats):
        """Callback cuando termina la validación"""
        self.progress.setVisible(False)
        self.btn_validar.setEnabled(True)
        
        if exito:
            # Actualizar estadísticas
            self.lbl_total.setText(f"Total pólizas: {stats['total']:,}")
            self.lbl_mensuales.setText(f"Pólizas MENSUALES: {stats['mensuales']:,}")
            self.lbl_problematicas.setText(f"MENSUALES sin prima: {stats['problematicas']:,}")
            self.lbl_porcentaje.setText(f"Porcentaje: {stats['porcentaje']:.1f}%")
            
            # Guardar resultados
            self.df_resultados = stats.get('df_problema')
            self.btn_exportar.setEnabled(len(self.df_resultados) > 0 if self.df_resultados is not None else False)
            
            if stats['problematicas'] > 0:
                self.agregar_log(f"\n⚠️ Se encontraron {stats['problematicas']} pólizas problemáticas")
            else:
                self.agregar_log("\n✅ No se encontraron problemas")
        else:
            QMessageBox.critical(self, "Error", mensaje)
    
    def exportar_resultados(self):
        """Exporta los resultados a Excel"""
        if self.df_resultados is None or len(self.df_resultados) == 0:
            QMessageBox.warning(self, "Advertencia", "No hay resultados para exportar")
            return
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archivo_salida = Path('../output') / f'mensuales_sin_prima_{timestamp}.xlsx'
            
            # Preparar datos
            resultado = pd.DataFrame({
                'Póliza': self.df_resultados.iloc[:, 0],
                'Prima Neta': self.df_resultados.iloc[:, 14],
                'Modalidad': self.df_resultados.iloc[:, 21],
                'Fila Excel': self.df_resultados.index + 2,
                'Placa': self.df_resultados.iloc[:, 1],
                'Asegurado': self.df_resultados.iloc[:, 2],
                'Fecha Inicio': self.df_resultados.iloc[:, 10],
                'Fecha Fin': self.df_resultados.iloc[:, 11],
            })
            
            resultado.to_excel(archivo_salida, index=False)
            self.agregar_log(f"\n💾 Reporte exportado: {archivo_salida.name}")
            
            QMessageBox.information(
                self,
                "Éxito",
                f"Reporte exportado correctamente:\n{archivo_salida}"
            )
            
            # Abrir archivo
            if os.name == 'nt':
                os.startfile(archivo_salida)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar:\n{str(e)}")
    
    def agregar_log(self, mensaje):
        """Agrega mensaje al log"""
        self.log_text.append(mensaje)
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )


# ==================================================================================
# Ventana Principal
# ==================================================================================

class ComparadorGUI(QMainWindow):
    """Ventana principal con pestañas"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Comparador de Archivos - SoftSeguros")
        self.setMinimumSize(1000, 800)
        
        # Aplicar estilos
        self.setStyleSheet(DARK_STYLE)
        
        # Widget central
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Crear pestañas
        self.tabs = QTabWidget()
        
        # Pestaña 1: Maviso vs Maviso
        self.tab_maviso = TabMavisoVsMaviso()
        self.tabs.addTab(self.tab_maviso, "📊 Maviso Manual vs Generado")
        
        # Pestaña 2: Maviso vs CELER
        self.tab_celer = TabMavisoVsCeler()
        self.tabs.addTab(self.tab_celer, "🔄 Maviso vs CELER")
        
        # Pestaña 3: Validar Mensuales sin Prima
        self.tab_validar = TabValidarMensuales()
        self.tabs.addTab(self.tab_validar, "⚠️ Validar Mensuales sin Prima")
        
        layout.addWidget(self.tabs)


def main():
    app = QApplication(sys.argv)
    
    # Configurar fuente por defecto
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    
    ventana = ComparadorGUI()
    ventana.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
