"""
Comparador Maviso vs CELER
==========================
Compara el archivo Maviso (manual) contra el archivo fuente CELER
para validar que los datos originales estén correctos.

Campos a comparar:
- Clave: Número de póliza
- Prima sin IVA
- Fecha inicio
- Fecha fin
- Modalidad (Mensual/Anual)
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Tuple, Dict, List
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================
# COLUMNAS CELER (con skiprows=3)
# ============================================
CELER_COL_POLIZA = 20       # Póliza
CELER_COL_PRIMA = 42        # Prima sin IVA
CELER_COL_FECHA_INICIO = 22 # F_Inicio
CELER_COL_FECHA_FIN = 23    # F_Fin
CELER_COL_MODALIDAD = 26    # Modalidad

# ============================================
# COLUMNAS MAVISO
# ============================================
MAVISO_COL_POLIZA = 0       # NÚMERO DE PÓLIZA
MAVISO_COL_PRIMA = 14       # PRIMA NETA
MAVISO_COL_FECHA_INICIO = 10 # FECHA INICIO
MAVISO_COL_FECHA_FIN = 11    # FECHA FIN
MAVISO_COL_FORMA_PAGO = 22   # FORMA PAGO

# Estilos para Excel
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ComparadorMavisoCeler:
    """Clase para comparar Maviso manual contra archivo fuente CELER"""
    
    def __init__(self):
        self.df_maviso = None
        self.df_celer = None
        self.resultados = None
        self.estadisticas = {}
    
    def cargar_maviso(self, ruta: str) -> Tuple[bool, str]:
        """Carga el archivo Maviso (modificado manualmente)"""
        try:
            ruta_path = Path(ruta)
            if not ruta_path.exists():
                return False, f"Archivo no encontrado: {ruta}"
            
            self.df_maviso = pd.read_excel(ruta, header=0)
            
            if len(self.df_maviso.columns) < 23:
                return False, "El archivo Maviso no tiene suficientes columnas"
            
            logger.info(f"✅ Maviso cargado: {len(self.df_maviso)} filas, {len(self.df_maviso.columns)} columnas")
            return True, f"Maviso cargado: {len(self.df_maviso)} registros"
            
        except Exception as e:
            return False, f"Error cargando Maviso: {str(e)}"
    
    def cargar_celer(self, ruta: str) -> Tuple[bool, str]:
        """Carga el archivo CELER (fuente original)"""
        try:
            ruta_path = Path(ruta)
            if not ruta_path.exists():
                return False, f"Archivo no encontrado: {ruta}"
            
            # CELER tiene 3 filas de encabezado
            self.df_celer = pd.read_excel(ruta, skiprows=3, header=0)
            
            if len(self.df_celer.columns) < 43:
                return False, "El archivo CELER no tiene suficientes columnas"
            
            logger.info(f"✅ CELER cargado: {len(self.df_celer)} filas, {len(self.df_celer.columns)} columnas")
            return True, f"CELER cargado: {len(self.df_celer)} registros"
            
        except Exception as e:
            return False, f"Error cargando CELER: {str(e)}"
    
    def _normalizar_valor(self, valor):
        """Normaliza un valor para comparación"""
        if pd.isna(valor):
            return None
        
        if isinstance(valor, (datetime, pd.Timestamp)):
            return valor.strftime('%Y-%m-%d')
        
        if isinstance(valor, (int, float)):
            return round(float(valor), 2)
        
        return str(valor).strip().upper()
    
    def _normalizar_poliza(self, valor):
        """Normaliza número de póliza para comparación"""
        if pd.isna(valor):
            return None
        return str(valor).strip()
    
    def _comparar_valores(self, val1, val2) -> bool:
        """Compara dos valores normalizados"""
        norm1 = self._normalizar_valor(val1)
        norm2 = self._normalizar_valor(val2)
        
        if norm1 is None and norm2 is None:
            return True
        if norm1 is None or norm2 is None:
            return False
        return norm1 == norm2
    
    def _comparar_prima(self, prima_maviso, prima_celer) -> bool:
        """
        Compara primas considerando que:
        - Los valores pueden ser negativos o positivos
        - Lo importante es que el valor absoluto sea equivalente
        - Se redondea a 2 decimales
        """
        if pd.isna(prima_maviso) and pd.isna(prima_celer):
            return True
        if pd.isna(prima_maviso) or pd.isna(prima_celer):
            return False
        
        try:
            # Convertir a float y obtener valor absoluto
            val_maviso = abs(round(float(prima_maviso), 2))
            val_celer = abs(round(float(prima_celer), 2))
            
            return val_maviso == val_celer
        except (ValueError, TypeError):
            return False
    
    def comparar(self) -> Tuple[bool, str, Dict]:
        """Ejecuta la comparación entre Maviso y CELER"""
        if self.df_maviso is None:
            return False, "No se ha cargado el archivo Maviso", None
        
        if self.df_celer is None:
            return False, "No se ha cargado el archivo CELER", None
        
        logger.info("=" * 60)
        logger.info("COMPARANDO MAVISO vs CELER")
        logger.info("=" * 60)
        
        # Crear diccionario de CELER por número de póliza
        polizas_celer = {}
        for idx, row in self.df_celer.iterrows():
            poliza = self._normalizar_poliza(row.iloc[CELER_COL_POLIZA])
            if poliza:
                polizas_celer[poliza] = {
                    'fila': idx + 5,  # +5 por las filas de encabezado (skiprows=3 + header + 0-index)
                    'prima': row.iloc[CELER_COL_PRIMA],
                    'fecha_inicio': row.iloc[CELER_COL_FECHA_INICIO],
                    'fecha_fin': row.iloc[CELER_COL_FECHA_FIN],
                    'modalidad': row.iloc[CELER_COL_MODALIDAD]
                }
        
        # Crear diccionario de MAVISO por número de póliza
        polizas_maviso = {}
        for idx, row in self.df_maviso.iterrows():
            poliza = self._normalizar_poliza(row.iloc[MAVISO_COL_POLIZA])
            if poliza:
                polizas_maviso[poliza] = {
                    'fila': idx + 2,
                    'prima': row.iloc[MAVISO_COL_PRIMA],
                    'fecha_inicio': row.iloc[MAVISO_COL_FECHA_INICIO],
                    'fecha_fin': row.iloc[MAVISO_COL_FECHA_FIN],
                    'forma_pago': row.iloc[MAVISO_COL_FORMA_PAGO]
                }
        
        # Estadísticas
        total_maviso = len(polizas_maviso)
        total_celer = len(polizas_celer)
        
        # Comparar
        coincidencias = []
        discrepancias = []
        solo_en_maviso = []
        solo_en_celer = []
        
        # Buscar pólizas de Maviso en CELER
        for poliza, datos_maviso in polizas_maviso.items():
            if poliza in polizas_celer:
                datos_celer = polizas_celer[poliza]
                errores = []
                
                # Prima (compara valor absoluto)
                if not self._comparar_prima(datos_maviso['prima'], datos_celer['prima']):
                    errores.append({
                        'campo': 'Prima',
                        'maviso': self._normalizar_valor(datos_maviso['prima']),
                        'celer': self._normalizar_valor(datos_celer['prima'])
                    })
                
                # Fecha inicio
                if not self._comparar_valores(datos_maviso['fecha_inicio'], datos_celer['fecha_inicio']):
                    errores.append({
                        'campo': 'Fecha Inicio',
                        'maviso': self._normalizar_valor(datos_maviso['fecha_inicio']),
                        'celer': self._normalizar_valor(datos_celer['fecha_inicio'])
                    })
                
                # Fecha fin
                if not self._comparar_valores(datos_maviso['fecha_fin'], datos_celer['fecha_fin']):
                    errores.append({
                        'campo': 'Fecha Fin',
                        'maviso': self._normalizar_valor(datos_maviso['fecha_fin']),
                        'celer': self._normalizar_valor(datos_celer['fecha_fin'])
                    })
                
                # NOTA: Modalidad/Forma de pago NO se compara
                
                if errores:
                    discrepancias.append({
                        'poliza': poliza,
                        'fila_maviso': datos_maviso['fila'],
                        'fila_celer': datos_celer['fila'],
                        'errores': errores
                    })
                    logger.warning(f"❌ DISCREPANCIA - Póliza: {poliza}")
                    logger.warning(f"   Fila Maviso: {datos_maviso['fila']} | Fila CELER: {datos_celer['fila']}")
                    for err in errores:
                        logger.warning(f"   → {err['campo']}: Maviso='{err['maviso']}' vs CELER='{err['celer']}'")
                else:
                    coincidencias.append(poliza)
            else:
                solo_en_maviso.append({
                    'poliza': poliza,
                    'fila': datos_maviso['fila']
                })
                logger.info(f"⚠️ Solo en MAVISO - Póliza: {poliza} (Fila {datos_maviso['fila']})")
        
        # Buscar pólizas solo en CELER
        for poliza in polizas_celer:
            if poliza not in polizas_maviso:
                solo_en_celer.append({
                    'poliza': poliza,
                    'fila': polizas_celer[poliza]['fila']
                })
                logger.info(f"⚠️ Solo en CELER - Póliza: {poliza} (Fila {polizas_celer[poliza]['fila']})")
        
        # Guardar resultados
        self.resultados = {
            'coincidencias': coincidencias,
            'discrepancias': discrepancias,
            'solo_maviso': solo_en_maviso,
            'solo_celer': solo_en_celer
        }
        
        self.estadisticas = {
            'total_maviso': total_maviso,
            'total_celer': total_celer,
            'coincidencias': len(coincidencias),
            'discrepancias': len(discrepancias),
            'solo_maviso': len(solo_en_maviso),
            'solo_celer': len(solo_en_celer)
        }
        
        # Log resumen
        logger.info("=" * 60)
        logger.info("RESUMEN DE COMPARACIÓN MAVISO vs CELER")
        logger.info("=" * 60)
        logger.info(f"  Total pólizas Maviso: {total_maviso}")
        logger.info(f"  Total pólizas CELER: {total_celer}")
        logger.info(f"  ✅ Coincidencias exactas: {len(coincidencias)}")
        logger.info(f"  ❌ Discrepancias: {len(discrepancias)}")
        logger.info(f"  ⚠️ Solo en Maviso: {len(solo_en_maviso)}")
        logger.info(f"  ⚠️ Solo en CELER: {len(solo_en_celer)}")
        logger.info("=" * 60)
        
        return True, "Comparación completada", self.resultados
    
    def obtener_estadisticas(self) -> Dict:
        return self.estadisticas
    
    def obtener_discrepancias(self) -> List[Dict]:
        if self.resultados:
            return self.resultados['discrepancias']
        return []
    
    def _aplicar_formato_header(self, worksheet):
        """Aplica formato estándar a encabezados"""
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _aplicar_formato_datos(self, worksheet, fila_inicio=2):
        """Aplica bordes a todas las celdas de datos"""
        for row in worksheet.iter_rows(min_row=fila_inicio, max_row=worksheet.max_row):
            for cell in row:
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
    
    def exportar_reporte(self, ruta_salida: str = None) -> str:
        """Exporta reporte de comparación a Excel"""
        if not self.resultados:
            raise ValueError("No hay resultados. Ejecute comparar() primero.")
        
        if not ruta_salida:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            carpeta = Path(__file__).parent / 'output'
            carpeta.mkdir(exist_ok=True)
            ruta_salida = carpeta / f'comparacion_maviso_celer_{timestamp}.xlsx'
        
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            resumen_data = {
                'Métrica': [
                    'Total pólizas Maviso',
                    'Total pólizas CELER',
                    'Coincidencias exactas',
                    'Discrepancias',
                    'Solo en Maviso',
                    'Solo en CELER'
                ],
                'Valor': [
                    self.estadisticas['total_maviso'],
                    self.estadisticas['total_celer'],
                    self.estadisticas['coincidencias'],
                    self.estadisticas['discrepancias'],
                    self.estadisticas['solo_maviso'],
                    self.estadisticas['solo_celer']
                ],
                'Estado': [
                    '📊',
                    '📊',
                    '✅',
                    '❌' if self.estadisticas['discrepancias'] > 0 else '✅',
                    '⚠️' if self.estadisticas['solo_maviso'] > 0 else '✅',
                    '⚠️' if self.estadisticas['solo_celer'] > 0 else '✅'
                ]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            ws_resumen = writer.sheets['Resumen']
            self._aplicar_formato_header(ws_resumen)
            self._aplicar_formato_datos(ws_resumen)
            
            for row_idx in range(2, ws_resumen.max_row + 1):
                estado = ws_resumen.cell(row=row_idx, column=3).value
                if estado == '✅':
                    for col in range(1, 4):
                        ws_resumen.cell(row=row_idx, column=col).fill = SUCCESS_FILL
                elif estado == '❌':
                    for col in range(1, 4):
                        ws_resumen.cell(row=row_idx, column=col).fill = ERROR_FILL
                elif estado == '⚠️':
                    for col in range(1, 4):
                        ws_resumen.cell(row=row_idx, column=col).fill = WARNING_FILL
            
            # Hoja 2: Discrepancias
            if self.resultados['discrepancias']:
                discrepancias_flat = []
                for d in self.resultados['discrepancias']:
                    for error in d['errores']:
                        discrepancias_flat.append({
                            'Póliza': d['poliza'],
                            'Fila Maviso': d['fila_maviso'],
                            'Fila CELER': d['fila_celer'],
                            'Campo': error['campo'],
                            'Valor Maviso': error['maviso'],
                            'Valor CELER': error['celer'],
                            'Diferencia': self._calcular_diferencia(error['maviso'], error['celer'])
                        })
                df_disc = pd.DataFrame(discrepancias_flat)
                df_disc.to_excel(writer, sheet_name='Discrepancias', index=False)
                
                ws_disc = writer.sheets['Discrepancias']
                self._aplicar_formato_header(ws_disc)
                self._aplicar_formato_datos(ws_disc)
                
                for row in range(2, ws_disc.max_row + 1):
                    for col in range(1, 8):
                        ws_disc.cell(row=row, column=col).fill = ERROR_FILL
            
            # Hoja 3: Solo en Maviso
            if self.resultados['solo_maviso']:
                df_solo_maviso = pd.DataFrame(self.resultados['solo_maviso'])
                df_solo_maviso.columns = ['Póliza', 'Fila']
                df_solo_maviso['Estado'] = '⚠️ No existe en CELER'
                df_solo_maviso.to_excel(writer, sheet_name='Solo_en_Maviso', index=False)
                
                ws_maviso = writer.sheets['Solo_en_Maviso']
                self._aplicar_formato_header(ws_maviso)
                self._aplicar_formato_datos(ws_maviso)
                
                for row in range(2, ws_maviso.max_row + 1):
                    for col in range(1, 4):
                        ws_maviso.cell(row=row, column=col).fill = WARNING_FILL
            
            # Hoja 4: Solo en CELER
            if self.resultados['solo_celer']:
                df_solo_celer = pd.DataFrame(self.resultados['solo_celer'])
                df_solo_celer.columns = ['Póliza', 'Fila']
                df_solo_celer['Estado'] = '⚠️ No existe en Maviso'
                df_solo_celer.to_excel(writer, sheet_name='Solo_en_CELER', index=False)
                
                ws_celer = writer.sheets['Solo_en_CELER']
                self._aplicar_formato_header(ws_celer)
                self._aplicar_formato_datos(ws_celer)
                
                for row in range(2, ws_celer.max_row + 1):
                    for col in range(1, 4):
                        ws_celer.cell(row=row, column=col).fill = WARNING_FILL
            
            # Hoja 5: Coincidencias
            if self.resultados['coincidencias']:
                df_coincidencias = pd.DataFrame({
                    'Póliza': self.resultados['coincidencias'],
                    'Estado': ['✅ OK'] * len(self.resultados['coincidencias'])
                })
                df_coincidencias.to_excel(writer, sheet_name='Coincidencias', index=False)
                
                ws_coinc = writer.sheets['Coincidencias']
                self._aplicar_formato_header(ws_coinc)
                self._aplicar_formato_datos(ws_coinc)
                
                for row in range(2, ws_coinc.max_row + 1):
                    for col in range(1, 3):
                        ws_coinc.cell(row=row, column=col).fill = SUCCESS_FILL
        
        logger.info(f"📊 Reporte exportado: {ruta_salida}")
        
        # Abrir automáticamente
        os.startfile(ruta_salida)
        
        return str(ruta_salida)
    
    def _calcular_diferencia(self, val1, val2):
        """Calcula diferencia numérica si ambos son números"""
        try:
            if val1 is None or val2 is None:
                return "N/A"
            num1 = float(val1)
            num2 = float(val2)
            diff = num1 - num2
            if diff > 0:
                return f"+{diff:,.2f}"
            else:
                return f"{diff:,.2f}"
        except (ValueError, TypeError):
            return "N/A (texto/fecha)"


# Test rápido
if __name__ == "__main__":
    print("=" * 60)
    print("COMPARADOR MAVISO vs CELER")
    print("=" * 60)
    print("\nUso:")
    print("  from comparar_maviso_celer import ComparadorMavisoCeler")
    print("  comp = ComparadorMavisoCeler()")
    print("  comp.cargar_maviso('maviso.xlsx')")
    print("  comp.cargar_celer('celer.xlsx')")
    print("  comp.comparar()")
    print("  comp.exportar_reporte()")
    print("=" * 60)
