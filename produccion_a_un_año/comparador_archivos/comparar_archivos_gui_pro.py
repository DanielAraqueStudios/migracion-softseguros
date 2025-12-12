"""
Comparador de Archivos - Interfaz Profesional Dark Mode
=======================================================
GUI moderna con PyQt6 para comparar archivos Excel
- Dark Mode profesional
- UI/UX mejorada
- Animaciones sutiles
- Diseño moderno
"""

import sys
import subprocess
import os
import time
import re
import requests
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QFileDialog, QGroupBox,
    QProgressBar, QFrame, QMessageBox, QTabWidget, QScrollArea
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QPropertyAnimation, QEasingCurve, QTimer
from PyQt6.QtGui import QFont, QColor, QPalette, QIcon, QPainter, QLinearGradient
import pandas as pd

from comparar_archivos import ComparadorMaviso
from comparar_maviso_celer import ComparadorMavisoCeler


# ==================== ESTILOS DARK MODE ====================
DARK_STYLESHEET = """
QMainWindow {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                stop:0 #1a1a2e, stop:1 #16213e);
}

QWidget {
    background-color: transparent;
    color: #e4e4e7;
    font-family: 'Segoe UI', Arial;
    font-size: 9pt;
}

QTabWidget::pane {
    border: 2px solid #3f3f46;
    border-radius: 8px;
    background-color: #18181b;
}

QTabBar::tab {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #27272a, stop:1 #18181b);
    border: 1px solid #3f3f46;
    padding: 12px 24px;
    margin-right: 4px;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    color: #a1a1aa;
    font-weight: 500;
}

QTabBar::tab:selected {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #3b82f6, stop:1 #2563eb);
    color: white;
    border-bottom: 2px solid #3b82f6;
}

QTabBar::tab:hover:!selected {
    background: #27272a;
    color: #e4e4e7;
}

QGroupBox {
    background-color: rgba(39, 39, 42, 0.7);
    border: 2px solid #3f3f46;
    border-radius: 10px;
    margin-top: 12px;
    padding: 15px;
    font-weight: 600;
    color: #e4e4e7;
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 5px 15px;
    background-color: #3b82f6;
    border-radius: 5px;
    color: white;
}

QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #3b82f6, stop:1 #2563eb);
    color: white;
    border: none;
    border-radius: 8px;
    padding: 12px 20px;
    font-weight: 600;
    font-size: 9pt;
    min-height: 20px;
}

QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #60a5fa, stop:1 #3b82f6);
}

QPushButton:pressed {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #2563eb, stop:1 #1e40af);
}

QPushButton:disabled {
    background: #3f3f46;
    color: #71717a;
}

QPushButton#btnSuccess {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #10b981, stop:1 #059669);
}

QPushButton#btnSuccess:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #34d399, stop:1 #10b981);
}

QPushButton#btnWarning {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #f59e0b, stop:1 #d97706);
}

QPushButton#btnWarning:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #fbbf24, stop:1 #f59e0b);
}

QPushButton#btnDanger {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #ef4444, stop:1 #dc2626);
}

QPushButton#btnDanger:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #f87171, stop:1 #ef4444);
}

QLabel {
    color: #e4e4e7;
    background: transparent;
}

QLabel#titleLabel {
    font-size: 14pt;
    font-weight: 700;
    color: #3b82f6;
}

QLabel#subtitleLabel {
    font-size: 10pt;
    color: #a1a1aa;
}

QLabel#statValue {
    font-size: 20pt;
    font-weight: 700;
    color: #3b82f6;
}

QTextEdit {
    background-color: #09090b;
    border: 2px solid #27272a;
    border-radius: 8px;
    color: #e4e4e7;
    padding: 10px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 9pt;
    selection-background-color: #3b82f6;
}

QTextEdit:focus {
    border: 2px solid #3b82f6;
}

QProgressBar {
    background-color: #27272a;
    border: 2px solid #3f3f46;
    border-radius: 8px;
    text-align: center;
    color: white;
    font-weight: 600;
    height: 25px;
}

QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 #3b82f6, stop:0.5 #8b5cf6, stop:1 #ec4899);
    border-radius: 6px;
}

QScrollBar:vertical {
    background: #18181b;
    width: 14px;
    margin: 0px;
    border-radius: 7px;
}

QScrollBar::handle:vertical {
    background: #3f3f46;
    min-height: 30px;
    border-radius: 7px;
}

QScrollBar::handle:vertical:hover {
    background: #52525b;
}

QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}

QScrollBar:horizontal {
    background: #18181b;
    height: 14px;
    margin: 0px;
    border-radius: 7px;
}

QScrollBar::handle:horizontal {
    background: #3f3f46;
    min-width: 30px;
    border-radius: 7px;
}

QScrollBar::handle:horizontal:hover {
    background: #52525b;
}

QFrame {
    background: transparent;
}

QFrame#separator {
    background-color: #3f3f46;
    max-height: 2px;
}
"""


class ComparadorThread(QThread):
    """Thread para ejecutar comparación sin bloquear UI"""
    progreso = pyqtSignal(str)
    terminado = pyqtSignal(bool, str, dict)
    
    def __init__(self, comparador, archivo1, archivo2=None):
        super().__init__()
        self.comparador = comparador
        self.archivo1 = archivo1
        self.archivo2 = archivo2
    
    def run(self):
        try:
            self.progreso.emit("🔄 Iniciando comparación...")
            
            if self.archivo2:
                resultados = self.comparador.comparar(self.archivo1, self.archivo2)
            else:
                resultados = self.comparador.comparar(self.archivo1)
            
            estadisticas = self.comparador.obtener_estadisticas()
            self.terminado.emit(True, "Comparación completada", estadisticas)
            
        except Exception as e:
            self.terminado.emit(False, str(e), {})


class ValidadorThread(QThread):
    """Thread para ejecutar validación sin bloquear UI"""
    progreso = pyqtSignal(str)
    terminado = pyqtSignal(bool, str, dict)
    
    def __init__(self, archivo):
        super().__init__()
        self.archivo = archivo
    
    def run(self):
        try:
            self.progreso.emit("🔄 Validando mensuales sin prima...")
            
            # Implementación básica sin importar el validador
            df = pd.read_excel(self.archivo)
            mensuales_sin_prima = df[(df.iloc[:, 21].str.upper() == 'MENSUAL') & (df.iloc[:, 14] == 0)]
            
            estadisticas = {
                'mensuales_sin_prima': len(mensuales_sin_prima)
            }
            
            self.terminado.emit(True, "Validación completada", estadisticas)
            
        except Exception as e:
            self.terminado.emit(False, str(e), {})


class TabMavisoVsCeler(QWidget):
    """Pestaña 2: Comparación Maviso vs CELER"""
    
    def __init__(self):
        super().__init__()
        self.comparador = ComparadorMavisoCeler()
        self.ruta_maviso = None
        self.ruta_celer = None
        self.thread = None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Título
        title = QLabel("📊 Comparador MAVISO vs CELER")
        title.setObjectName("titleLabel")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        subtitle = QLabel("Valida datos del archivo MAVISO contra la fuente original CELER")
        subtitle.setObjectName("subtitleLabel")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(subtitle)
        
        layout.addSpacing(10)
        
        # Grupo de archivos
        grupo_archivos = QGroupBox("📁 Archivos")
        layout_archivos = QVBoxLayout()
        
        # MAVISO
        layout_maviso = QHBoxLayout()
        self.btn_maviso = QPushButton("📂 Seleccionar MAVISO")
        self.btn_maviso.clicked.connect(self.seleccionar_maviso)
        self.lbl_maviso = QLabel("No seleccionado")
        self.lbl_maviso.setStyleSheet("color: #71717a; font-style: italic;")
        layout_maviso.addWidget(self.btn_maviso)
        layout_maviso.addWidget(self.lbl_maviso, 1)
        layout_archivos.addLayout(layout_maviso)
        
        # CELER
        layout_celer = QHBoxLayout()
        self.btn_celer = QPushButton("📂 Seleccionar CELER")
        self.btn_celer.clicked.connect(self.seleccionar_celer)
        self.lbl_celer = QLabel("No seleccionado")
        self.lbl_celer.setStyleSheet("color: #71717a; font-style: italic;")
        layout_celer.addWidget(self.btn_celer)
        layout_celer.addWidget(self.lbl_celer, 1)
        layout_archivos.addLayout(layout_celer)
        
        grupo_archivos.setLayout(layout_archivos)
        layout.addWidget(grupo_archivos)
        
        # Botón comparar
        self.btn_comparar = QPushButton("🔍 INICIAR COMPARACIÓN")
        self.btn_comparar.setMinimumHeight(50)
        self.btn_comparar.setObjectName("btnSuccess")
        self.btn_comparar.clicked.connect(self.comparar)
        self.btn_comparar.setEnabled(False)
        layout.addWidget(self.btn_comparar)
        
        # Progreso
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)
        
        # Estadísticas
        grupo_stats = QGroupBox("📈 Estadísticas")
        layout_stats = QHBoxLayout()
        
        self.stats_widgets = {}
        stats_info = [
            ('coincidencias', 'Coincidencias', '#10b981'),
            ('discrepancias_prima', 'Disc. Prima', '#f59e0b'),
            ('discrepancias_fechas', 'Disc. Fechas', '#ef4444'),
            ('discrepancias_modalidad', 'Disc. Modalidad', '#8b5cf6')
        ]
        
        for key, label, color in stats_info:
            widget_stat = QWidget()
            layout_stat = QVBoxLayout()
            layout_stat.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout_stat.setSpacing(5)
            
            lbl_value = QLabel("0")
            lbl_value.setObjectName("statValue")
            lbl_value.setStyleSheet(f"color: {color};")
            lbl_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            lbl_name = QLabel(label)
            lbl_name.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_name.setStyleSheet("color: #a1a1aa; font-size: 8pt;")
            
            layout_stat.addWidget(lbl_value)
            layout_stat.addWidget(lbl_name)
            widget_stat.setLayout(layout_stat)
            
            layout_stats.addWidget(widget_stat)
            self.stats_widgets[key] = lbl_value
        
        grupo_stats.setLayout(layout_stats)
        layout.addWidget(grupo_stats)
        
        # Logs
        grupo_logs = QGroupBox("📋 Logs")
        layout_logs = QVBoxLayout()
        
        self.txt_logs = QTextEdit()
        self.txt_logs.setReadOnly(True)
        self.txt_logs.setMinimumHeight(200)
        layout_logs.addWidget(self.txt_logs)
        
        # Botones de acción
        layout_btns = QHBoxLayout()
        
        btn_limpiar = QPushButton("🗑️ Limpiar Logs")
        btn_limpiar.clicked.connect(self.limpiar_logs)
        
        self.btn_corregir = QPushButton("🔧 Corregir Modalidades")
        self.btn_corregir.clicked.connect(self.corregir_modalidades)
        self.btn_corregir.setEnabled(False)
        self.btn_corregir.setObjectName("btnWarning")
        
        self.btn_corregir_primas = QPushButton("💰 Corregir Primas")
        self.btn_corregir_primas.clicked.connect(self.corregir_primas)
        self.btn_corregir_primas.setEnabled(False)
        self.btn_corregir_primas.setObjectName("btnWarning")
        
        self.btn_primas_cero = QPushButton("🔄 Primas→0")
        self.btn_primas_cero.clicked.connect(self.primas_mensuales_a_cero)
        self.btn_primas_cero.setEnabled(False)
        
        self.btn_nits = QPushButton("🔧 NITs Completos")
        self.btn_nits.clicked.connect(self.colocar_nits_completos)
        self.btn_nits.setEnabled(False)
        self.btn_nits.setObjectName("btnSuccess")
        
        self.btn_exportar = QPushButton("📊 Exportar")
        self.btn_exportar.clicked.connect(self.exportar_reporte)
        self.btn_exportar.setEnabled(False)
        
        layout_btns.addWidget(btn_limpiar)
        layout_btns.addWidget(self.btn_corregir)
        layout_btns.addWidget(self.btn_corregir_primas)
        layout_btns.addWidget(self.btn_primas_cero)
        layout_btns.addWidget(self.btn_nits)
        layout_btns.addStretch()
        layout_btns.addWidget(self.btn_exportar)
        
        layout_logs.addLayout(layout_btns)
        grupo_logs.setLayout(layout_logs)
        layout.addWidget(grupo_logs, 1)
        
        self.setLayout(layout)
    
    def log(self, mensaje, tipo="info"):
        """Agrega mensaje a los logs con color"""
        colores = {
            'info': '#3b82f6',
            'success': '#10b981',
            'warning': '#f59e0b',
            'error': '#ef4444'
        }
        color = colores.get(tipo, '#e4e4e7')
        timestamp = datetime.now().strftime('%H:%M:%S')
        html = f'<span style="color: #71717a;">[{timestamp}]</span> <span style="color: {color};">{mensaje}</span>'
        self.txt_logs.append(html)
    
    def limpiar_logs(self):
        self.txt_logs.clear()
        self.log("📝 Logs limpiados", "info")
    
    def seleccionar_maviso(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo MAVISO", "",
            "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            self.ruta_maviso = archivo
            self.lbl_maviso.setText(Path(archivo).name)
            self.lbl_maviso.setStyleSheet("color: #10b981; font-weight: 600;")
            self.log(f"✅ MAVISO cargado: {Path(archivo).name}", "success")
            self.verificar_archivos()
    
    def seleccionar_celer(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo CELER", "",
            "Excel Files (*.xlsx *.xls)"
        )
        if archivo:
            self.ruta_celer = archivo
            self.lbl_celer.setText(Path(archivo).name)
            self.lbl_celer.setStyleSheet("color: #10b981; font-weight: 600;")
            self.log(f"✅ CELER cargado: {Path(archivo).name}", "success")
            self.verificar_archivos()
    
    def verificar_archivos(self):
        if self.ruta_maviso and self.ruta_celer:
            self.btn_comparar.setEnabled(True)
    
    def comparar(self):
        self.btn_comparar.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.log("🚀 Iniciando comparación...", "info")
        
        self.thread = ComparadorThread(self.comparador, self.ruta_maviso, self.ruta_celer)
        self.thread.progreso.connect(lambda msg: self.log(msg, "info"))
        self.thread.terminado.connect(self.comparacion_terminada)
        self.thread.start()
    
    def comparacion_terminada(self, exito, mensaje, estadisticas):
        self.btn_comparar.setEnabled(True)
        self.progress.setVisible(False)
        
        if exito:
            for key, widget in self.stats_widgets.items():
                valor = estadisticas.get(key, 0)
                widget.setText(str(valor))
            
            self.log("", "info")
            self.log("═" * 50, "success")
            self.log("✅ COMPARACIÓN COMPLETADA", "success")
            self.log("═" * 50, "success")
            
            self.btn_exportar.setEnabled(True)
            self.btn_corregir.setEnabled(True)
            self.btn_corregir_primas.setEnabled(True)
            self.btn_primas_cero.setEnabled(True)
            self.btn_nits.setEnabled(True)
        else:
            self.log(f"❌ Error: {mensaje}", "error")
    
    def exportar_reporte(self):
        try:
            ruta = self.comparador.exportar_reporte()
            self.log(f"✅ Reporte exportado: {Path(ruta).name}", "success")
            QMessageBox.information(self, "Éxito", f"Reporte exportado:\n{ruta}")
            if os.name == 'nt':
                os.startfile(ruta)
        except Exception as e:
            self.log(f"❌ Error al exportar: {str(e)}", "error")
            QMessageBox.critical(self, "Error", str(e))
    
    def corregir_modalidades(self):
        """Corrige modalidades en MAVISO usando CELER como fuente"""
        if not self.ruta_maviso or not self.comparador.resultados:
            QMessageBox.warning(self, "Advertencia", "Ejecute la comparación primero")
            return
        
        respuesta = QMessageBox.question(
            self, "Confirmar",
            "¿Actualizar modalidades en MAVISO con los valores de CELER?\n\n"
            "⚠️ Se modificará el archivo original manteniendo su formato.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("🔧 Corrigiendo modalidades...", "info")
            
            # Cargar CELER
            df_celer = pd.read_excel(self.ruta_celer, skiprows=3)
            
            # Columnas
            MAVISO_COL_POLIZA = 0
            MAVISO_COL_MODALIDAD = 21
            CELER_COL_POLIZA = 20
            CELER_COL_MODALIDAD = 27
            
            # Crear diccionario CELER
            celer_dict = {}
            for idx, row in df_celer.iterrows():
                poliza = str(row.iloc[CELER_COL_POLIZA]).strip()
                modalidad = str(row.iloc[CELER_COL_MODALIDAD]).strip().upper()
                celer_dict[poliza] = modalidad
            
            # Cargar MAVISO con openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            cambios = []
            for fila in range(2, ws.max_row + 1):
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_modalidad = ws.cell(row=fila, column=MAVISO_COL_MODALIDAD + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                modalidad_maviso = str(celda_modalidad.value).strip().upper() if celda_modalidad.value else ""
                
                if poliza in celer_dict:
                    modalidad_celer = celer_dict[poliza]
                    if modalidad_maviso != modalidad_celer and modalidad_celer in ['MENSUAL', 'ANUAL']:
                        celda_modalidad.value = modalidad_celer
                        cambios.append({
                            'poliza': poliza,
                            'fila': fila,
                            'anterior': modalidad_maviso,
                            'nuevo': modalidad_celer
                        })
            
            if len(cambios) == 0:
                self.log("✅ No hay discrepancias de modalidad", "success")
                QMessageBox.information(self, "Info", "No hay cambios necesarios.")
                wb.close()
                return
            
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"✅ Corregidas {len(cambios)} modalidades", "success")
            QMessageBox.information(self, "Éxito", f"Se corrigieron {len(cambios)} modalidades.")
            
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
                
        except Exception as e:
            self.log(f"❌ Error: {str(e)}", "error")
            QMessageBox.critical(self, "Error", str(e))
    
    def corregir_primas(self):
        """Corrige primas en MAVISO usando CELER como fuente"""
        if not self.ruta_maviso or not self.comparador.resultados:
            QMessageBox.warning(self, "Advertencia", "Ejecute la comparación primero")
            return
        
        respuesta = QMessageBox.question(
            self, "Confirmar",
            "¿Actualizar primas en MAVISO con los valores de CELER?\n\n"
            "⚠️ Se modificará el archivo original manteniendo su formato.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("💰 Corrigiendo primas...", "info")
            
            # Cargar CELER
            df_celer = pd.read_excel(self.ruta_celer, skiprows=3)
            
            # Columnas
            MAVISO_COL_POLIZA = 0
            MAVISO_COL_PRIMA = 14
            CELER_COL_POLIZA = 20
            CELER_COL_PRIMA = 42
            
            # Crear diccionario CELER
            celer_dict = {}
            for idx, row in df_celer.iterrows():
                poliza = str(row.iloc[CELER_COL_POLIZA]).strip()
                try:
                    prima = abs(round(float(row.iloc[CELER_COL_PRIMA]), 2))
                    celer_dict[poliza] = prima
                except (ValueError, TypeError):
                    continue
            
            # Cargar MAVISO con openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            cambios = []
            for fila in range(2, ws.max_row + 1):
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_prima = ws.cell(row=fila, column=MAVISO_COL_PRIMA + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                
                try:
                    prima_maviso = abs(round(float(celda_prima.value), 2)) if celda_prima.value else 0
                except (ValueError, TypeError):
                    prima_maviso = 0
                
                if poliza in celer_dict:
                    prima_celer = celer_dict[poliza]
                    if prima_maviso != prima_celer:
                        celda_prima.value = prima_celer
                        cambios.append({
                            'poliza': poliza,
                            'fila': fila,
                            'anterior': prima_maviso,
                            'nuevo': prima_celer
                        })
            
            if len(cambios) == 0:
                self.log("✅ No hay discrepancias de prima", "success")
                QMessageBox.information(self, "Info", "No hay cambios necesarios.")
                wb.close()
                return
            
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"✅ Corregidas {len(cambios)} primas", "success")
            QMessageBox.information(self, "Éxito", f"Se corrigieron {len(cambios)} primas.")
            
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
                
        except Exception as e:
            self.log(f"❌ Error: {str(e)}", "error")
            QMessageBox.critical(self, "Error", str(e))
    
    def primas_mensuales_a_cero(self):
        """Pone primas en 0 para pólizas MENSUALES"""
        if not self.ruta_maviso:
            QMessageBox.warning(self, "Advertencia", "Cargue el archivo MAVISO primero")
            return
        
        respuesta = QMessageBox.question(
            self, "Confirmar",
            "¿Poner en CERO las primas de pólizas MENSUALES?\n\n"
            "⚠️ Se modificará el archivo original manteniendo su formato.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("🔄 Ajustando primas mensuales...", "info")
            
            # Cargar MAVISO con openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            MAVISO_COL_POLIZA = 0
            MAVISO_COL_PRIMA = 14
            MAVISO_COL_MODALIDAD = 21
            
            cambios = []
            for fila in range(2, ws.max_row + 1):
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_prima = ws.cell(row=fila, column=MAVISO_COL_PRIMA + 1)
                celda_modalidad = ws.cell(row=fila, column=MAVISO_COL_MODALIDAD + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                modalidad = str(celda_modalidad.value).strip().upper() if celda_modalidad.value else ""
                
                try:
                    prima_actual = abs(round(float(celda_prima.value), 2)) if celda_prima.value else 0
                except (ValueError, TypeError):
                    prima_actual = 0
                
                if modalidad == "MENSUAL" and prima_actual != 0:
                    celda_prima.value = 0
                    cambios.append({'poliza': poliza, 'fila': fila, 'prima_anterior': prima_actual})
            
            if len(cambios) == 0:
                self.log("✅ No hay mensuales con prima ≠ 0", "success")
                QMessageBox.information(self, "Info", "No hay cambios necesarios.")
                wb.close()
                return
            
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"✅ Ajustadas {len(cambios)} primas a cero", "success")
            QMessageBox.information(self, "Éxito", f"Se ajustaron {len(cambios)} primas mensuales a 0.")
            
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
                
        except Exception as e:
            self.log(f"❌ Error: {str(e)}", "error")
            QMessageBox.critical(self, "Error", str(e))
    
    def _verificar_api_dian(self):
        """Verifica si la API DIAN está corriendo"""
        try:
            response = requests.get("http://localhost:8000/health", timeout=2)
            return response.status_code == 200
        except:
            return False
    
    def _iniciar_api_dian(self):
        """Inicia la API DIAN en segundo plano"""
        if self._verificar_api_dian():
            return True
        
        self.log("🚀 Iniciando API DIAN...", "info")
        
        # Ruta al directorio backend
        script_dir = Path(__file__).parent.parent.parent
        backend_dir = script_dir / 'backend'
        
        if not backend_dir.exists():
            self.log(f"❌ No se encuentra backend: {backend_dir}", "error")
            return False
        
        try:
            subprocess.Popen(
                [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8000"],
                cwd=str(backend_dir),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            for i in range(10):
                time.sleep(1)
                if self._verificar_api_dian():
                    self.log("✅ API DIAN iniciada", "success")
                    return True
                self.log(f"   Esperando API... ({i+1}/10)", "info")
            
            self.log("❌ No se pudo iniciar API DIAN", "error")
            return False
            
        except Exception as e:
            self.log(f"❌ Error iniciando API: {str(e)}", "error")
            return False
    
    def _calcular_dv_dian(self, nit, reintentos=3):
        """Calcula el DV usando API DIAN con reintentos"""
        nit_str = re.sub(r'\D', '', str(nit))
        
        if not nit_str or len(nit_str) > 15 or len(nit_str) < 5:
            return None
        
        for intento in range(reintentos):
            try:
                response = requests.post(
                    "http://localhost:8000/calcular",
                    json={"nit": nit_str},
                    timeout=10
                )
                
                if response.status_code == 200:
                    data = response.json()
                    return str(data["digito_verificacion"])
                else:
                    if intento < reintentos - 1:
                        time.sleep(0.5)
                        continue
                    return None
            
            except requests.exceptions.ConnectionError:
                if intento < reintentos - 1:
                    if not self._verificar_api_dian():
                        self._iniciar_api_dian()
                        time.sleep(2)
                    continue
                return None
            except Exception:
                if intento < reintentos - 1:
                    time.sleep(0.5)
                    continue
                return None
        
        return None
    
    def colocar_nits_completos(self):
        """Calcula y agrega DV a NITs de personas jurídicas"""
        if not self.ruta_maviso:
            QMessageBox.warning(self, "Advertencia", "Cargue el archivo MAVISO primero")
            return
        
        respuesta = QMessageBox.question(
            self, "Confirmar",
            "¿Calcular y agregar DV a NITs de personas jurídicas?\n\n"
            "• Usa API DIAN para calcular el DV\n"
            "• Solo procesa personas JURÍDICAS (columna AC = 'J')\n"
            "• Ignora personas NATURALES (columna AC = 'N')\n"
            "⚠️ Se modificará el archivo original manteniendo su formato.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        try:
            self.log("", "info")
            self.log("🔧 Iniciando cálculo de NITs...", "info")
            
            # Verificar/Iniciar API DIAN
            if not self._verificar_api_dian():
                if not self._iniciar_api_dian():
                    QMessageBox.critical(
                        self, "Error API",
                        "No se pudo iniciar la API DIAN.\n\n"
                        "Verifique que el directorio 'backend' existe."
                    )
                    return
            else:
                self.log("✅ API DIAN disponible", "success")
            
            # Columnas
            MAVISO_COL_POLIZA = 0
            MAVISO_COL_DOCUMENTO = 27
            MAVISO_COL_TIPO_PERSONA = 28
            
            # Cargar MAVISO con openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(self.ruta_maviso)
            ws = wb.active
            
            cambios = []
            errores = []
            naturales_ignoradas = 0
            
            for fila in range(2, ws.max_row + 1):
                celda_poliza = ws.cell(row=fila, column=MAVISO_COL_POLIZA + 1)
                celda_documento = ws.cell(row=fila, column=MAVISO_COL_DOCUMENTO + 1)
                celda_tipo = ws.cell(row=fila, column=MAVISO_COL_TIPO_PERSONA + 1)
                
                poliza = str(celda_poliza.value).strip() if celda_poliza.value else ""
                documento = str(celda_documento.value).strip() if celda_documento.value else ""
                tipo_persona = str(celda_tipo.value).strip().upper() if celda_tipo.value else ""
                
                # Solo procesar JURÍDICAS
                if tipo_persona != 'J':
                    if tipo_persona == 'N':
                        naturales_ignoradas += 1
                    continue
                
                # Verificar si ya tiene DV
                if '-' in documento:
                    continue
                
                # Verificar que sea número válido
                documento_limpio = re.sub(r'\D', '', documento)
                if not documento_limpio or len(documento_limpio) < 5:
                    continue
                
                # Calcular DV
                dv = self._calcular_dv_dian(documento_limpio)
                
                if dv is not None:
                    nit_completo = f"{documento_limpio}-{dv}"
                    celda_documento.value = nit_completo
                    cambios.append({
                        'poliza': poliza,
                        'fila': fila,
                        'anterior': documento,
                        'nuevo': nit_completo
                    })
                else:
                    errores.append({'poliza': poliza, 'fila': fila, 'documento': documento})
                
                # Mostrar progreso cada 50
                if len(cambios) % 50 == 0 and len(cambios) > 0:
                    self.log(f"   Procesados: {len(cambios)} NITs...", "info")
            
            self.log(f"\n📊 Personas naturales ignoradas: {naturales_ignoradas}", "info")
            self.log(f"✅ NITs procesados: {len(cambios)}", "success" if len(cambios) > 0 else "info")
            
            if len(errores) > 0:
                self.log(f"⚠️ Errores al calcular: {len(errores)}", "warning")
            
            if len(cambios) == 0:
                self.log("✅ No hay NITs jurídicos sin DV", "success")
                QMessageBox.information(self, "Info", "No hay cambios necesarios.")
                wb.close()
                return
            
            wb.save(self.ruta_maviso)
            wb.close()
            
            self.log(f"✅ Actualizados {len(cambios)} NITs", "success")
            
            mensaje = f"Se actualizaron {len(cambios)} NITs con su dígito de verificación."
            if len(errores) > 0:
                mensaje += f"\n\n⚠️ {len(errores)} NITs no pudieron ser calculados."
            
            QMessageBox.information(self, "Éxito", mensaje)
            
            if os.name == 'nt':
                os.startfile(self.ruta_maviso)
                
        except Exception as e:
            self.log(f"❌ Error: {str(e)}", "error")
            QMessageBox.critical(self, "Error", str(e))


class ComparadorGUIPro(QMainWindow):
    """Ventana principal profesional con dark mode"""
    
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Comparador de Archivos MAVISO - Professional Edition")
        self.setMinimumSize(1400, 900)
        
        # Widget central
        central = QWidget()
        self.setCentralWidget(central)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header
        header = QWidget()
        header.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3b82f6, stop:1 #8b5cf6); padding: 20px;")
        header_layout = QVBoxLayout()
        
        title = QLabel("📊 COMPARADOR DE ARCHIVOS MAVISO")
        title.setStyleSheet("color: white; font-size: 24pt; font-weight: 700; background: transparent;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        subtitle = QLabel("Sistema Profesional de Validación y Corrección de Datos")
        subtitle.setStyleSheet("color: rgba(255, 255, 255, 0.9); font-size: 11pt; background: transparent;")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        header.setLayout(header_layout)
        
        layout.addWidget(header)
        
        # Tabs
        tabs = QTabWidget()
        tabs.setTabPosition(QTabWidget.TabPosition.North)
        
        # Solo agregar la pestaña principal por ahora
        tab_maviso_celer = TabMavisoVsCeler()
        tabs.addTab(tab_maviso_celer, "📊 MAVISO vs CELER")
        
        layout.addWidget(tabs)
        
        central.setLayout(layout)
        
        # Centrar ventana
        self.center()
    
    def center(self):
        """Centra la ventana en la pantalla"""
        screen = QApplication.primaryScreen().geometry()
        size = self.geometry()
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setStyleSheet(DARK_STYLESHEET)
    
    ventana = ComparadorGUIPro()
    ventana.show()
    
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
