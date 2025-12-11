"""
Comparador de Archivos Maviso
=============================
Compara dos archivos Excel de Maviso para validar que los datos coincidan.

Lógica:
- Clave: Número de póliza (columna A)
- Campos a validar: Prima (O), Fecha inicio (K), Fecha fin (L)
- Si la póliza existe en ambos archivos, los campos deben coincidir
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Tuple, Dict, List
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Columnas de Maviso (0-indexed para pandas)
COL_POLIZA = 0       # A - Número de póliza (CLAVE)
COL_FECHA_INICIO = 10  # K - Fecha inicio
COL_FECHA_FIN = 11     # L - Fecha fin
COL_PRIMA = 14         # O - Prima sin IVA

# Estilos para Excel (estilo Maviso)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ComparadorMaviso:
    """Clase para comparar dos archivos Maviso"""
    
    def __init__(self):
        self.df_manual = None
        self.df_generado = None
        self.resultados = None
        self.estadisticas = {}
    
    def cargar_archivo(self, ruta: str, nombre: str = "archivo") -> Tuple[bool, str, pd.DataFrame]:
        """
        Carga un archivo Excel de Maviso.
        
        Args:
            ruta: Ruta al archivo Excel
            nombre: Nombre descriptivo para logs
            
        Returns:
            Tuple (éxito, mensaje, dataframe)
        """
        try:
            ruta_path = Path(ruta)
            if not ruta_path.exists():
                return False, f"Archivo no encontrado: {ruta}", None
            
            df = pd.read_excel(ruta, header=0)
            
            # Verificar que tenga las columnas necesarias
            if len(df.columns) < 15:
                return False, f"El archivo {nombre} no tiene suficientes columnas (mínimo 15)", None
            
            logger.info(f"✅ {nombre} cargado: {len(df)} filas, {len(df.columns)} columnas")
            return True, f"Cargado correctamente: {len(df)} registros", df
            
        except Exception as e:
            return False, f"Error cargando {nombre}: {str(e)}", None
    
    def cargar_archivo_manual(self, ruta: str) -> Tuple[bool, str]:
        """Carga el archivo modificado manualmente"""
        exito, mensaje, df = self.cargar_archivo(ruta, "Archivo Manual")
        if exito:
            self.df_manual = df
        return exito, mensaje
    
    def cargar_archivo_generado(self, ruta: str) -> Tuple[bool, str]:
        """Carga el archivo generado por el script"""
        exito, mensaje, df = self.cargar_archivo(ruta, "Archivo Generado")
        if exito:
            self.df_generado = df
        return exito, mensaje
    
    def _normalizar_valor(self, valor):
        """Normaliza un valor para comparación"""
        if pd.isna(valor):
            return None
        
        # Si es fecha, convertir a string formato estándar
        if isinstance(valor, (datetime, pd.Timestamp)):
            return valor.strftime('%Y-%m-%d')
        
        # Si es número, redondear a 2 decimales
        if isinstance(valor, (int, float)):
            return round(float(valor), 2)
        
        # Convertir a string y limpiar
        return str(valor).strip()
    
    def _comparar_valores(self, val1, val2) -> bool:
        """Compara dos valores normalizados"""
        norm1 = self._normalizar_valor(val1)
        norm2 = self._normalizar_valor(val2)
        
        # Ambos None = iguales
        if norm1 is None and norm2 is None:
            return True
        
        # Uno None y otro no = diferentes
        if norm1 is None or norm2 is None:
            return False
        
        return norm1 == norm2
    
    def comparar(self) -> Tuple[bool, str, Dict]:
        """
        Ejecuta la comparación entre los dos archivos.
        
        Returns:
            Tuple (éxito, mensaje, resultados)
        """
        if self.df_manual is None:
            return False, "No se ha cargado el archivo manual", None
        
        if self.df_generado is None:
            return False, "No se ha cargado el archivo generado", None
        
        logger.info("=" * 60)
        logger.info("INICIANDO COMPARACIÓN")
        logger.info("=" * 60)
        
        # Obtener nombres de columnas
        cols_manual = self.df_manual.columns.tolist()
        cols_generado = self.df_generado.columns.tolist()
        
        # Crear diccionario del archivo manual por número de póliza
        polizas_manual = {}
        for idx, row in self.df_manual.iterrows():
            poliza = self._normalizar_valor(row.iloc[COL_POLIZA])
            if poliza:
                polizas_manual[poliza] = {
                    'fila': idx + 2,  # +2 por encabezado y 0-index
                    'prima': row.iloc[COL_PRIMA],
                    'fecha_inicio': row.iloc[COL_FECHA_INICIO],
                    'fecha_fin': row.iloc[COL_FECHA_FIN]
                }
        
        # Crear diccionario del archivo generado por número de póliza
        polizas_generado = {}
        for idx, row in self.df_generado.iterrows():
            poliza = self._normalizar_valor(row.iloc[COL_POLIZA])
            if poliza:
                polizas_generado[poliza] = {
                    'fila': idx + 2,
                    'prima': row.iloc[COL_PRIMA],
                    'fecha_inicio': row.iloc[COL_FECHA_INICIO],
                    'fecha_fin': row.iloc[COL_FECHA_FIN]
                }
        
        # Estadísticas
        total_manual = len(polizas_manual)
        total_generado = len(polizas_generado)
        
        # Comparar
        coincidencias = []
        discrepancias = []
        solo_en_manual = []
        solo_en_generado = []
        
        # Buscar pólizas del archivo manual en el generado
        for poliza, datos_manual in polizas_manual.items():
            if poliza in polizas_generado:
                datos_generado = polizas_generado[poliza]
                
                # Comparar campos
                errores = []
                
                # Prima
                if not self._comparar_valores(datos_manual['prima'], datos_generado['prima']):
                    errores.append({
                        'campo': 'Prima (O)',
                        'manual': self._normalizar_valor(datos_manual['prima']),
                        'generado': self._normalizar_valor(datos_generado['prima'])
                    })
                
                # Fecha inicio
                if not self._comparar_valores(datos_manual['fecha_inicio'], datos_generado['fecha_inicio']):
                    errores.append({
                        'campo': 'Fecha Inicio (K)',
                        'manual': self._normalizar_valor(datos_manual['fecha_inicio']),
                        'generado': self._normalizar_valor(datos_generado['fecha_inicio'])
                    })
                
                # Fecha fin
                if not self._comparar_valores(datos_manual['fecha_fin'], datos_generado['fecha_fin']):
                    errores.append({
                        'campo': 'Fecha Fin (L)',
                        'manual': self._normalizar_valor(datos_manual['fecha_fin']),
                        'generado': self._normalizar_valor(datos_generado['fecha_fin'])
                    })
                
                if errores:
                    discrepancias.append({
                        'poliza': poliza,
                        'fila_manual': datos_manual['fila'],
                        'fila_generado': datos_generado['fila'],
                        'errores': errores
                    })
                    # Log detallado de la discrepancia
                    logger.warning(f"❌ DISCREPANCIA - Póliza: {poliza}")
                    logger.warning(f"   Fila manual: {datos_manual['fila']} | Fila generado: {datos_generado['fila']}")
                    for err in errores:
                        logger.warning(f"   → {err['campo']}: Manual='{err['manual']}' vs Generado='{err['generado']}'")
                else:
                    coincidencias.append(poliza)
            else:
                solo_en_manual.append({
                    'poliza': poliza,
                    'fila': datos_manual['fila']
                })
                logger.info(f"⚠️ Solo en MANUAL - Póliza: {poliza} (Fila {datos_manual['fila']})")
        
        # Buscar pólizas solo en generado
        for poliza in polizas_generado:
            if poliza not in polizas_manual:
                solo_en_generado.append({
                    'poliza': poliza,
                    'fila': polizas_generado[poliza]['fila']
                })
                logger.info(f"⚠️ Solo en GENERADO - Póliza: {poliza} (Fila {polizas_generado[poliza]['fila']})")
        
        # Guardar resultados
        self.resultados = {
            'coincidencias': coincidencias,
            'discrepancias': discrepancias,
            'solo_manual': solo_en_manual,
            'solo_generado': solo_en_generado
        }
        
        self.estadisticas = {
            'total_manual': total_manual,
            'total_generado': total_generado,
            'coincidencias': len(coincidencias),
            'discrepancias': len(discrepancias),
            'solo_manual': len(solo_en_manual),
            'solo_generado': len(solo_en_generado)
        }
        
        # Log resumen
        logger.info("=" * 60)
        logger.info("RESUMEN DE COMPARACIÓN")
        logger.info("=" * 60)
        logger.info(f"  Total pólizas archivo manual: {total_manual}")
        logger.info(f"  Total pólizas archivo generado: {total_generado}")
        logger.info(f"  ✅ Coincidencias exactas: {len(coincidencias)}")
        logger.info(f"  ❌ Discrepancias: {len(discrepancias)}")
        logger.info(f"  ⚠️ Solo en manual: {len(solo_en_manual)}")
        logger.info(f"  ⚠️ Solo en generado: {len(solo_en_generado)}")
        logger.info("=" * 60)
        
        return True, "Comparación completada", self.resultados
    
    def obtener_estadisticas(self) -> Dict:
        """Retorna las estadísticas de la última comparación"""
        return self.estadisticas
    
    def obtener_discrepancias(self) -> List[Dict]:
        """Retorna la lista de discrepancias encontradas"""
        if self.resultados:
            return self.resultados['discrepancias']
        return []
    
    def _aplicar_formato_header(self, worksheet):
        """Aplica formato estándar a encabezados estilo Maviso"""
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        
        # Auto-ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _aplicar_formato_datos(self, worksheet, fila_inicio=2):
        """Aplica bordes y alineación a todas las celdas de datos"""
        for row in worksheet.iter_rows(min_row=fila_inicio, max_row=worksheet.max_row):
            for cell in row:
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
    
    def exportar_reporte(self, ruta_salida: str = None) -> str:
        """
        Exporta un reporte detallado de la comparación a Excel con formato Maviso.
        
        Args:
            ruta_salida: Ruta para el archivo de salida
            
        Returns:
            Ruta del archivo generado
        """
        if not self.resultados:
            raise ValueError("No hay resultados de comparación. Ejecute comparar() primero.")
        
        # Generar nombre de archivo
        if not ruta_salida:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            carpeta = Path(__file__).parent / 'output'
            carpeta.mkdir(exist_ok=True)
            ruta_salida = carpeta / f'comparacion_maviso_{timestamp}.xlsx'
        
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            # ========================================
            # Hoja 1: Resumen
            # ========================================
            resumen_data = {
                'Métrica': [
                    'Total pólizas archivo manual',
                    'Total pólizas archivo generado',
                    'Coincidencias exactas',
                    'Discrepancias',
                    'Solo en archivo manual',
                    'Solo en archivo generado'
                ],
                'Valor': [
                    self.estadisticas['total_manual'],
                    self.estadisticas['total_generado'],
                    self.estadisticas['coincidencias'],
                    self.estadisticas['discrepancias'],
                    self.estadisticas['solo_manual'],
                    self.estadisticas['solo_generado']
                ],
                'Estado': [
                    '📊',
                    '📊',
                    '✅',
                    '❌' if self.estadisticas['discrepancias'] > 0 else '✅',
                    '⚠️' if self.estadisticas['solo_manual'] > 0 else '✅',
                    '⚠️' if self.estadisticas['solo_generado'] > 0 else '✅'
                ]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Aplicar formato a Resumen
            ws_resumen = writer.sheets['Resumen']
            self._aplicar_formato_header(ws_resumen)
            self._aplicar_formato_datos(ws_resumen)
            
            # Colorear filas según tipo
            for row_idx in range(2, ws_resumen.max_row + 1):
                estado = ws_resumen.cell(row=row_idx, column=3).value
                if estado == '✅':
                    for col in range(1, 4):
                        ws_resumen.cell(row=row_idx, column=col).fill = SUCCESS_FILL
                elif estado == '❌':
                    for col in range(1, 4):
                        ws_resumen.cell(row=row_idx, column=col).fill = ERROR_FILL
                elif estado == '⚠️':
                    for col in range(1, 4):
                        ws_resumen.cell(row=row_idx, column=col).fill = WARNING_FILL
            
            # ========================================
            # Hoja 2: Discrepancias
            # ========================================
            if self.resultados['discrepancias']:
                discrepancias_flat = []
                for d in self.resultados['discrepancias']:
                    for error in d['errores']:
                        discrepancias_flat.append({
                            'Póliza': d['poliza'],
                            'Fila Manual': d['fila_manual'],
                            'Fila Generado': d['fila_generado'],
                            'Campo': error['campo'],
                            'Valor Manual': error['manual'],
                            'Valor Generado': error['generado'],
                            'Diferencia': self._calcular_diferencia(error['manual'], error['generado'])
                        })
                df_disc = pd.DataFrame(discrepancias_flat)
                df_disc.to_excel(writer, sheet_name='Discrepancias', index=False)
                
                # Aplicar formato
                ws_disc = writer.sheets['Discrepancias']
                self._aplicar_formato_header(ws_disc)
                self._aplicar_formato_datos(ws_disc)
                
                # Colorear todas las filas de discrepancias en rojo claro
                for row in range(2, ws_disc.max_row + 1):
                    for col in range(1, 8):
                        ws_disc.cell(row=row, column=col).fill = ERROR_FILL
            
            # ========================================
            # Hoja 3: Solo en Manual
            # ========================================
            if self.resultados['solo_manual']:
                df_solo_manual = pd.DataFrame(self.resultados['solo_manual'])
                df_solo_manual.columns = ['Póliza', 'Fila']
                df_solo_manual['Estado'] = '⚠️ No existe en archivo generado'
                df_solo_manual.to_excel(writer, sheet_name='Solo_en_Manual', index=False)
                
                # Aplicar formato
                ws_manual = writer.sheets['Solo_en_Manual']
                self._aplicar_formato_header(ws_manual)
                self._aplicar_formato_datos(ws_manual)
                
                # Colorear en amarillo
                for row in range(2, ws_manual.max_row + 1):
                    for col in range(1, 4):
                        ws_manual.cell(row=row, column=col).fill = WARNING_FILL
            
            # ========================================
            # Hoja 4: Solo en Generado
            # ========================================
            if self.resultados['solo_generado']:
                df_solo_gen = pd.DataFrame(self.resultados['solo_generado'])
                df_solo_gen.columns = ['Póliza', 'Fila']
                df_solo_gen['Estado'] = '⚠️ No existe en archivo manual'
                df_solo_gen.to_excel(writer, sheet_name='Solo_en_Generado', index=False)
                
                # Aplicar formato
                ws_gen = writer.sheets['Solo_en_Generado']
                self._aplicar_formato_header(ws_gen)
                self._aplicar_formato_datos(ws_gen)
                
                # Colorear en amarillo
                for row in range(2, ws_gen.max_row + 1):
                    for col in range(1, 4):
                        ws_gen.cell(row=row, column=col).fill = WARNING_FILL
            
            # ========================================
            # Hoja 5: Coincidencias (opcional, para verificación)
            # ========================================
            if self.resultados['coincidencias']:
                df_coincidencias = pd.DataFrame({
                    'Póliza': self.resultados['coincidencias'],
                    'Estado': ['✅ OK'] * len(self.resultados['coincidencias'])
                })
                df_coincidencias.to_excel(writer, sheet_name='Coincidencias', index=False)
                
                # Aplicar formato
                ws_coinc = writer.sheets['Coincidencias']
                self._aplicar_formato_header(ws_coinc)
                self._aplicar_formato_datos(ws_coinc)
                
                # Colorear en verde
                for row in range(2, ws_coinc.max_row + 1):
                    for col in range(1, 3):
                        ws_coinc.cell(row=row, column=col).fill = SUCCESS_FILL
        
        logger.info(f"📊 Reporte exportado: {ruta_salida}")
        return str(ruta_salida)
    
    def _calcular_diferencia(self, val1, val2):
        """Calcula la diferencia numérica si ambos valores son números"""
        try:
            if val1 is None or val2 is None:
                return "N/A"
            num1 = float(val1)
            num2 = float(val2)
            diff = num1 - num2
            if diff > 0:
                return f"+{diff:,.2f}"
            else:
                return f"{diff:,.2f}"
        except (ValueError, TypeError):
            return "N/A (fechas)"


def comparar_archivos(ruta_manual: str, ruta_generado: str) -> Tuple[Dict, str]:
    """
    Función principal para comparar dos archivos.
    
    Args:
        ruta_manual: Ruta al archivo modificado manualmente
        ruta_generado: Ruta al archivo generado por el script
        
    Returns:
        Tuple (estadísticas, ruta_reporte)
    """
    comparador = ComparadorMaviso()
    
    # Cargar archivos
    exito, mensaje = comparador.cargar_archivo_manual(ruta_manual)
    if not exito:
        raise ValueError(mensaje)
    
    exito, mensaje = comparador.cargar_archivo_generado(ruta_generado)
    if not exito:
        raise ValueError(mensaje)
    
    # Comparar
    exito, mensaje, resultados = comparador.comparar()
    if not exito:
        raise ValueError(mensaje)
    
    # Exportar reporte
    ruta_reporte = comparador.exportar_reporte()
    
    return comparador.obtener_estadisticas(), ruta_reporte


# Test básico
if __name__ == "__main__":
    print("=" * 60)
    print("COMPARADOR DE ARCHIVOS MAVISO")
    print("=" * 60)
    print("\nUso:")
    print("  from comparar_archivos import ComparadorMaviso, comparar_archivos")
    print("\n  # Opción 1: Función directa")
    print("  stats, reporte = comparar_archivos('manual.xlsx', 'generado.xlsx')")
    print("\n  # Opción 2: Clase para más control")
    print("  comp = ComparadorMaviso()")
    print("  comp.cargar_archivo_manual('manual.xlsx')")
    print("  comp.cargar_archivo_generado('generado.xlsx')")
    print("  comp.comparar()")
    print("  comp.exportar_reporte()")
    print("\n" + "=" * 60)
