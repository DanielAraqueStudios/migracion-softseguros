"""
Ejecutar Migración Completa
===========================
Script que ejecuta en secuencia:
1. Migración CELER → MAVISO
2. Cálculo de Dígito de Verificación (API DIAN)

Uso: python ejecutar_migracion_completa.py
"""

import sys
import os
from pathlib import Path
from datetime import datetime

# Agregar el directorio actual al path
sys.path.insert(0, str(Path(__file__).parent))

from llenar_maviso import (
    ARCHIVO_CELER, 
    ARCHIVO_MAVISO, 
    CARPETA_OUTPUT,
    aplicar_digito_verificacion,
    iniciar_api,
    detener_api
)
import pandas as pd
from openpyxl import load_workbook
import copy


def print_header(titulo):
    """Imprime un encabezado formateado"""
    print("\n" + "=" * 60)
    print(f"  {titulo}")
    print("=" * 60)


def print_success(msg):
    """Imprime mensaje de éxito"""
    print(f"✅ {msg}")


def print_error(msg):
    """Imprime mensaje de error"""
    print(f"❌ {msg}")


def print_info(msg):
    """Imprime mensaje informativo"""
    print(f"📋 {msg}")


def print_warning(msg):
    """Imprime advertencia"""
    print(f"⚠️ {msg}")


def log_callback(msg, tipo):
    """Callback para mostrar logs en consola"""
    if tipo == 'error':
        print(f"❌ {msg}")
    elif tipo == 'success':
        print(f"✅ {msg}")
    elif tipo == 'warning':
        print(f"⚠️ {msg}")
    else:
        print(f"   {msg}")


def ejecutar_migracion():
    """
    Ejecuta la migración CELER → MAVISO
    Retorna la ruta del archivo generado o None si hay error
    """
    print_header("PASO 1: MIGRACIÓN CELER → MAVISO")
    
    # Verificar archivos de entrada
    if not ARCHIVO_CELER.exists():
        print_error(f"No se encuentra archivo CELER: {ARCHIVO_CELER}")
        return None
    
    if not ARCHIVO_MAVISO.exists():
        print_error(f"No se encuentra plantilla Maviso: {ARCHIVO_MAVISO}")
        return None
    
    print_info(f"Archivo CELER: {ARCHIVO_CELER.name}")
    print_info(f"Plantilla Maviso: {ARCHIVO_MAVISO.name}")
    print()
    
    try:
        # 1. Cargar archivo CELER
        print("📂 Cargando archivo CELER (skiprows=3)...")
        df_celer = pd.read_excel(ARCHIVO_CELER, skiprows=3)
        print_success(f"CELER cargado: {len(df_celer):,} filas, {len(df_celer.columns)} columnas")
        
        # 2. Cargar plantilla Maviso (preservando formato)
        print("📂 Cargando plantilla Maviso...")
        wb = load_workbook(ARCHIVO_MAVISO)
        ws = wb.active
        print_success(f"Maviso cargado: {ws.max_row:,} filas, {ws.max_column} columnas")
        
        # 3. Guardar estilos de la fila 2 como plantilla
        print("💾 Preservando estilos originales...")
        estilos_plantilla = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            estilos_plantilla[col] = {
                'font': copy.copy(cell.font),
                'fill': copy.copy(cell.fill),
                'border': copy.copy(cell.border),
                'alignment': copy.copy(cell.alignment),
                'number_format': cell.number_format
            }
        
        # 4. Mapeo de columnas
        mapeo = {
            1: 'U',    # A ← U (Póliza)
            2: 'AE',   # B ← AE (Placa)
            3: 'R',    # C ← R (Aseguradora)
            5: 'S',    # E ← S (Ramo)
            10: 'BE',  # J ← BE (Ejecutivos)
            11: 'W',   # K ← W (F_Inicio)
            12: 'X',   # L ← X (F_Fin)
            15: 'AQ',  # O ← AQ (Prima sin IVA)
            24: 'AP',  # X ← AP (V_Asegurado)
            28: 'C',   # AB ← C (Identificación)
            29: 'A',   # AC ← A (Tipo_Persona)
            30: 'B',   # AD ← B (Tomador)
            31: 'C',   # AE ← C (Identificación)
            32: 'AS',  # AF ← AS (Asegurado)
            33: 'AT',  # AG ← AT (Iden_Asegurado)
            34: 'AW',  # AH ← AW (Beneficiario)
            35: 'AX',  # AI ← AX (Iden_Beneficiario)
        }
        
        # Función auxiliar para letra a índice
        def letra_a_indice(letra):
            resultado = 0
            for char in letra.upper():
                resultado = resultado * 26 + (ord(char) - ord('A') + 1)
            return resultado - 1
        
        # Función para convertir forma de pago
        def convertir_forma_pago(valor):
            if pd.isna(valor):
                return ''
            valor_str = str(valor).strip().upper()
            if valor_str == 'MENSUAL':
                return 'Fraccionado'
            elif valor_str == 'ANUAL':
                return 'Contado'
            else:
                return 'Fraccionado'
        
        # 5. Procesar cada fila
        print()
        print("🔄 Procesando filas...")
        filas_procesadas = 0
        col_w_maviso = 23  # Columna W (forma de pago)
        col_forma_pago_celer = letra_a_indice('Y')
        
        for idx_celer, row_celer in df_celer.iterrows():
            fila_maviso = idx_celer + 2
            
            # Aplicar mapeo de columnas
            for col_maviso, col_celer_letra in mapeo.items():
                col_celer_idx = letra_a_indice(col_celer_letra)
                valor = row_celer.iloc[col_celer_idx] if col_celer_idx < len(row_celer) else ''
                
                if pd.isna(valor):
                    valor = ''
                
                cell = ws.cell(row=fila_maviso, column=col_maviso)
                cell.value = valor
                
                # Aplicar estilo
                if col_maviso in estilos_plantilla:
                    estilo = estilos_plantilla[col_maviso]
                    cell.font = copy.copy(estilo['font'])
                    cell.fill = copy.copy(estilo['fill'])
                    cell.border = copy.copy(estilo['border'])
                    cell.alignment = copy.copy(estilo['alignment'])
                    cell.number_format = estilo['number_format']
            
            # Columna W - Forma de pago
            forma_pago_celer = row_celer.iloc[col_forma_pago_celer] if col_forma_pago_celer < len(row_celer) else ''
            forma_pago_maviso = convertir_forma_pago(forma_pago_celer)
            
            cell_w = ws.cell(row=fila_maviso, column=col_w_maviso)
            cell_w.value = forma_pago_maviso
            
            if col_w_maviso in estilos_plantilla:
                estilo = estilos_plantilla[col_w_maviso]
                cell_w.font = copy.copy(estilo['font'])
                cell_w.fill = copy.copy(estilo['fill'])
                cell_w.border = copy.copy(estilo['border'])
                cell_w.alignment = copy.copy(estilo['alignment'])
                cell_w.number_format = estilo['number_format']
            
            filas_procesadas += 1
            
            # Progreso cada 500 filas
            if filas_procesadas % 500 == 0:
                print(f"   Procesadas {filas_procesadas:,} filas...")
        
        # 6. Guardar archivo
        print()
        print("💾 Guardando archivo...")
        CARPETA_OUTPUT.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_salida = CARPETA_OUTPUT / f'Maviso_llenado_{timestamp}.xlsx'
        
        wb.save(archivo_salida)
        
        print()
        print_header("MIGRACIÓN COMPLETADA")
        print_success(f"Archivo generado: {archivo_salida.name}")
        print_success(f"Total filas procesadas: {filas_procesadas:,}")
        print_info(f"Ubicación: {archivo_salida}")
        
        return archivo_salida
        
    except Exception as e:
        print_error(f"Error en migración: {e}")
        import traceback
        traceback.print_exc()
        return None


def ejecutar_digito_verificacion(archivo_excel):
    """
    Ejecuta el cálculo de dígito de verificación
    """
    print_header("PASO 2: CÁLCULO DÍGITO DE VERIFICACIÓN (API DIAN)")
    
    if not archivo_excel or not Path(archivo_excel).exists():
        print_error("No se proporcionó archivo válido")
        return False
    
    print_info(f"Archivo: {Path(archivo_excel).name}")
    print()
    
    try:
        # Ejecutar aplicación de DV con callback para logs
        stats = aplicar_digito_verificacion(str(archivo_excel), log_callback)
        
        if stats.get('errores', 0) == 0 or stats.get('modificadas', 0) > 0:
            return True
        else:
            return False
            
    except Exception as e:
        print_error(f"Error en DV: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Función principal - Ejecuta todo el proceso"""
    print("\n" + "╔" + "═" * 58 + "╗")
    print("║" + "  MIGRACIÓN COMPLETA CELER → MAVISO + DÍGITO VERIFICACIÓN  ".center(58) + "║")
    print("╚" + "═" * 58 + "╝")
    print()
    print(f"📅 Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Paso 1: Migración
    archivo_generado = ejecutar_migracion()
    
    if not archivo_generado:
        print()
        print_error("La migración falló. No se puede continuar.")
        input("\nPresione ENTER para salir...")
        return 1
    
    print()
    print("-" * 60)
    
    # Paso 2: Dígito de Verificación
    exito_dv = ejecutar_digito_verificacion(archivo_generado)
    
    # Resumen final
    print()
    print_header("RESUMEN FINAL")
    print_success(f"Migración: COMPLETADA")
    if exito_dv:
        print_success(f"Dígito de Verificación: APLICADO")
    else:
        print_warning(f"Dígito de Verificación: CON ERRORES")
    
    print()
    print_info(f"Archivo final: {archivo_generado}")
    print()
    
    # Detener API al finalizar
    try:
        detener_api()
    except:
        pass
    
    input("Presione ENTER para salir...")
    return 0


if __name__ == "__main__":
    sys.exit(main())
