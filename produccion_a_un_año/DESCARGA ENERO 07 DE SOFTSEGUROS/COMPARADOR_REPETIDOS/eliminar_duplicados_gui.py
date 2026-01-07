"""
GUI para Eliminar Pólizas Duplicadas en Archivos SoftSeguros
Compara dos archivos Excel por número de póliza y genera archivo sin duplicados
"""

import sys
import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QFileDialog, QMessageBox,
    QGroupBox, QProgressBar
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font as OpenpyxlFont, Border, Side, Alignment


# ==================================================================================
# ESTILOS - TEMA OSCURO PROFESIONAL
# ==================================================================================

ESTILO_DARK = """
QMainWindow {
    background-color: #1e1e1e;
}

QWidget {
    background-color: #1e1e1e;
    color: #d4d4d4;
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    font-size: 10pt;
}

QLabel {
    color: #d4d4d4;
    font-size: 10pt;
    padding: 5px;
}

QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #3a3a3a, stop:1 #2d2d2d);
    color: #ffffff;
    border: 1px solid #555555;
    border-radius: 6px;
    padding: 10px 20px;
    font-size: 11pt;
    font-weight: bold;
    min-height: 30px;
}

QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #4a4a4a, stop:1 #3d3d3d);
    border: 1px solid #0078d4;
}

QPushButton:pressed {
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 #2d2d2d, stop:1 #1e1e1e);
}

QPushButton:disabled {
    background: #2d2d2d;
    color: #666666;
    border: 1px solid #3a3a3a;
}

QTextEdit {
    background-color: #252526;
    color: #d4d4d4;
    border: 1px solid #3a3a3a;
    border-radius: 4px;
    padding: 10px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 9pt;
}

QGroupBox {
    border: 2px solid #3a3a3a;
    border-radius: 6px;
    margin-top: 10px;
    padding-top: 15px;
    font-weight: bold;
    color: #0078d4;
}

QGroupBox::title {
    subcontrol-origin: margin;
    left: 15px;
    padding: 0 5px;
}

QProgressBar {
    border: 1px solid #3a3a3a;
    border-radius: 4px;
    background-color: #252526;
    text-align: center;
    color: #d4d4d4;
    height: 25px;
}

QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 #0078d4, stop:1 #005a9e);
    border-radius: 3px;
}
"""


# ==================================================================================
# THREAD PARA ELIMINAR DUPLICADOS
# ==================================================================================

class EliminarDuplicadosThread(QThread):
    """Thread para procesar archivos y eliminar duplicados"""
    log_signal = pyqtSignal(str, str)  # mensaje, tipo
    finished_signal = pyqtSignal(bool, str, dict)  # exito, mensaje, estadisticas
    progress_signal = pyqtSignal(int)  # progreso
    
    def __init__(self, archivo1, archivo2, archivo_salida):
        super().__init__()
        self.archivo1 = archivo1
        self.archivo2 = archivo2
        self.archivo_salida = archivo_salida
    
    def run(self):
        try:
            # Leer archivos
            self.log_signal.emit("📂 Leyendo primer archivo...", "info")
            self.progress_signal.emit(10)
            df1 = pd.read_excel(self.archivo1, dtype=str)
            self.log_signal.emit(f"✅ Archivo 1: {len(df1)} registros", "success")
            
            self.log_signal.emit("📂 Leyendo segundo archivo...", "info")
            self.progress_signal.emit(30)
            df2 = pd.read_excel(self.archivo2, dtype=str)
            self.log_signal.emit(f"✅ Archivo 2: {len(df2)} registros", "success")
            
            # Obtener columna de póliza (columna A = índice 0)
            COL_POLIZA = 0
            
            # Normalizar números de póliza
            self.log_signal.emit("🔍 Identificando duplicados...", "info")
            self.progress_signal.emit(50)
            
            polizas1 = set()
            for idx, row in df1.iterrows():
                poliza = str(row.iloc[COL_POLIZA]).strip().upper()
                if poliza and poliza != 'NAN':
                    polizas1.add(poliza)
            
            polizas2 = set()
            duplicados = []
            indices_duplicados = []
            
            for idx, row in df2.iterrows():
                poliza = str(row.iloc[COL_POLIZA]).strip().upper()
                if poliza and poliza != 'NAN':
                    polizas2.add(poliza)
                    if poliza in polizas1:
                        duplicados.append(poliza)
                        indices_duplicados.append(idx)
            
            self.log_signal.emit(f"⚠️ Duplicados encontrados: {len(duplicados)}", "warning")
            
            # Mostrar algunos ejemplos
            if duplicados:
                self.log_signal.emit("", "info")
                self.log_signal.emit("📋 Ejemplos de pólizas duplicadas:", "warning")
                for poliza in duplicados[:10]:
                    self.log_signal.emit(f"   • {poliza}", "warning")
                if len(duplicados) > 10:
                    self.log_signal.emit(f"   ... y {len(duplicados) - 10} más", "warning")
            
            # Eliminar duplicados del segundo archivo
            self.log_signal.emit("", "info")
            self.log_signal.emit("🗑️ Eliminando duplicados del archivo 2...", "info")
            self.progress_signal.emit(70)
            
            df2_sin_duplicados = df2.drop(indices_duplicados)
            
            self.log_signal.emit(f"✅ Registros después de eliminar: {len(df2_sin_duplicados)}", "success")
            
            # Guardar con formato preservado
            self.log_signal.emit("", "info")
            self.log_signal.emit("💾 Guardando archivo sin duplicados...", "info")
            self.progress_signal.emit(85)
            
            # Cargar workbook original para preservar formato
            wb = load_workbook(self.archivo2)
            ws = wb.active
            
            # Eliminar filas duplicadas (de abajo hacia arriba para no alterar índices)
            for idx in sorted(indices_duplicados, reverse=True):
                # +2 porque: +1 para header, +1 porque openpyxl es 1-based
                ws.delete_rows(idx + 2)
            
            # Guardar
            wb.save(self.archivo_salida)
            self.log_signal.emit(f"✅ Guardado: {os.path.basename(self.archivo_salida)}", "success")
            
            self.progress_signal.emit(100)
            
            # Estadísticas
            estadisticas = {
                'total_archivo1': len(df1),
                'total_archivo2': len(df2),
                'duplicados': len(duplicados),
                'final': len(df2_sin_duplicados),
                'polizas_unicas_archivo1': len(polizas1),
                'polizas_unicas_archivo2': len(polizas2)
            }
            
            self.log_signal.emit("", "success")
            self.log_signal.emit("="*60, "success")
            self.log_signal.emit("✅ PROCESO COMPLETADO EXITOSAMENTE", "success")
            self.log_signal.emit("="*60, "success")
            
            self.finished_signal.emit(True, "Duplicados eliminados correctamente", estadisticas)
            
        except Exception as e:
            self.log_signal.emit(f"❌ Error: {str(e)}", "error")
            import traceback
            traceback.print_exc()
            self.finished_signal.emit(False, str(e), {})


# ==================================================================================
# VENTANA PRINCIPAL
# ==================================================================================

class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.archivo1 = None
        self.archivo2 = None
        self.initUI()
    
    def initUI(self):
        self.setWindowTitle("🗑️ Eliminar Pólizas Duplicadas - SoftSeguros")
        self.setGeometry(100, 100, 1200, 800)
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout_principal = QVBoxLayout(central_widget)
        layout_principal.setSpacing(15)
        layout_principal.setContentsMargins(20, 20, 20, 20)
        
        # Título
        titulo = QLabel("🗑️ ELIMINAR PÓLIZAS DUPLICADAS")
        titulo.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #0078d4; padding: 15px;")
        layout_principal.addWidget(titulo)
        
        # Subtítulo
        subtitulo = QLabel("Compara dos archivos Excel y elimina pólizas duplicadas del segundo archivo")
        subtitulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitulo.setStyleSheet("color: #a0a0a0; font-size: 11pt; padding-bottom: 10px;")
        layout_principal.addWidget(subtitulo)
        
        # Grupo de archivos
        grupo_archivos = QGroupBox("📂 SELECCIÓN DE ARCHIVOS")
        layout_archivos = QVBoxLayout()
        
        # Archivo 1
        layout_archivo1 = QHBoxLayout()
        self.lbl_archivo1 = QLabel("📄 Archivo 1 (Referencia):")
        self.lbl_archivo1.setMinimumWidth(180)
        self.txt_archivo1 = QLabel("No seleccionado")
        self.txt_archivo1.setStyleSheet("color: #888888; font-style: italic;")
        self.btn_archivo1 = QPushButton("📁 Seleccionar Archivo 1")
        self.btn_archivo1.clicked.connect(self.seleccionar_archivo1)
        layout_archivo1.addWidget(self.lbl_archivo1)
        layout_archivo1.addWidget(self.txt_archivo1, 1)
        layout_archivo1.addWidget(self.btn_archivo1)
        layout_archivos.addLayout(layout_archivo1)
        
        # Archivo 2
        layout_archivo2 = QHBoxLayout()
        self.lbl_archivo2 = QLabel("📄 Archivo 2 (Para limpiar):")
        self.lbl_archivo2.setMinimumWidth(180)
        self.txt_archivo2 = QLabel("No seleccionado")
        self.txt_archivo2.setStyleSheet("color: #888888; font-style: italic;")
        self.btn_archivo2 = QPushButton("📁 Seleccionar Archivo 2")
        self.btn_archivo2.clicked.connect(self.seleccionar_archivo2)
        layout_archivo2.addWidget(self.lbl_archivo2)
        layout_archivo2.addWidget(self.txt_archivo2, 1)
        layout_archivo2.addWidget(self.btn_archivo2)
        layout_archivos.addLayout(layout_archivo2)
        
        grupo_archivos.setLayout(layout_archivos)
        layout_principal.addWidget(grupo_archivos)
        
        # Botón procesar
        self.btn_procesar = QPushButton("🗑️ ELIMINAR DUPLICADOS")
        self.btn_procesar.setEnabled(False)
        self.btn_procesar.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #c84444, stop:1 #a83232);
                font-size: 13pt;
                padding: 15px;
                min-height: 40px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #d85555, stop:1 #b84343);
            }
            QPushButton:disabled {
                background: #2d2d2d;
                color: #666666;
            }
        """)
        self.btn_procesar.clicked.connect(self.procesar)
        layout_principal.addWidget(self.btn_procesar)
        
        # Barra de progreso
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout_principal.addWidget(self.progress)
        
        # Grupo de estadísticas
        grupo_stats = QGroupBox("📊 ESTADÍSTICAS")
        layout_stats = QHBoxLayout()
        
        self.stats_widgets = {}
        stats_labels = [
            ('total_archivo1', 'Archivo 1'),
            ('total_archivo2', 'Archivo 2'),
            ('duplicados', 'Duplicados'),
            ('final', 'Final')
        ]
        
        for key, label in stats_labels:
            vbox = QVBoxLayout()
            lbl_titulo = QLabel(label)
            lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_titulo.setStyleSheet("font-weight: bold; color: #0078d4;")
            lbl_valor = QLabel("0")
            lbl_valor.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_valor.setStyleSheet("font-size: 24pt; font-weight: bold; color: #4ec9b0;")
            vbox.addWidget(lbl_titulo)
            vbox.addWidget(lbl_valor)
            layout_stats.addLayout(vbox)
            self.stats_widgets[key] = lbl_valor
        
        grupo_stats.setLayout(layout_stats)
        layout_principal.addWidget(grupo_stats)
        
        # Grupo de logs
        grupo_logs = QGroupBox("📋 LOGS DE PROCESO")
        layout_logs = QVBoxLayout()
        
        self.txt_logs = QTextEdit()
        self.txt_logs.setReadOnly(True)
        self.txt_logs.setMinimumHeight(250)
        layout_logs.addWidget(self.txt_logs)
        
        # Botones de logs
        layout_botones_logs = QHBoxLayout()
        self.btn_limpiar_logs = QPushButton("🗑️ Limpiar Logs")
        self.btn_limpiar_logs.clicked.connect(self.limpiar_logs)
        self.btn_abrir_archivo = QPushButton("📂 Abrir Archivo Generado")
        self.btn_abrir_archivo.setEnabled(False)
        self.btn_abrir_archivo.clicked.connect(self.abrir_archivo)
        layout_botones_logs.addWidget(self.btn_limpiar_logs)
        layout_botones_logs.addWidget(self.btn_abrir_archivo)
        layout_logs.addLayout(layout_botones_logs)
        
        grupo_logs.setLayout(layout_logs)
        layout_principal.addWidget(grupo_logs)
        
        # Aplicar estilo
        self.setStyleSheet(ESTILO_DARK)
        
        # Log inicial
        self.log("="*60, "info")
        self.log("🗑️ ELIMINADOR DE DUPLICADOS - SOFTSEGUROS", "info")
        self.log("="*60, "info")
        self.log("📌 Instrucciones:", "info")
        self.log("1. Seleccione el archivo de referencia (Archivo 1)", "info")
        self.log("2. Seleccione el archivo a limpiar (Archivo 2)", "info")
        self.log("3. Click en 'ELIMINAR DUPLICADOS'", "info")
        self.log("4. Se generará 'archivo2_sin_duplicados.xlsx'", "info")
        self.log("="*60, "info")
    
    def seleccionar_archivo1(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Archivo 1 (Referencia)",
            "",
            "Archivos Excel (*.xlsx *.xls)"
        )
        if archivo:
            self.archivo1 = archivo
            self.txt_archivo1.setText(os.path.basename(archivo))
            self.txt_archivo1.setStyleSheet("color: #4ec9b0; font-weight: bold;")
            self.log(f"✅ Archivo 1 seleccionado: {os.path.basename(archivo)}", "success")
            self.verificar_archivos()
    
    def seleccionar_archivo2(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Archivo 2 (Para limpiar)",
            "",
            "Archivos Excel (*.xlsx *.xls)"
        )
        if archivo:
            self.archivo2 = archivo
            self.txt_archivo2.setText(os.path.basename(archivo))
            self.txt_archivo2.setStyleSheet("color: #4ec9b0; font-weight: bold;")
            self.log(f"✅ Archivo 2 seleccionado: {os.path.basename(archivo)}", "success")
            self.verificar_archivos()
    
    def verificar_archivos(self):
        if self.archivo1 and self.archivo2:
            self.btn_procesar.setEnabled(True)
            self.log("✅ Ambos archivos seleccionados - Listo para procesar", "success")
    
    def procesar(self):
        # Confirmación
        respuesta = QMessageBox.question(
            self,
            "Confirmar Operación",
            f"¿Eliminar pólizas duplicadas del archivo 2?\n\n"
            f"Archivo 1 (Referencia):\n{os.path.basename(self.archivo1)}\n\n"
            f"Archivo 2 (Para limpiar):\n{os.path.basename(self.archivo2)}\n\n"
            f"Se generará: {os.path.basename(self.archivo2).replace('.xlsx', '_sin_duplicados.xlsx')}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if respuesta != QMessageBox.StandardButton.Yes:
            return
        
        # Preparar archivo de salida
        base_dir = os.path.dirname(self.archivo2)
        nombre_base = os.path.basename(self.archivo2).replace('.xlsx', '').replace('.xls', '')
        archivo_salida = os.path.join(base_dir, f"{nombre_base}_sin_duplicados.xlsx")
        
        # Deshabilitar controles
        self.btn_procesar.setEnabled(False)
        self.btn_archivo1.setEnabled(False)
        self.btn_archivo2.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        
        # Log inicio
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log("", "info")
        self.log("="*60, "info")
        self.log(f"🚀 INICIANDO PROCESO - {timestamp}", "info")
        self.log("="*60, "info")
        self.log(f"📂 Archivo 1: {os.path.basename(self.archivo1)}", "info")
        self.log(f"📂 Archivo 2: {os.path.basename(self.archivo2)}", "info")
        self.log(f"💾 Salida: {os.path.basename(archivo_salida)}", "info")
        self.log("="*60, "info")
        
        # Iniciar thread
        self.worker = EliminarDuplicadosThread(
            self.archivo1,
            self.archivo2,
            archivo_salida
        )
        self.archivo_salida = archivo_salida
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.progress.setValue)
        self.worker.finished_signal.connect(self.proceso_completado)
        self.worker.start()
    
    def proceso_completado(self, exito, mensaje, estadisticas):
        self.progress.setVisible(False)
        self.btn_procesar.setEnabled(True)
        self.btn_archivo1.setEnabled(True)
        self.btn_archivo2.setEnabled(True)
        
        if exito:
            # Actualizar estadísticas
            for key, valor in estadisticas.items():
                if key in self.stats_widgets:
                    self.stats_widgets[key].setText(str(valor))
            
            self.btn_abrir_archivo.setEnabled(True)
            
            # Log resumen
            self.log("", "success")
            self.log("📊 RESUMEN:", "success")
            self.log(f"   Archivo 1 (Referencia): {estadisticas.get('total_archivo1', 0)} pólizas", "info")
            self.log(f"   Archivo 2 (Original): {estadisticas.get('total_archivo2', 0)} pólizas", "info")
            self.log(f"   🗑️ Duplicados eliminados: {estadisticas.get('duplicados', 0)}", "warning")
            self.log(f"   ✅ Archivo final: {estadisticas.get('final', 0)} pólizas", "success")
            self.log("="*60, "success")
            
            QMessageBox.information(
                self,
                "Éxito",
                f"✅ Proceso completado exitosamente\n\n"
                f"Duplicados eliminados: {estadisticas.get('duplicados', 0)}\n"
                f"Registros finales: {estadisticas.get('final', 0)}"
            )
        else:
            self.log(f"❌ Error: {mensaje}", "error")
            QMessageBox.critical(self, "Error", f"❌ Error al procesar:\n\n{mensaje}")
    
    def abrir_archivo(self):
        if hasattr(self, 'archivo_salida') and os.path.exists(self.archivo_salida):
            if os.name == 'nt':
                os.startfile(self.archivo_salida)
            self.log(f"📂 Abriendo: {os.path.basename(self.archivo_salida)}", "info")
    
    def limpiar_logs(self):
        self.txt_logs.clear()
    
    def log(self, mensaje: str, tipo: str = "info"):
        colores = {
            "info": "#d4d4d4",
            "success": "#6a9955",
            "warning": "#dcdcaa",
            "error": "#f44747"
        }
        color = colores.get(tipo, "#d4d4d4")
        self.txt_logs.append(f'<span style="color: {color};">{mensaje}</span>')
        self.txt_logs.verticalScrollBar().setValue(
            self.txt_logs.verticalScrollBar().maximum()
        )


# ==================================================================================
# MAIN
# ==================================================================================

def main():
    app = QApplication(sys.argv)
    ventana = VentanaPrincipal()
    ventana.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
