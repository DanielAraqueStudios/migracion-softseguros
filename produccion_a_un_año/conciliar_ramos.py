"""
Conciliador de Ramos - Maviso
=============================
Actualiza la columna de ramos (E) del archivo nuevo con los ramos
corregidos del archivo de referencia, usando la póliza (A) como clave.
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import logging

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Rutas
CARPETA_BASE = Path(__file__).parent
CARPETA_OUTPUT = CARPETA_BASE / 'output'

# Archivos
ARCHIVO_REFERENCIA = CARPETA_OUTPUT / 'Copia de Maviso_llenado.xlsx'  # Con ramos corregidos
ARCHIVO_NUEVO = CARPETA_OUTPUT / 'Maviso_llenado_20251209_153342.xlsx'  # Recién generado

# Columnas
COL_POLIZA = 1   # A - Clave para match
COL_RAMO = 5     # E - Ramo a actualizar


def conciliar_ramos():
    """
    Concilia los ramos entre el archivo de referencia y el nuevo.
    Actualiza la columna E del archivo nuevo con los ramos del archivo de referencia.
    """
    print()
    print("=" * 60)
    print("   CONCILIADOR DE RAMOS - MAVISO")
    print("=" * 60)
    print()
    
    # Verificar que existan los archivos
    if not ARCHIVO_REFERENCIA.exists():
        logger.error(f"❌ No se encuentra el archivo de referencia: {ARCHIVO_REFERENCIA}")
        return
    
    if not ARCHIVO_NUEVO.exists():
        logger.error(f"❌ No se encuentra el archivo nuevo: {ARCHIVO_NUEVO}")
        return
    
    logger.info(f"📂 Archivo referencia (ramos corregidos): {ARCHIVO_REFERENCIA.name}")
    logger.info(f"📂 Archivo a actualizar: {ARCHIVO_NUEVO.name}")
    print()
    
    # 1. Cargar archivo de referencia y crear diccionario póliza → ramo
    logger.info("Cargando archivo de referencia...")
    wb_ref = load_workbook(ARCHIVO_REFERENCIA)
    ws_ref = wb_ref.active
    
    ramos_referencia = {}
    for fila in range(2, ws_ref.max_row + 1):
        poliza = ws_ref.cell(row=fila, column=COL_POLIZA).value
        ramo = ws_ref.cell(row=fila, column=COL_RAMO).value
        
        if poliza and ramo:
            poliza_str = str(poliza).strip()
            ramos_referencia[poliza_str] = str(ramo).strip()
    
    logger.info(f"   ✅ {len(ramos_referencia)} pólizas con ramo en referencia")
    wb_ref.close()
    
    # 2. Cargar archivo nuevo
    logger.info("Cargando archivo a actualizar...")
    wb_nuevo = load_workbook(ARCHIVO_NUEVO)
    ws_nuevo = wb_nuevo.active
    
    total_filas = ws_nuevo.max_row - 1
    logger.info(f"   ✅ {total_filas} filas en archivo nuevo")
    print()
    
    # 3. Actualizar ramos
    logger.info("Actualizando ramos...")
    actualizados = 0
    sin_match = 0
    sin_cambio = 0
    diferencias = []
    
    for fila in range(2, ws_nuevo.max_row + 1):
        poliza = ws_nuevo.cell(row=fila, column=COL_POLIZA).value
        ramo_actual = ws_nuevo.cell(row=fila, column=COL_RAMO).value
        
        if poliza:
            poliza_str = str(poliza).strip()
            
            if poliza_str in ramos_referencia:
                ramo_nuevo = ramos_referencia[poliza_str]
                ramo_actual_str = str(ramo_actual).strip() if ramo_actual else ''
                
                if ramo_actual_str != ramo_nuevo:
                    # Guardar diferencia para reporte
                    diferencias.append({
                        'fila': fila,
                        'poliza': poliza_str,
                        'ramo_anterior': ramo_actual_str,
                        'ramo_nuevo': ramo_nuevo
                    })
                    
                    # Actualizar
                    ws_nuevo.cell(row=fila, column=COL_RAMO).value = ramo_nuevo
                    actualizados += 1
                else:
                    sin_cambio += 1
            else:
                sin_match += 1
    
    # 4. Guardar archivo actualizado
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    archivo_salida = CARPETA_OUTPUT / f'Maviso_conciliado_{timestamp}.xlsx'
    
    logger.info(f"Guardando archivo: {archivo_salida.name}")
    wb_nuevo.save(archivo_salida)
    wb_nuevo.close()
    
    # 5. Mostrar resumen
    print()
    print("=" * 60)
    print("   RESUMEN DE CONCILIACIÓN")
    print("=" * 60)
    print(f"   ✅ Ramos actualizados: {actualizados}")
    print(f"   ➖ Sin cambios (ya iguales): {sin_cambio}")
    print(f"   ⚠️  Sin match en referencia: {sin_match}")
    print(f"   📄 Archivo generado: {archivo_salida.name}")
    print("=" * 60)
    
    # 6. Mostrar primeras diferencias encontradas
    if diferencias:
        print()
        print("📋 Primeras 20 diferencias encontradas:")
        print("-" * 80)
        print(f"{'Fila':<6} {'Póliza':<20} {'Anterior':<25} {'Nuevo':<25}")
        print("-" * 80)
        
        for d in diferencias[:20]:
            anterior = d['ramo_anterior'][:23] + '..' if len(d['ramo_anterior']) > 25 else d['ramo_anterior']
            nuevo = d['ramo_nuevo'][:23] + '..' if len(d['ramo_nuevo']) > 25 else d['ramo_nuevo']
            print(f"{d['fila']:<6} {d['poliza']:<20} {anterior:<25} {nuevo:<25}")
        
        if len(diferencias) > 20:
            print(f"   ... y {len(diferencias) - 20} diferencias más")
    
    print()
    return archivo_salida


if __name__ == '__main__':
    conciliar_ramos()
