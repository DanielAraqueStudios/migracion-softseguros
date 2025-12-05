"""
Extraer Dropdowns de Maviso
===========================
Script para extraer las listas de valores (rangos con nombre) 
que alimentan los dropdowns de Subramo por cada Aseguradora.
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# Rutas
CARPETA_BASE = Path(__file__).parent
ARCHIVO_MAVISO = CARPETA_BASE / 'Copy of Maviso.xlsx'


def extraer_rangos_con_nombre():
    """Extrae todos los rangos con nombre definidos en el libro Excel"""
    print("=" * 80)
    print("EXTRAYENDO RANGOS CON NOMBRE (DROPDOWNS) DE MAVISO")
    print("=" * 80)
    
    wb = load_workbook(ARCHIVO_MAVISO, data_only=True)
    
    # Obtener todos los rangos con nombre definidos
    print(f"\n📋 Rangos con nombre encontrados: {len(wb.defined_names)}")
    print("-" * 80)
    
    # Diccionario para guardar aseguradora -> subramos
    aseguradora_subramos = {}
    
    for nombre in sorted(wb.defined_names):
        name = wb.defined_names[nombre]
        destino = name.attr_text  # Referencia al rango
        
        # Los nombres de aseguradoras tienen "space" en lugar de espacios
        if 'space' in nombre.lower() or nombre[0].isupper():
            # Intentar leer los valores del rango
            try:
                # El destino tiene formato: 'NombreHoja'!$A$1:$A$10
                if '!' in destino:
                    hoja_ref, rango_ref = destino.split('!')
                    hoja_ref = hoja_ref.replace("'", "")
                    
                    if hoja_ref in wb.sheetnames:
                        ws = wb[hoja_ref]
                        # Obtener valores del rango
                        valores = []
                        for row in ws[rango_ref.replace('$', '')]:
                            if isinstance(row, tuple):
                                for cell in row:
                                    if cell.value:
                                        valores.append(str(cell.value).strip())
                            else:
                                if row.value:
                                    valores.append(str(row.value).strip())
                        
                        if valores:
                            # Convertir nombre a formato legible
                            nombre_legible = nombre.replace('space', ' ')
                            aseguradora_subramos[nombre_legible] = valores
                            print(f"\n📋 {nombre_legible}")
                            print(f"   Referencia: {destino}")
                            print(f"   Valores ({len(valores)}):")
                            for v in valores:
                                print(f"      └── {v}")
            except Exception as e:
                print(f"   ⚠️  Error leyendo {nombre}: {e}")
    
    return aseguradora_subramos, wb


def listar_hojas(wb):
    """Lista todas las hojas del libro"""
    print("\n" + "=" * 80)
    print("HOJAS EN EL LIBRO MAVISO")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📄 {sheet_name}")
        print(f"   Dimensiones: {ws.dimensions}")
        # Mostrar primeras filas si no es la hoja principal
        if sheet_name != wb.active.title:
            print("   Primeras filas:")
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                valores = [str(v)[:30] if v else '' for v in row[:5]]
                print(f"      {i+1}: {valores}")


def main():
    print("\n" + "█" * 80)
    print("  ANÁLISIS DE DROPDOWNS EN MAVISO")
    print("█" * 80)
    
    aseg_subramos, wb = extraer_rangos_con_nombre()
    listar_hojas(wb)
    
    # Resumen
    print("\n" + "=" * 80)
    print("RESUMEN DE ASEGURADORAS CON SUBRAMOS CONFIGURADOS")
    print("=" * 80)
    
    if aseg_subramos:
        for aseg in sorted(aseg_subramos.keys()):
            subramos = aseg_subramos[aseg]
            print(f"\n{aseg}: {len(subramos)} subramos")
    else:
        print("\n⚠️  No se encontraron rangos con nombre para subramos.")
        print("   Los dropdowns pueden estar en una hoja separada o usar otra técnica.")
    
    print("\n" + "█" * 80)
    print("  FIN DEL ANÁLISIS")
    print("█" * 80)


if __name__ == "__main__":
    main()
