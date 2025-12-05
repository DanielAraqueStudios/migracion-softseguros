"""
Llenar Maviso desde CELER
=========================
Script para migrar datos de pólizas desde CELER hacia la plantilla Maviso,
manteniendo el formato y estilos del archivo original.

Mapeo de columnas (letra Excel):
- Maviso A ← CELER U (Póliza)
- Maviso B ← CELER AE (Placa)
- Maviso C ← CELER R (Aseguradora)
- Maviso E ← CELER S (Ramo)
- Maviso J ← CELER BE (Ejecutivos -> Nombre Vendedor)
- Maviso K ← CELER W (F_Inicio)
- Maviso L ← CELER X (F_Fin)
- Maviso O ← CELER AQ (prima sin iva)
- Maviso W ← Condicional: MENSUAL→Fraccionado, ANUAL→Contado
- Maviso X ← CELER AP (V_Asegurado)
- Maviso AB ← CELER C (Identificacion)
- Maviso AC ← CELER A (Tipo_Persona)
- Maviso AD ← CELER B (Tomador)
- Maviso AE ← CELER C (Identificacion)
- Maviso AF ← CELER AS (Asegurado)
- Maviso AG ← CELER AT (Iden_Asegurado)
- Maviso AH ← CELER AW (Beneficiario)
- Maviso AI ← CELER AX (Iden_Beneficiario)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import copy
import logging
import re
import requests
import subprocess
import sys
import os
import time

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Rutas
CARPETA_BASE = Path(__file__).parent
ARCHIVO_MAVISO = CARPETA_BASE / 'Copy of Maviso.xlsx'
ARCHIVO_CELER = CARPETA_BASE / 'Copy of polizas vigentes celer.xlsx'
CARPETA_OUTPUT = CARPETA_BASE / 'output'

# URL de la API DIAN (backend local)
API_URL = "http://localhost:8000"

# Crear carpeta output si no existe
CARPETA_OUTPUT.mkdir(exist_ok=True)

# Variable global para el proceso de la API
_api_proceso = None

# =============================================================================
# MAPEO DE ASEGURADORAS (CELER → MAVISO)
# =============================================================================
MAPEO_ASEGURADORAS = {
    'LIBERTY SEGUROS S A': 'ALLIANZ SEGUROS SA',  # Liberty = Allianz
    'ALLIANZ SEGUROS S.A': 'ALLIANZ SEGUROS SA',
    'ASEGURADORA SOLIDARIA DE COLOMBIA': 'ASEGURADORA SOLIDARIA DE COLOMBIA',
    'SURAMERICANA S.A.': 'SEGUROS GENERALES SURAMERICANA S A',
    'COMPAÑÍA MUNDIAL DE SEGUROS S A': 'SEGUROS MUNDIAL',
    'SBS SEGUROS COLOMBIA S.A': 'SBS SEGUROS COLOMBIA SA',
    'SEGUROS DEL ESTADO S A': 'SEGUROS DEL ESTADO SA',
    'HDI SEGUROS SA': 'HDI SEGUROS',
    'SEGUROS BOLIVAR': 'COMPAÑIA DE SEGUROS BOLIVAR SA',
    'AXA COLPATRIA SEGUROS S.A.': 'AXA COLPATRIA SEGUROS SA',
    'CEM': 'COOMEVA EXPERIENCIA MEDICA SAS',
    'COOMEVA': 'COOMEVA MEDICINA PREPAGADA SA',
    'LA PREVISORA S A COMPAÑÍA DE SEGUROS': 'LA PREVISORA S A COMPAÑIA DE SEGUROS',
    'ASSIST CARD': 'ASSIST CARD DE COLOMBIA SAS',
    'MAGENTA SEGUROS LTDA': 'MAGENTA ASISTANCE SAS',
    'COLMENA VIDA Y RIESGOS PROFESIONES SA': 'COLMENA SEGUROS',
    'LA EQUIDAD SEGUROS OC': 'LA EQUIDAD SEGUROS GENERALES',
    'MAPFRE SEGUROS DE COLOMBIA S A': 'MAPFRE SEGUROS GENERALES',
    'ZURICH COLOMBIA SEGUROS S.A': 'ZURICH COLOMBIA SEGUROS SA',
    'CHUBB DE COLOMBIA COMPAÑÍA SEGUROS S A': 'CHUBB SEGUROS COLOMBIA SA',
    'COMPAÑIA DE MEDICINA PREPAGADA COLSANITAS S.A': 'COMPAÑIA DE MEDICINA PREPAGADA COLSANITAS SA',
    'POSITIVA COMPAÑIA DE SEGUROS S.A.': 'POSITIVA COMPAÑIA DE SEGUROS SA',
    'ASEGURADORA GRANCOLOMBIANA S.A.': 'GRANCOLOMBIANA DE FIANZAS SAS',
    'FUNER SAN VICENTE': 'FUNERARIA SAN VICENTE SA',
    'EMERMÉDICA S.A': 'EMERMEDICA SA SERVICIOS DE AMBULANCIA PREPAGADOS',
    'MEDISANITAS': 'MEDISANITAS SAS COMPAÑIA DE MEDICINA PREPAGADA',
}

# =============================================================================
# MAPEO DE RAMOS (CELER → MAVISO SUBRAMO)
# =============================================================================
MAPEO_RAMOS = {
    # Generales
    'AUTOMOVILES': 'AUTOS INDIVIDUAL',
    'MULTIRIESGO RESIDENCIAL': 'HOGAR',
    'MULTIRIESGO EMPRESARIAL': 'MULTIRRIESGO EMPRESARIAL',
    'MI PYME': 'MI PYME',  # Corregido: era MULTIRIESGO EMPRESARIAL
    'RESPONSABILIDAD CIVIL': 'RC DERIVADA DE CUMPLIMIENTO',
    'TRANSPORTES DE MERCANCIAS': 'TRANSPORTES DE MERCANCIAS',
    'CUMPLIMIENTO': 'CUMPLIMIENTO',
    'SOAT': 'SOAT',
    'TODO RIESGO DAÑOS MATERIALES': 'TODO RIESGO DAÑOS MATERIALES',
    'MANEJO': 'MANEJO ENTIDADES FINANCIERAS',  # Corregido en Solidaria
    'MANEJO ENTIDADES FINANCIERAS': 'MANEJO ENTIDADES FINANCIERAS',
    'MAQUINARIA Y EQUIPO': 'MAQUINARIA Y EQUIPO',
    'TRANSPORTE DE VALORES': 'TRANSPORTE DE VALORES',
    'MULTIRIESGO COPROPIEDADES': 'COPROPIEDADES',
    'INCENDIO': 'MULTIRRIESGO EMPRESARIAL',
    'ARRENDAMIENTO': 'ARRENDAMIENTO',
    'PROTECCION DIGITAL': 'PROTECCION DIGITAL',
    'RC CLINICAS Y HOSPITALES': 'RC CLINICAS Y HOSPITALES',
    'AERONAVES CASCO': 'AERONAVES CASCO',  # Corregido: antes no tenía mapeo
    
    # Vida y Salud
    'ACCIDENTES PERSONALES': 'ACCIDENTES PERSONALES',
    'ACCIDENTES DE PASAJEROS': 'ACCIDENTES PERSONALES',
    'ACCIDENTES JUVENILES': 'ACCIDENTES JUVENILES',
    'ACCIDENTES ESCOLARES': 'ACCIDENTES ESCOLARES',
    'VIDA INDIVIDUAL': 'VIDA ACTUAL',
    'VIDA COLECTIVO': 'VIDA GRUPO CONTRIBUTIVA',
    'VIDA GRUPO COLECTIVO': 'VIDA GRUPO CONTRIBUTIVO',
    'SALUD FAMILIAR': 'SALUD CLASICO',
    'SALUD PARA TODOS': 'SALUD PARA TODOS',
    'SALUD COLECTIVA': 'SALUD COLECTIVA CLASICO',
    'PLAN COMPLEMENTARIO': 'PLAN COMPLEMENTARIO',
    'PLAN COMPLEMENTARIO COLECTIVO': 'PLAN COMPLEMENTARIO COLECTIVO',
    'PLAN COMPLEMENTARIO FAMILIAR': 'PLAN COMPLEMENTARIO',
    'ARL': 'ARL',
    'RENTA EDUCATIVA': 'RENTA EDUCATIVA',
    'MAS VIDA': 'MAS VIDA',
    'SEGURO EXEQUIAL': 'SEGUROS EXEQUIALES',
    'SEGUROS EXEQUIALES': 'SEGUROS EXEQUIALES',
    
    # Medicina Prepagada
    'MEDICINA PREPAGADA FAMILIAR': 'MEDICINA PREPAGADA FAMILIAR',
    'MEDICINA PREPAGADA COLECTIV': 'MEDICINA PREPAGADA COLECTIV',
    'EMERGENCIAS MÉDICAS': 'EMERGENCIAS MÉDICAS',
    'AREA PROTEGIDA': 'CEM',
    'TELEMEDICINA': 'EMERGENCIAS MÉDICAS',
    
    # Otros
    'ASIST CARD': 'ASSIST CARD',
    'RC SERVIDORES PUBLICOS': 'RC PREDIOS LABORES Y OPERACIONES',
}

# =============================================================================
# FILAS A RESALTAR EN VERDE (Requieren revisión manual)
# =============================================================================
# Combinaciones (Aseguradora, Ramo) que deben resaltarse en verde claro
RESALTAR_VERDE = {
    # ALLIANZ
    ('ALLIANZ SEGUROS S.A', 'VIDA INDIVIDUAL'),
    ('ALLIANZ SEGUROS S.A', 'VIDA COLECTIVO'),
    ('LIBERTY SEGUROS S A', 'VIDA INDIVIDUAL'),
    ('LIBERTY SEGUROS S A', 'VIDA COLECTIVO'),
    
    # SURA
    ('SURAMERICANA S.A.', 'VIDA INDIVIDUAL'),
    ('SURAMERICANA S.A.', 'SALUD FAMILIAR'),
    ('SURAMERICANA S.A.', 'VIDA GRUPO COLECTIVO'),
    
    # MUNDIAL
    ('COMPAÑÍA MUNDIAL DE SEGUROS S A', 'RESPONSABILIDAD CIVIL'),
    
    # SBS
    ('SBS SEGUROS COLOMBIA S.A', 'RESPONSABILIDAD CIVIL'),
    ('SBS SEGUROS COLOMBIA S.A', 'INCENDIO'),
    
    # SEGUROS DEL ESTADO
    ('SEGUROS DEL ESTADO S A', 'INCENDIO'),
    
    # HDI
    ('HDI SEGUROS SA', 'VIDA INDIVIDUAL'),
    
    # AXA
    ('AXA COLPATRIA SEGUROS S.A.', 'SOAT'),
    
    # COOMEVA
    ('COOMEVA', 'EMERGENCIAS MÉDICAS'),
    
    # PREVISORA
    ('LA PREVISORA S A COMPAÑÍA DE SEGUROS', 'RESPONSABILIDAD CIVIL'),
    ('LA PREVISORA S A COMPAÑÍA DE SEGUROS', 'RC SERVIDORES PUBLICOS'),
}

# Color verde claro para resaltar
VERDE_CLARO = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")


def letra_a_indice(letra):
    """Convierte letra de columna Excel a índice (0-based)"""
    resultado = 0
    for char in letra.upper():
        resultado = resultado * 26 + (ord(char) - ord('A') + 1)
    return resultado - 1


def verificar_api():
    """Verifica si la API está corriendo"""
    try:
        response = requests.get(f"{API_URL}/health", timeout=2)
        return response.status_code == 200
    except:
        return False


def matar_proceso_puerto(puerto=8000):
    """Mata cualquier proceso usando el puerto especificado (Windows)"""
    try:
        resultado = subprocess.run(
            ["netstat", "-ano"],
            capture_output=True,
            text=True,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
        )
        
        for linea in resultado.stdout.split('\n'):
            if f":{puerto}" in linea and "LISTENING" in linea:
                partes = linea.split()
                if partes:
                    pid = partes[-1]
                    if pid.isdigit():
                        subprocess.run(
                            ["taskkill", "/F", "/PID", pid],
                            capture_output=True,
                            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
                        )
                        time.sleep(1)
                        return True
    except:
        pass
    return False


def iniciar_api(callback=None):
    """
    Inicia la API de DIAN en segundo plano si no está corriendo.
    
    Args:
        callback: Función opcional para reportar mensajes (mensaje, tipo)
    
    Returns:
        bool: True si la API está lista
    """
    global _api_proceso
    
    def log(msg, tipo='info'):
        if callback:
            callback(msg, tipo)
        else:
            logger.info(msg)
    
    # Verificar si ya está corriendo
    if verificar_api():
        log("✅ API DIAN ya está corriendo", "success")
        return True
    
    log("🚀 Iniciando API DIAN...", "info")
    
    # Matar proceso anterior si existe
    matar_proceso_puerto(8000)
    time.sleep(1)
    
    # Ruta al directorio backend
    script_dir = Path(__file__).parent.parent
    backend_dir = script_dir / 'backend'
    
    if not backend_dir.exists():
        log(f"❌ No se encuentra el directorio backend: {backend_dir}", "error")
        return False
    
    try:
        # Iniciar uvicorn en segundo plano
        _api_proceso = subprocess.Popen(
            [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8000"],
            cwd=str(backend_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
        )
        
        # Esperar a que la API esté lista
        for i in range(10):
            time.sleep(1)
            if verificar_api():
                log("✅ API DIAN iniciada correctamente", "success")
                return True
            log(f"   Esperando API... ({i+1}/10)", "info")
        
        log("❌ No se pudo iniciar la API DIAN", "error")
        return False
        
    except Exception as e:
        log(f"❌ Error iniciando API: {e}", "error")
        return False


def detener_api():
    """Detiene la API si fue iniciada por este módulo"""
    global _api_proceso
    if _api_proceso:
        _api_proceso.terminate()
        _api_proceso = None
        logger.info("API DIAN detenida")


def calcular_digito_verificacion(nit):
    """
    Calcula el dígito de verificación de un NIT colombiano usando la API DIAN.
    
    Args:
        nit: Número de NIT (string o int)
    
    Returns:
        str: Dígito de verificación (0-9) o None si hay error
    """
    # Limpiar NIT - solo números
    nit_limpio = re.sub(r'[^0-9]', '', str(nit))
    
    if not nit_limpio or len(nit_limpio) > 15:
        return None
    
    try:
        response = requests.post(
            f"{API_URL}/calcular",
            json={"nit": nit_limpio},
            timeout=5
        )
        
        if response.status_code == 200:
            data = response.json()
            return str(data["digito_verificacion"])
        else:
            return None
            
    except Exception as e:
        logger.warning(f"Error API DIAN para NIT {nit}: {e}")
        return None


def aplicar_digito_verificacion(archivo_excel, callback=None):
    """
    Aplica el dígito de verificación a los NITs de personas jurídicas.
    Usa la API DIAN para calcular el DV.
    
    Args:
        archivo_excel: Ruta al archivo Excel generado
        callback: Función opcional para reportar progreso (mensaje, tipo)
    
    Returns:
        dict: Estadísticas del proceso
    """
    def log(msg, tipo='info'):
        if callback:
            callback(msg, tipo)
        else:
            logger.info(msg)
    
    log("═" * 50, "info")
    log("🔢 APLICANDO DÍGITO DE VERIFICACIÓN A NITs", "info")
    log("═" * 50, "info")
    
    # Iniciar API DIAN si no está corriendo
    if not iniciar_api(callback):
        log("❌ No se puede continuar sin la API DIAN", "error")
        return {
            'total': 0,
            'juridicas': 0,
            'naturales': 0,
            'modificadas': 0,
            'ya_tiene_dv': 0,
            'errores': 1
        }
    
    # Cargar archivo
    log(f"📂 Cargando archivo: {Path(archivo_excel).name}", "info")
    wb = load_workbook(archivo_excel)
    ws = wb.active
    
    # Columnas relevantes (1-indexed)
    COL_AB = 28  # Identificación (NIT)
    COL_AC = 29  # Tipo persona (J/N)
    
    # Estadísticas
    total_filas = ws.max_row - 1  # Excluir encabezado
    juridicas = 0
    naturales = 0
    ya_tiene_dv = 0
    modificadas = 0
    errores = 0
    
    log(f"📊 Total filas a procesar: {total_filas:,}", "info")
    log("", "info")
    
    # Procesar cada fila (empezando en 2 para saltar encabezado)
    for fila in range(2, ws.max_row + 1):
        tipo_persona = ws.cell(row=fila, column=COL_AC).value
        identificacion = ws.cell(row=fila, column=COL_AB).value
        
        # Solo procesar personas jurídicas (J)
        if tipo_persona and str(tipo_persona).strip().upper() == 'J':
            juridicas += 1
            
            if identificacion:
                nit_str = str(identificacion).strip()
                
                # Limpiar posibles .0 de números flotantes
                if nit_str.endswith('.0'):
                    nit_str = nit_str[:-2]
                
                # Verificar si ya tiene DV (formato NIT-DV)
                if '-' in nit_str:
                    ya_tiene_dv += 1
                    continue
                
                try:
                    # Calcular DV usando API DIAN
                    dv = calcular_digito_verificacion(nit_str)
                    
                    if dv:
                        # Aplicar formato NIT-DV
                        nuevo_valor = f"{nit_str}-{dv}"
                        ws.cell(row=fila, column=COL_AB).value = nuevo_valor
                        modificadas += 1
                        
                        # Mostrar primeros 10 cambios
                        if modificadas <= 10:
                            log(f"   ✓ Fila {fila}: {nit_str} → {nuevo_valor}", "info")
                    else:
                        errores += 1
                except Exception as e:
                    errores += 1
                    log(f"   ⚠️ Error fila {fila}: {e}", "warning")
        else:
            naturales += 1
        
        # Reportar progreso cada 500 filas
        if (fila - 1) % 500 == 0:
            log(f"   📝 Procesadas {fila - 1:,} / {total_filas:,} filas...", "info")
    
    # Guardar cambios
    log("", "info")
    log("💾 Guardando cambios...", "info")
    wb.save(archivo_excel)
    
    # Resumen
    log("", "info")
    log("═" * 50, "success")
    log("✅ DÍGITO DE VERIFICACIÓN APLICADO (API DIAN)", "success")
    log(f"   📊 Personas Jurídicas: {juridicas:,}", "success")
    log(f"   📊 Personas Naturales: {naturales:,}", "info")
    log(f"   ✅ NITs modificados: {modificadas:,}", "success")
    log(f"   ⏭️ Ya tenían DV: {ya_tiene_dv:,}", "info")
    if errores > 0:
        log(f"   ⚠️ Errores: {errores:,}", "warning")
    log("═" * 50, "success")
    
    return {
        'total': total_filas,
        'juridicas': juridicas,
        'naturales': naturales,
        'modificadas': modificadas,
        'ya_tiene_dv': ya_tiene_dv,
        'errores': errores
    }


def convertir_forma_pago(valor_celer):
    """
    Convierte forma de pago de CELER a formato Maviso.
    MENSUAL → Fraccionado
    ANUAL → Contado
    """
    if pd.isna(valor_celer):
        return ''
    valor = str(valor_celer).strip().upper()
    if valor == 'MENSUAL':
        return 'Fraccionado'
    elif valor == 'ANUAL':
        return 'Contado'
    else:
        # Para otros valores (SEMESTRAL, TRIMESTRAL, etc.) → Fraccionado
        return 'Fraccionado'


def main():
    logger.info("=" * 60)
    logger.info("INICIANDO MIGRACIÓN CELER → MAVISO")
    logger.info("=" * 60)
    
    # 1. Cargar archivo CELER (skiprows=3)
    logger.info(f"Cargando archivo CELER: {ARCHIVO_CELER}")
    df_celer = pd.read_excel(ARCHIVO_CELER, skiprows=3)
    logger.info(f"✅ CELER cargado: {len(df_celer)} filas, {len(df_celer.columns)} columnas")
    
    # Obtener nombres de columnas por índice
    cols_celer = df_celer.columns.tolist()
    
    # 2. Cargar archivo Maviso con openpyxl para mantener formato
    logger.info(f"Cargando plantilla Maviso: {ARCHIVO_MAVISO}")
    wb = load_workbook(ARCHIVO_MAVISO)
    ws = wb.active
    
    # Obtener encabezados de Maviso (fila 1)
    headers_maviso = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    logger.info(f"✅ Maviso cargado: {ws.max_row} filas, {ws.max_column} columnas")
    
    # 3. Guardar número de columnas antes de modificar
    num_columnas = 39  # Maviso tiene 39 columnas
    
    # 4. Copiar estilos de la fila 2 (primera fila de datos) para aplicar a todas
    estilos_fila = {}
    for col in range(1, num_columnas + 1):
        cell = ws.cell(row=2, column=col)
        estilos_fila[col] = {
            'font': copy.copy(cell.font) if cell.font else None,
            'fill': copy.copy(cell.fill) if cell.fill else None,
            'border': copy.copy(cell.border) if cell.border else None,
            'alignment': copy.copy(cell.alignment) if cell.alignment else None,
            'number_format': cell.number_format
        }
    
    # 5. Mapeo de columnas (índice 0-based de CELER → letra columna Maviso)
    mapeo = {
        'A': (20, 'Póliza'),           # U
        'B': (30, 'Placa'),            # AE
        'C': (17, 'Aseguradora'),      # R
        'E': (18, 'Ramo'),             # S
        'J': (56, 'Ejecutivos'),       # BE (Nombre del vendedor)
        'K': (22, 'F_Inicio'),         # W
        'L': (23, 'F_Fin'),            # X
        'O': (42, 'prima sin iva'),    # AQ
        'X': (41, 'V_Asegurado'),      # AP
        'AB': (2, 'Identificacion'),   # C (Documento del cliente)
        'AC': (0, 'Tipo_Persona'),     # A
        'AD': (1, 'Tomador'),          # B
        'AE': (2, 'Identificacion'),   # C (Documento del tomador)
        'AF': (44, 'Asegurado'),       # AS
        'AG': (45, 'Iden_Asegurado'),  # AT
        'AH': (48, 'Beneficiario'),    # AW
        'AI': (49, 'Iden_Beneficiario') # AX
    }
    
    # Índices de columnas CELER importantes
    idx_forma_pago_celer = 27  # AB
    idx_aseguradora_celer = 17  # R
    idx_ramo_celer = 18  # S
    
    # Columnas Maviso para resaltar
    col_c_maviso = letra_a_indice('C') + 1  # Aseguradora
    col_e_maviso = letra_a_indice('E') + 1  # Ramo
    
    logger.info("Preparando datos para migración...")
    
    # 6. Llenar datos desde CELER
    logger.info("Llenando datos desde CELER...")
    filas_procesadas = 0
    filas_resaltadas = 0
    
    for idx_celer, row_celer in df_celer.iterrows():
        fila_maviso = idx_celer + 2  # +2 porque fila 1 es encabezado y df empieza en 0
        
        # Obtener aseguradora y ramo originales de CELER
        aseguradora_celer = str(row_celer.iloc[idx_aseguradora_celer]).strip() if not pd.isna(row_celer.iloc[idx_aseguradora_celer]) else ''
        ramo_celer = str(row_celer.iloc[idx_ramo_celer]).strip() if not pd.isna(row_celer.iloc[idx_ramo_celer]) else ''
        
        # Verificar si esta fila debe resaltarse en verde
        debe_resaltar = (aseguradora_celer, ramo_celer) in RESALTAR_VERDE
        
        if debe_resaltar:
            filas_resaltadas += 1
        
        # Aplicar mapeo directo
        for letra_maviso, (idx_col_celer, descripcion) in mapeo.items():
            col_maviso = letra_a_indice(letra_maviso) + 1
            
            if idx_col_celer < len(cols_celer):
                valor = row_celer.iloc[idx_col_celer]
                
                if pd.isna(valor):
                    valor = ''
                
                # ============================================
                # APLICAR MAPEOS DE ASEGURADORA Y RAMO
                # ============================================
                if letra_maviso == 'C':  # Columna Aseguradora
                    valor_str = str(valor).strip()
                    if valor_str in MAPEO_ASEGURADORAS:
                        valor = MAPEO_ASEGURADORAS[valor_str]
                
                elif letra_maviso == 'E':  # Columna Ramo
                    valor_str = str(valor).strip()
                    if valor_str in MAPEO_RAMOS:
                        valor = MAPEO_RAMOS[valor_str]
                
                # Escribir en Maviso
                cell = ws.cell(row=fila_maviso, column=col_maviso)
                cell.value = valor
                
                # Aplicar estilos base
                if col_maviso in estilos_fila:
                    cell.font = copy.copy(estilos_fila[col_maviso]['font'])
                    cell.border = copy.copy(estilos_fila[col_maviso]['border'])
                    cell.alignment = copy.copy(estilos_fila[col_maviso]['alignment'])
                    cell.number_format = estilos_fila[col_maviso]['number_format']
                    
                    # Si debe resaltar, aplicar verde; si no, estilo original
                    if debe_resaltar:
                        cell.fill = VERDE_CLARO
                    else:
                        cell.fill = copy.copy(estilos_fila[col_maviso]['fill'])
        
        # Columna W (FORMA PAGO) - Lógica condicional
        col_w_maviso = letra_a_indice('W') + 1
        forma_pago_celer = row_celer.iloc[idx_forma_pago_celer] if idx_forma_pago_celer < len(cols_celer) else ''
        forma_pago_maviso = convertir_forma_pago(forma_pago_celer)
        
        cell_w = ws.cell(row=fila_maviso, column=col_w_maviso)
        cell_w.value = forma_pago_maviso
        if col_w_maviso in estilos_fila:
            cell_w.font = copy.copy(estilos_fila[col_w_maviso]['font'])
            cell_w.border = copy.copy(estilos_fila[col_w_maviso]['border'])
            cell_w.alignment = copy.copy(estilos_fila[col_w_maviso]['alignment'])
            cell_w.number_format = estilos_fila[col_w_maviso]['number_format']
            
            if debe_resaltar:
                cell_w.fill = VERDE_CLARO
            else:
                cell_w.fill = copy.copy(estilos_fila[col_w_maviso]['fill'])
        
        filas_procesadas += 1
        
        if filas_procesadas % 500 == 0:
            logger.info(f"  Procesadas {filas_procesadas} filas...")
    
    # 7. Guardar archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    archivo_salida = CARPETA_OUTPUT / f'Maviso_llenado_{timestamp}.xlsx'
    
    logger.info(f"Guardando archivo: {archivo_salida}")
    wb.save(archivo_salida)
    
    logger.info("=" * 60)
    logger.info("✅ MIGRACIÓN COMPLETADA")
    logger.info(f"   Filas procesadas: {filas_procesadas}")
    logger.info(f"   🟢 Filas resaltadas (revisión manual): {filas_resaltadas}")
    logger.info(f"   Archivo generado: {archivo_salida}")
    logger.info("=" * 60)
    
    return archivo_salida


if __name__ == "__main__":
    main()
