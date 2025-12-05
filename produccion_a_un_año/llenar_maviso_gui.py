"""
Llenar Maviso GUI - Interfaz Profesional Dark Mode
===================================================
Interfaz PyQt6 moderna con tema oscuro para la migración
de datos CELER → MAVISO con visualización en tiempo real.
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy
import re

# Importar funciones del backend
from llenar_maviso import calcular_digito_verificacion, aplicar_digito_verificacion

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QPushButton, QLineEdit, QTextEdit, QProgressBar,
    QFileDialog, QFrame, QScrollArea, QSplitter, QTableWidget,
    QTableWidgetItem, QHeaderView, QGroupBox, QSpacerItem, QSizePolicy,
    QGraphicsDropShadowEffect, QStackedWidget, QMessageBox, QComboBox
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QPropertyAnimation, QEasingCurve, 
    QTimer, QSize, QPoint
)
from PyQt6.QtGui import (
    QFont, QColor, QPalette, QIcon, QPainter, QBrush, QPen,
    QLinearGradient, QPixmap, QFontDatabase
)


# =============================================================================
# DARK THEME STYLESHEET
# =============================================================================
DARK_STYLESHEET = """
/* ========== GLOBAL STYLES ========== */
* {
    font-family: 'Segoe UI', 'SF Pro Display', -apple-system, sans-serif;
}

QMainWindow {
    background-color: #0d0d0d;
}

QWidget {
    background-color: transparent;
    color: #e8e8e8;
}

/* ========== FRAMES & CONTAINERS ========== */
QFrame {
    background-color: transparent;
    border: none;
}

QFrame#mainContainer {
    background-color: #0d0d0d;
}

QFrame#sidebar {
    background-color: #151515;
    border-right: 1px solid #2a2a2a;
}

QFrame#contentArea {
    background-color: #0d0d0d;
}

QFrame#card {
    background-color: #1a1a1a;
    border: 1px solid #2a2a2a;
    border-radius: 16px;
}

QFrame#cardHover:hover {
    background-color: #202020;
    border: 1px solid #3a3a3a;
}

QFrame#statsCard {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
        stop:0 #1e1e2e, stop:1 #1a1a2a);
    border: 1px solid #2d2d3d;
    border-radius: 16px;
    padding: 20px;
}

QFrame#accentCard {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
        stop:0 #0f3460, stop:1 #16213e);
    border: 1px solid #1a4b8c;
    border-radius: 16px;
}

QFrame#separator {
    background-color: #2a2a2a;
    max-height: 1px;
    min-height: 1px;
}

/* ========== LABELS ========== */
QLabel {
    color: #e8e8e8;
    background: transparent;
    padding: 0px;
}

QLabel#title {
    font-size: 28px;
    font-weight: 700;
    color: #ffffff;
    padding: 0px;
}

QLabel#subtitle {
    font-size: 14px;
    font-weight: 400;
    color: #888888;
    padding: 0px;
}

QLabel#sectionTitle {
    font-size: 18px;
    font-weight: 600;
    color: #ffffff;
    padding: 8px 0px;
}

QLabel#cardTitle {
    font-size: 13px;
    font-weight: 500;
    color: #9ca3af;
    text-transform: uppercase;
    letter-spacing: 1px;
}

QLabel#cardValue {
    font-size: 32px;
    font-weight: 700;
    color: #ffffff;
}

QLabel#cardValueSmall {
    font-size: 24px;
    font-weight: 600;
    color: #ffffff;
}

QLabel#accentValue {
    font-size: 32px;
    font-weight: 700;
    color: #60a5fa;
}

QLabel#successText {
    color: #34d399;
    font-weight: 500;
}

QLabel#errorText {
    color: #f87171;
    font-weight: 500;
}

QLabel#warningText {
    color: #fbbf24;
    font-weight: 500;
}

QLabel#mutedText {
    color: #6b7280;
    font-size: 12px;
}

/* ========== BUTTONS ========== */
QPushButton {
    background-color: #2563eb;
    color: #ffffff;
    border: none;
    border-radius: 10px;
    padding: 12px 24px;
    font-size: 14px;
    font-weight: 600;
    min-height: 20px;
}

QPushButton:hover {
    background-color: #3b82f6;
}

QPushButton:pressed {
    background-color: #1d4ed8;
}

QPushButton:disabled {
    background-color: #374151;
    color: #6b7280;
}

QPushButton#primaryButton {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #2563eb, stop:1 #7c3aed);
    min-height: 48px;
    font-size: 16px;
    border-radius: 12px;
}

QPushButton#primaryButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #3b82f6, stop:1 #8b5cf6);
}

QPushButton#primaryButton:pressed {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #1d4ed8, stop:1 #6d28d9);
}

QPushButton#secondaryButton {
    background-color: #1f2937;
    border: 1px solid #374151;
}

QPushButton#secondaryButton:hover {
    background-color: #374151;
    border: 1px solid #4b5563;
}

QPushButton#outlineButton {
    background-color: transparent;
    border: 2px solid #3b82f6;
    color: #3b82f6;
}

QPushButton#outlineButton:hover {
    background-color: rgba(59, 130, 246, 0.1);
}

QPushButton#iconButton {
    background-color: #1f2937;
    border-radius: 8px;
    padding: 8px;
    min-width: 36px;
    max-width: 36px;
    min-height: 36px;
    max-height: 36px;
}

QPushButton#iconButton:hover {
    background-color: #374151;
}

QPushButton#dangerButton {
    background-color: #dc2626;
}

QPushButton#dangerButton:hover {
    background-color: #ef4444;
}

QPushButton#successButton {
    background-color: #059669;
}

QPushButton#successButton:hover {
    background-color: #10b981;
}

QPushButton#ghostButton {
    background-color: transparent;
    color: #9ca3af;
    padding: 8px 16px;
}

QPushButton#ghostButton:hover {
    background-color: #1f2937;
    color: #ffffff;
}

/* ========== LINE EDIT ========== */
QLineEdit {
    background-color: #1f2937;
    border: 1px solid #374151;
    border-radius: 10px;
    padding: 12px 16px;
    font-size: 14px;
    color: #ffffff;
    selection-background-color: #3b82f6;
}

QLineEdit:hover {
    border: 1px solid #4b5563;
}

QLineEdit:focus {
    border: 2px solid #3b82f6;
    background-color: #1a2332;
}

QLineEdit:disabled {
    background-color: #111827;
    color: #6b7280;
    border: 1px solid #1f2937;
}

QLineEdit#searchInput {
    background-color: #1a1a1a;
    border: 1px solid #2a2a2a;
    border-radius: 12px;
    padding-left: 40px;
}

/* ========== TEXT EDIT / LOG AREA ========== */
QTextEdit {
    background-color: #111111;
    border: 1px solid #2a2a2a;
    border-radius: 12px;
    padding: 16px;
    font-family: 'Cascadia Code', 'Fira Code', 'Consolas', monospace;
    font-size: 12px;
    color: #a8b2c3;
    selection-background-color: #3b82f6;
}

QTextEdit:focus {
    border: 1px solid #3b82f6;
}

/* ========== PROGRESS BAR ========== */
QProgressBar {
    background-color: #1f2937;
    border: none;
    border-radius: 8px;
    height: 12px;
    text-align: center;
    font-size: 0px;
}

QProgressBar::chunk {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 #3b82f6, stop:1 #8b5cf6);
    border-radius: 8px;
}

QProgressBar#largeProgress {
    height: 20px;
    border-radius: 10px;
    font-size: 11px;
    color: #ffffff;
}

QProgressBar#largeProgress::chunk {
    border-radius: 10px;
}

/* ========== TABLE WIDGET ========== */
QTableWidget {
    background-color: #111111;
    alternate-background-color: #151515;
    border: 1px solid #2a2a2a;
    border-radius: 12px;
    gridline-color: #1f2937;
    selection-background-color: rgba(59, 130, 246, 0.3);
    selection-color: #ffffff;
}

QTableWidget::item {
    padding: 12px;
    border-bottom: 1px solid #1f2937;
}

QTableWidget::item:selected {
    background-color: rgba(59, 130, 246, 0.3);
}

QTableWidget::item:hover {
    background-color: #1a1a2e;
}

QHeaderView::section {
    background-color: #1a1a1a;
    color: #9ca3af;
    padding: 14px 12px;
    border: none;
    border-bottom: 2px solid #2a2a2a;
    font-weight: 600;
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

QHeaderView::section:hover {
    background-color: #202020;
}

/* ========== SCROLL BAR ========== */
QScrollBar:vertical {
    background-color: #0d0d0d;
    width: 10px;
    margin: 0px;
    border-radius: 5px;
}

QScrollBar::handle:vertical {
    background-color: #374151;
    border-radius: 5px;
    min-height: 40px;
}

QScrollBar::handle:vertical:hover {
    background-color: #4b5563;
}

QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}

QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
    background: none;
}

QScrollBar:horizontal {
    background-color: #0d0d0d;
    height: 10px;
    margin: 0px;
    border-radius: 5px;
}

QScrollBar::handle:horizontal {
    background-color: #374151;
    border-radius: 5px;
    min-width: 40px;
}

QScrollBar::handle:horizontal:hover {
    background-color: #4b5563;
}

QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0px;
}

/* ========== COMBO BOX ========== */
QComboBox {
    background-color: #1f2937;
    border: 1px solid #374151;
    border-radius: 10px;
    padding: 10px 16px;
    font-size: 14px;
    color: #ffffff;
    min-height: 20px;
}

QComboBox:hover {
    border: 1px solid #4b5563;
}

QComboBox:focus {
    border: 2px solid #3b82f6;
}

QComboBox::drop-down {
    border: none;
    width: 30px;
}

QComboBox::down-arrow {
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid #9ca3af;
    margin-right: 10px;
}

QComboBox QAbstractItemView {
    background-color: #1f2937;
    border: 1px solid #374151;
    border-radius: 8px;
    selection-background-color: #3b82f6;
    padding: 4px;
}

/* ========== GROUP BOX ========== */
QGroupBox {
    background-color: #1a1a1a;
    border: 1px solid #2a2a2a;
    border-radius: 12px;
    margin-top: 16px;
    padding: 20px;
    padding-top: 36px;
    font-weight: 600;
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 16px;
    padding: 0 8px;
    color: #9ca3af;
    font-size: 13px;
    text-transform: uppercase;
    letter-spacing: 1px;
}

/* ========== TOOL TIP ========== */
QToolTip {
    background-color: #1f2937;
    color: #ffffff;
    border: 1px solid #374151;
    border-radius: 6px;
    padding: 8px 12px;
    font-size: 12px;
}

/* ========== MESSAGE BOX ========== */
QMessageBox {
    background-color: #1a1a1a;
}

QMessageBox QLabel {
    color: #e8e8e8;
    font-size: 14px;
}

QMessageBox QPushButton {
    min-width: 80px;
}

/* ========== SPLITTER ========== */
QSplitter::handle {
    background-color: #2a2a2a;
}

QSplitter::handle:horizontal {
    width: 2px;
}

QSplitter::handle:vertical {
    height: 2px;
}

QSplitter::handle:hover {
    background-color: #3b82f6;
}
"""


# =============================================================================
# WORKER THREAD FOR MIGRATION
# =============================================================================
class MigrationWorker(QThread):
    """Thread worker para ejecutar la migración sin bloquear la UI"""
    
    progress = pyqtSignal(int)
    log_message = pyqtSignal(str, str)  # mensaje, tipo (info/success/error/warning)
    stats_update = pyqtSignal(dict)
    finished_signal = pyqtSignal(bool, str)  # éxito, mensaje/ruta
    
    def __init__(self, archivo_celer, archivo_maviso, carpeta_output):
        super().__init__()
        self.archivo_celer = Path(archivo_celer)
        self.archivo_maviso = Path(archivo_maviso)
        self.carpeta_output = Path(carpeta_output)
        self._is_running = True
    
    def stop(self):
        self._is_running = False
    
    def letra_a_indice(self, letra):
        """Convierte letra de columna Excel a índice (0-based)"""
        resultado = 0
        for char in letra.upper():
            resultado = resultado * 26 + (ord(char) - ord('A') + 1)
        return resultado - 1
    
    def convertir_forma_pago(self, valor_celer):
        """Convierte forma de pago de CELER a formato Maviso"""
        if pd.isna(valor_celer):
            return ''
        valor = str(valor_celer).strip().upper()
        if valor == 'MENSUAL':
            return 'Fraccionado'
        elif valor == 'ANUAL':
            return 'Contado'
        else:
            return 'Fraccionado'
    
    def run(self):
        try:
            self.log_message.emit("═" * 50, "info")
            self.log_message.emit("🚀 INICIANDO MIGRACIÓN CELER → MAVISO", "info")
            self.log_message.emit("═" * 50, "info")
            
            # 1. Cargar CELER
            self.log_message.emit(f"📂 Cargando archivo CELER...", "info")
            df_celer = pd.read_excel(self.archivo_celer, skiprows=3)
            total_filas = len(df_celer)
            self.log_message.emit(f"✅ CELER cargado: {total_filas:,} filas", "success")
            
            self.stats_update.emit({
                'total_filas': total_filas,
                'columnas_celer': len(df_celer.columns)
            })
            
            cols_celer = df_celer.columns.tolist()
            
            # 2. Cargar Maviso
            self.log_message.emit(f"📂 Cargando plantilla Maviso...", "info")
            wb = load_workbook(self.archivo_maviso)
            ws = wb.active
            self.log_message.emit(f"✅ Maviso cargado: {ws.max_row:,} filas, {ws.max_column} columnas", "success")
            
            self.stats_update.emit({
                'columnas_maviso': ws.max_column
            })
            
            # 3. Copiar estilos
            self.log_message.emit("🎨 Copiando estilos de formato...", "info")
            num_columnas = 39
            estilos_fila = {}
            for col in range(1, num_columnas + 1):
                cell = ws.cell(row=2, column=col)
                estilos_fila[col] = {
                    'font': copy.copy(cell.font) if cell.font else None,
                    'fill': copy.copy(cell.fill) if cell.fill else None,
                    'border': copy.copy(cell.border) if cell.border else None,
                    'alignment': copy.copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format
                }
            
            # 4. Mapeo de columnas
            mapeo = {
                'A': (20, 'Póliza'),
                'B': (30, 'Placa'),
                'C': (17, 'Aseguradora'),
                'E': (18, 'Ramo'),
                'J': (56, 'Ejecutivos'),
                'K': (22, 'F_Inicio'),
                'L': (23, 'F_Fin'),
                'O': (42, 'prima sin iva'),
                'X': (41, 'V_Asegurado'),
                'AB': (2, 'Identificacion'),
                'AC': (0, 'Tipo_Persona'),
                'AD': (1, 'Tomador'),
                'AE': (2, 'Identificacion'),
                'AF': (44, 'Asegurado'),
                'AG': (45, 'Iden_Asegurado'),
                'AH': (48, 'Beneficiario'),
                'AI': (49, 'Iden_Beneficiario')
            }
            
            idx_forma_pago_celer = 27
            
            # 5. Procesar filas
            self.log_message.emit("", "info")
            self.log_message.emit("📊 Procesando datos...", "info")
            filas_procesadas = 0
            
            for idx_celer, row_celer in df_celer.iterrows():
                if not self._is_running:
                    self.log_message.emit("⚠️ Migración cancelada por el usuario", "warning")
                    self.finished_signal.emit(False, "Cancelado")
                    return
                
                fila_maviso = idx_celer + 2
                
                # Aplicar mapeo
                for letra_maviso, (idx_col_celer, descripcion) in mapeo.items():
                    col_maviso = self.letra_a_indice(letra_maviso) + 1
                    
                    if idx_col_celer < len(cols_celer):
                        valor = row_celer.iloc[idx_col_celer]
                        if pd.isna(valor):
                            valor = ''
                        
                        cell = ws.cell(row=fila_maviso, column=col_maviso)
                        cell.value = valor
                        
                        if col_maviso in estilos_fila:
                            cell.font = copy.copy(estilos_fila[col_maviso]['font'])
                            cell.fill = copy.copy(estilos_fila[col_maviso]['fill'])
                            cell.border = copy.copy(estilos_fila[col_maviso]['border'])
                            cell.alignment = copy.copy(estilos_fila[col_maviso]['alignment'])
                            cell.number_format = estilos_fila[col_maviso]['number_format']
                
                # Columna W (Forma Pago)
                col_w_maviso = self.letra_a_indice('W') + 1
                forma_pago_celer = row_celer.iloc[idx_forma_pago_celer] if idx_forma_pago_celer < len(cols_celer) else ''
                forma_pago_maviso = self.convertir_forma_pago(forma_pago_celer)
                
                cell_w = ws.cell(row=fila_maviso, column=col_w_maviso)
                cell_w.value = forma_pago_maviso
                if col_w_maviso in estilos_fila:
                    cell_w.font = copy.copy(estilos_fila[col_w_maviso]['font'])
                    cell_w.fill = copy.copy(estilos_fila[col_w_maviso]['fill'])
                    cell_w.border = copy.copy(estilos_fila[col_w_maviso]['border'])
                    cell_w.alignment = copy.copy(estilos_fila[col_w_maviso]['alignment'])
                    cell_w.number_format = estilos_fila[col_w_maviso]['number_format']
                
                filas_procesadas += 1
                
                # Actualizar progreso
                progreso = int((filas_procesadas / total_filas) * 100)
                self.progress.emit(progreso)
                
                if filas_procesadas % 500 == 0:
                    self.log_message.emit(f"   📝 Procesadas {filas_procesadas:,} / {total_filas:,} filas...", "info")
                    self.stats_update.emit({'procesadas': filas_procesadas})
            
            # 6. Guardar archivo
            self.log_message.emit("", "info")
            self.log_message.emit("💾 Guardando archivo...", "info")
            
            self.carpeta_output.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archivo_salida = self.carpeta_output / f'Maviso_llenado_{timestamp}.xlsx'
            
            wb.save(archivo_salida)
            
            self.log_message.emit("", "info")
            self.log_message.emit("═" * 50, "success")
            self.log_message.emit("✅ MIGRACIÓN COMPLETADA EXITOSAMENTE", "success")
            self.log_message.emit(f"   📁 Archivo: {archivo_salida.name}", "success")
            self.log_message.emit(f"   📊 Total filas: {filas_procesadas:,}", "success")
            self.log_message.emit("═" * 50, "success")
            
            self.stats_update.emit({
                'procesadas': filas_procesadas,
                'completado': True
            })
            
            self.finished_signal.emit(True, str(archivo_salida))
            
        except Exception as e:
            self.log_message.emit("", "error")
            self.log_message.emit(f"❌ ERROR: {str(e)}", "error")
            self.finished_signal.emit(False, str(e))


class DVWorker(QThread):
    """Thread worker para aplicar dígito de verificación a NITs"""
    
    log_message = pyqtSignal(str, str)
    finished_signal = pyqtSignal(bool, dict)  # éxito, estadísticas
    
    def __init__(self, archivo_excel):
        super().__init__()
        self.archivo_excel = archivo_excel
    
    def run(self):
        try:
            # Usar la función del backend con callback para logs
            def log_callback(msg, tipo):
                self.log_message.emit(msg, tipo)
            
            stats = aplicar_digito_verificacion(self.archivo_excel, log_callback)
            self.finished_signal.emit(True, stats)
            
        except Exception as e:
            self.log_message.emit("", "error")
            self.log_message.emit(f"❌ ERROR: {str(e)}", "error")
            self.finished_signal.emit(False, {'error': str(e)})


# =============================================================================
# CUSTOM WIDGETS
# =============================================================================
class AnimatedButton(QPushButton):
    """Botón con animaciones suaves"""
    
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self._animation = None
    
    def enterEvent(self, event):
        super().enterEvent(event)
    
    def leaveEvent(self, event):
        super().leaveEvent(event)


class StatsCard(QFrame):
    """Card para mostrar estadísticas"""
    
    def __init__(self, title, value="0", icon="📊", accent=False, parent=None):
        super().__init__(parent)
        self.setObjectName("accentCard" if accent else "statsCard")
        self.setMinimumHeight(120)
        self.setup_ui(title, value, icon)
        self.add_shadow()
    
    def setup_ui(self, title, value, icon):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(8)
        
        # Icono y título
        header = QHBoxLayout()
        
        icon_label = QLabel(icon)
        icon_label.setStyleSheet("font-size: 24px;")
        header.addWidget(icon_label)
        
        title_label = QLabel(title)
        title_label.setObjectName("cardTitle")
        header.addWidget(title_label)
        header.addStretch()
        
        layout.addLayout(header)
        
        # Valor
        self.value_label = QLabel(value)
        self.value_label.setObjectName("cardValue")
        layout.addWidget(self.value_label)
        
        layout.addStretch()
    
    def add_shadow(self):
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setOffset(0, 4)
        shadow.setColor(QColor(0, 0, 0, 80))
        self.setGraphicsEffect(shadow)
    
    def set_value(self, value):
        self.value_label.setText(str(value))


class FileSelector(QFrame):
    """Selector de archivo estilizado"""
    
    file_selected = pyqtSignal(str)
    
    def __init__(self, label, placeholder="Seleccionar archivo...", file_filter="", parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.file_filter = file_filter
        self.setup_ui(label, placeholder)
    
    def setup_ui(self, label, placeholder):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(12)
        
        # Label
        title = QLabel(label)
        title.setObjectName("cardTitle")
        layout.addWidget(title)
        
        # Input row
        input_row = QHBoxLayout()
        input_row.setSpacing(12)
        
        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText(placeholder)
        self.path_input.setReadOnly(True)
        input_row.addWidget(self.path_input)
        
        browse_btn = QPushButton("📁 Examinar")
        browse_btn.setObjectName("secondaryButton")
        browse_btn.setFixedWidth(120)
        browse_btn.clicked.connect(self.browse_file)
        input_row.addWidget(browse_btn)
        
        layout.addLayout(input_row)
    
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo", "", self.file_filter
        )
        if file_path:
            self.path_input.setText(file_path)
            self.file_selected.emit(file_path)
    
    def get_path(self):
        return self.path_input.text()
    
    def set_path(self, path):
        self.path_input.setText(path)


class FolderSelector(QFrame):
    """Selector de carpeta estilizado"""
    
    folder_selected = pyqtSignal(str)
    
    def __init__(self, label, placeholder="Seleccionar carpeta...", parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setup_ui(label, placeholder)
    
    def setup_ui(self, label, placeholder):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(12)
        
        title = QLabel(label)
        title.setObjectName("cardTitle")
        layout.addWidget(title)
        
        input_row = QHBoxLayout()
        input_row.setSpacing(12)
        
        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText(placeholder)
        self.path_input.setReadOnly(True)
        input_row.addWidget(self.path_input)
        
        browse_btn = QPushButton("📂 Examinar")
        browse_btn.setObjectName("secondaryButton")
        browse_btn.setFixedWidth(120)
        browse_btn.clicked.connect(self.browse_folder)
        input_row.addWidget(browse_btn)
        
        layout.addLayout(input_row)
    
    def browse_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Seleccionar carpeta")
        if folder_path:
            self.path_input.setText(folder_path)
            self.folder_selected.emit(folder_path)
    
    def get_path(self):
        return self.path_input.text()
    
    def set_path(self, path):
        self.path_input.setText(path)


class LogViewer(QFrame):
    """Visor de logs estilizado"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header
        header = QFrame()
        header.setStyleSheet("""
            background-color: #1a1a1a;
            border-bottom: 1px solid #2a2a2a;
            border-radius: 12px 12px 0 0;
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(20, 12, 20, 12)
        
        title = QLabel("📋 Registro de Actividad")
        title.setObjectName("sectionTitle")
        title.setStyleSheet("font-size: 14px; font-weight: 600;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        clear_btn = QPushButton("Limpiar")
        clear_btn.setObjectName("ghostButton")
        clear_btn.clicked.connect(self.clear_log)
        header_layout.addWidget(clear_btn)
        
        layout.addWidget(header)
        
        # Log area
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                border-radius: 0 0 12px 12px;
                border-top: none;
            }
        """)
        layout.addWidget(self.log_text)
    
    def append_log(self, message, msg_type="info"):
        color_map = {
            "info": "#a8b2c3",
            "success": "#34d399",
            "error": "#f87171",
            "warning": "#fbbf24"
        }
        color = color_map.get(msg_type, "#a8b2c3")
        
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted = f'<span style="color: #6b7280;">[{timestamp}]</span> <span style="color: {color};">{message}</span>'
        
        self.log_text.append(formatted)
        
        # Auto-scroll
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def clear_log(self):
        self.log_text.clear()


# =============================================================================
# MAIN WINDOW
# =============================================================================
class MainWindow(QMainWindow):
    """Ventana principal de la aplicación"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SoftSeguros - Migración CELER → MAVISO")
        self.setMinimumSize(1200, 800)
        self.resize(1400, 900)
        
        # Variables
        self.worker = None
        self.dv_worker = None
        self.ultimo_archivo_generado = None
        self.carpeta_base = Path(__file__).parent
        
        self.setup_ui()
        self.load_defaults()
    
    def setup_ui(self):
        """Configurar la interfaz de usuario"""
        
        # Widget central
        central_widget = QWidget()
        central_widget.setObjectName("mainContainer")
        self.setCentralWidget(central_widget)
        
        # Layout principal
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # ===== SIDEBAR =====
        sidebar = self.create_sidebar()
        main_layout.addWidget(sidebar)
        
        # ===== CONTENT AREA =====
        content = self.create_content_area()
        main_layout.addWidget(content, 1)
    
    def create_sidebar(self):
        """Crear barra lateral"""
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(280)
        
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(20, 24, 20, 24)
        layout.setSpacing(16)
        
        # Logo / Título
        logo_container = QVBoxLayout()
        logo_container.setSpacing(4)
        
        logo = QLabel("🔄")
        logo.setStyleSheet("font-size: 48px;")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_container.addWidget(logo)
        
        title = QLabel("SoftSeguros")
        title.setObjectName("title")
        title.setStyleSheet("font-size: 22px;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_container.addWidget(title)
        
        subtitle = QLabel("Migración CELER → MAVISO")
        subtitle.setObjectName("subtitle")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_container.addWidget(subtitle)
        
        layout.addLayout(logo_container)
        
        # Separador
        sep = QFrame()
        sep.setObjectName("separator")
        layout.addWidget(sep)
        
        layout.addSpacing(8)
        
        # Estadísticas
        stats_title = QLabel("📊 ESTADÍSTICAS")
        stats_title.setObjectName("cardTitle")
        layout.addWidget(stats_title)
        
        # Stats cards (compactas para sidebar)
        self.stat_total = self.create_mini_stat("Total Filas", "0", "#3b82f6")
        layout.addWidget(self.stat_total)
        
        self.stat_procesadas = self.create_mini_stat("Procesadas", "0", "#8b5cf6")
        layout.addWidget(self.stat_procesadas)
        
        self.stat_estado = self.create_mini_stat("Estado", "Listo", "#10b981")
        layout.addWidget(self.stat_estado)
        
        layout.addStretch()
        
        # Versión
        version = QLabel("v1.0.0 • PyQt6")
        version.setObjectName("mutedText")
        version.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(version)
        
        return sidebar
    
    def create_mini_stat(self, label, value, color):
        """Crear estadística compacta para sidebar"""
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background-color: #1a1a1a;
                border: 1px solid #2a2a2a;
                border-radius: 10px;
                border-left: 3px solid {color};
            }}
        """)
        
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(14, 12, 14, 12)
        
        label_widget = QLabel(label)
        label_widget.setStyleSheet("color: #9ca3af; font-size: 12px;")
        layout.addWidget(label_widget)
        
        value_widget = QLabel(value)
        value_widget.setStyleSheet(f"color: {color}; font-size: 16px; font-weight: 700;")
        value_widget.setAlignment(Qt.AlignmentFlag.AlignRight)
        layout.addWidget(value_widget)
        
        frame.value_label = value_widget
        return frame
    
    def create_content_area(self):
        """Crear área de contenido principal"""
        content = QFrame()
        content.setObjectName("contentArea")
        
        layout = QVBoxLayout(content)
        layout.setContentsMargins(32, 24, 32, 24)
        layout.setSpacing(24)
        
        # ===== HEADER =====
        header = QHBoxLayout()
        
        header_text = QVBoxLayout()
        title = QLabel("Migración de Datos")
        title.setObjectName("title")
        header_text.addWidget(title)
        
        subtitle = QLabel("Transfiere datos de pólizas desde CELER hacia la plantilla MAVISO")
        subtitle.setObjectName("subtitle")
        header_text.addWidget(subtitle)
        
        header.addLayout(header_text)
        header.addStretch()
        
        # Botón de ayuda
        help_btn = QPushButton("❓ Ayuda")
        help_btn.setObjectName("ghostButton")
        help_btn.clicked.connect(self.show_help)
        header.addWidget(help_btn)
        
        layout.addLayout(header)
        
        # ===== CONFIGURACIÓN =====
        config_section = QVBoxLayout()
        config_section.setSpacing(16)
        
        config_title = QLabel("⚙️ Configuración de Archivos")
        config_title.setObjectName("sectionTitle")
        config_section.addWidget(config_title)
        
        # Grid de selectores
        selectors_grid = QGridLayout()
        selectors_grid.setSpacing(16)
        
        self.celer_selector = FileSelector(
            "📥 ARCHIVO CELER (ORIGEN)",
            "Seleccionar archivo Excel de CELER...",
            "Excel Files (*.xlsx *.xls)"
        )
        selectors_grid.addWidget(self.celer_selector, 0, 0)
        
        self.maviso_selector = FileSelector(
            "📄 PLANTILLA MAVISO (DESTINO)",
            "Seleccionar plantilla Maviso...",
            "Excel Files (*.xlsx)"
        )
        selectors_grid.addWidget(self.maviso_selector, 0, 1)
        
        self.output_selector = FolderSelector(
            "📂 CARPETA DE SALIDA",
            "Seleccionar carpeta para guardar resultado..."
        )
        selectors_grid.addWidget(self.output_selector, 1, 0, 1, 2)
        
        config_section.addLayout(selectors_grid)
        layout.addLayout(config_section)
        
        # ===== PROGRESO =====
        progress_section = QVBoxLayout()
        progress_section.setSpacing(12)
        
        progress_header = QHBoxLayout()
        progress_title = QLabel("📈 Progreso de Migración")
        progress_title.setObjectName("sectionTitle")
        progress_header.addWidget(progress_title)
        
        progress_header.addStretch()
        
        self.progress_percent = QLabel("0%")
        self.progress_percent.setStyleSheet("color: #3b82f6; font-size: 18px; font-weight: 700;")
        progress_header.addWidget(self.progress_percent)
        
        progress_section.addLayout(progress_header)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("largeProgress")
        self.progress_bar.setValue(0)
        progress_section.addWidget(self.progress_bar)
        
        layout.addLayout(progress_section)
        
        # ===== LOG VIEWER =====
        self.log_viewer = LogViewer()
        self.log_viewer.setMinimumHeight(250)
        layout.addWidget(self.log_viewer, 1)
        
        # ===== BOTONES DE ACCIÓN =====
        actions = QHBoxLayout()
        actions.setSpacing(16)
        
        self.btn_cancel = QPushButton("✖️ Cancelar")
        self.btn_cancel.setObjectName("dangerButton")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self.cancel_migration)
        actions.addWidget(self.btn_cancel)
        
        actions.addStretch()
        
        # Botón para aplicar Dígito de Verificación
        self.btn_apply_dv = QPushButton("🔢 Aplicar DV a NITs")
        self.btn_apply_dv.setObjectName("secondaryButton")
        self.btn_apply_dv.setToolTip("Calcula y aplica el dígito de verificación a los NITs de personas jurídicas (columna AB)")
        self.btn_apply_dv.setEnabled(False)
        self.btn_apply_dv.clicked.connect(self.apply_dv)
        actions.addWidget(self.btn_apply_dv)
        
        self.btn_open_output = QPushButton("📂 Abrir Carpeta")
        self.btn_open_output.setObjectName("secondaryButton")
        self.btn_open_output.clicked.connect(self.open_output_folder)
        actions.addWidget(self.btn_open_output)
        
        self.btn_start = QPushButton("🚀 Iniciar Migración")
        self.btn_start.setObjectName("primaryButton")
        self.btn_start.setMinimumWidth(200)
        self.btn_start.clicked.connect(self.start_migration)
        actions.addWidget(self.btn_start)
        
        layout.addLayout(actions)
        
        return content
    
    def load_defaults(self):
        """Cargar rutas por defecto"""
        archivo_celer = self.carpeta_base / 'Copy of polizas vigentes celer.xlsx'
        archivo_maviso = self.carpeta_base / 'Copy of Maviso.xlsx'
        carpeta_output = self.carpeta_base / 'output'
        
        if archivo_celer.exists():
            self.celer_selector.set_path(str(archivo_celer))
        
        if archivo_maviso.exists():
            self.maviso_selector.set_path(str(archivo_maviso))
        
        self.output_selector.set_path(str(carpeta_output))
        
        self.log_viewer.append_log("🎉 Aplicación iniciada correctamente", "success")
        self.log_viewer.append_log("📌 Rutas por defecto cargadas", "info")
    
    def start_migration(self):
        """Iniciar el proceso de migración"""
        # Validar archivos
        celer_path = self.celer_selector.get_path()
        maviso_path = self.maviso_selector.get_path()
        output_path = self.output_selector.get_path()
        
        if not celer_path or not Path(celer_path).exists():
            self.show_error("Por favor seleccione un archivo CELER válido")
            return
        
        if not maviso_path or not Path(maviso_path).exists():
            self.show_error("Por favor seleccione una plantilla Maviso válida")
            return
        
        if not output_path:
            self.show_error("Por favor seleccione una carpeta de salida")
            return
        
        # Actualizar UI
        self.btn_start.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.progress_bar.setValue(0)
        self.update_stat(self.stat_estado, "Procesando...", "#fbbf24")
        
        # Crear y ejecutar worker
        self.worker = MigrationWorker(celer_path, maviso_path, output_path)
        self.worker.progress.connect(self.update_progress)
        self.worker.log_message.connect(self.log_viewer.append_log)
        self.worker.stats_update.connect(self.update_stats)
        self.worker.finished_signal.connect(self.on_migration_finished)
        self.worker.start()
    
    def cancel_migration(self):
        """Cancelar la migración en curso"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log_viewer.append_log("⚠️ Cancelando migración...", "warning")
    
    def update_progress(self, value):
        """Actualizar barra de progreso"""
        self.progress_bar.setValue(value)
        self.progress_percent.setText(f"{value}%")
    
    def update_stats(self, stats):
        """Actualizar estadísticas"""
        if 'total_filas' in stats:
            self.update_stat(self.stat_total, f"{stats['total_filas']:,}", "#3b82f6")
        
        if 'procesadas' in stats:
            self.update_stat(self.stat_procesadas, f"{stats['procesadas']:,}", "#8b5cf6")
    
    def update_stat(self, stat_widget, value, color):
        """Actualizar valor de una estadística"""
        stat_widget.value_label.setText(str(value))
        stat_widget.value_label.setStyleSheet(f"color: {color}; font-size: 16px; font-weight: 700;")
    
    def on_migration_finished(self, success, message):
        """Callback cuando termina la migración"""
        self.btn_start.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        
        if success:
            self.ultimo_archivo_generado = message
            self.update_stat(self.stat_estado, "✅ Completado", "#10b981")
            self.progress_bar.setValue(100)
            self.progress_percent.setText("100%")
            
            # Habilitar botón de DV
            self.btn_apply_dv.setEnabled(True)
            
            # Mostrar mensaje de éxito
            QMessageBox.information(
                self,
                "Migración Exitosa",
                f"✅ La migración se completó correctamente.\n\n"
                f"📁 Archivo generado:\n{Path(message).name}\n\n"
                f"💡 Puede aplicar el Dígito de Verificación a los NITs\n"
                f"usando el botón '🔢 Aplicar DV a NITs'"
            )
        else:
            self.update_stat(self.stat_estado, "❌ Error", "#ef4444")
            if message != "Cancelado":
                QMessageBox.critical(
                    self,
                    "Error en Migración",
                    f"❌ Ocurrió un error durante la migración:\n\n{message}"
                )
    
    def open_output_folder(self):
        """Abrir carpeta de salida en el explorador"""
        output_path = self.output_selector.get_path()
        if output_path and Path(output_path).exists():
            os.startfile(output_path)
        else:
            self.show_error("La carpeta de salida no existe")
    
    def show_error(self, message):
        """Mostrar mensaje de error"""
        QMessageBox.warning(self, "Advertencia", message)
    
    def apply_dv(self):
        """Aplicar dígito de verificación a NITs de personas jurídicas"""
        if not self.ultimo_archivo_generado or not Path(self.ultimo_archivo_generado).exists():
            self.show_error("No hay archivo generado para procesar.\nPrimero ejecute la migración.")
            return
        
        # Confirmar acción
        reply = QMessageBox.question(
            self,
            "Confirmar Acción",
            f"¿Desea aplicar el Dígito de Verificación a los NITs?\n\n"
            f"Esto modificará la columna AB del archivo:\n{Path(self.ultimo_archivo_generado).name}\n\n"
            f"• Personas Jurídicas (J): Se calculará el DV\n"
            f"• Personas Naturales (N): Sin cambios",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Deshabilitar botones durante proceso
        self.btn_apply_dv.setEnabled(False)
        self.btn_start.setEnabled(False)
        self.update_stat(self.stat_estado, "Aplicando DV...", "#fbbf24")
        
        # Ejecutar en thread separado
        self.dv_worker = DVWorker(self.ultimo_archivo_generado)
        self.dv_worker.log_message.connect(self.log_viewer.append_log)
        self.dv_worker.finished_signal.connect(self.on_dv_finished)
        self.dv_worker.start()
    
    def on_dv_finished(self, success, stats):
        """Callback cuando termina el proceso de DV"""
        self.btn_start.setEnabled(True)
        
        if success:
            self.update_stat(self.stat_estado, "✅ DV Aplicado", "#10b981")
            self.btn_apply_dv.setEnabled(False)  # Ya no se necesita
            
            QMessageBox.information(
                self,
                "Dígito de Verificación Aplicado",
                f"✅ Proceso completado exitosamente\n\n"
                f"📊 Personas Jurídicas: {stats.get('juridicas', 0):,}\n"
                f"📊 Personas Naturales: {stats.get('naturales', 0):,}\n"
                f"✅ NITs modificados: {stats.get('modificadas', 0):,}\n"
                f"⏭️ Ya tenían DV: {stats.get('ya_tiene_dv', 0):,}"
            )
        else:
            self.update_stat(self.stat_estado, "❌ Error DV", "#ef4444")
            self.btn_apply_dv.setEnabled(True)
            QMessageBox.critical(
                self,
                "Error en Dígito de Verificación",
                f"❌ Ocurrió un error:\n\n{stats.get('error', 'Error desconocido')}"
            )
    
    def show_help(self):
        """Mostrar diálogo de ayuda"""
        help_text = """
<h2>🔄 Migración CELER → MAVISO</h2>

<h3>📋 Descripción</h3>
<p>Esta herramienta permite migrar datos de pólizas desde archivos Excel 
del sistema CELER hacia la plantilla del sistema MAVISO (SoftSeguros).</p>

<h3>📁 Archivos Requeridos</h3>
<ul>
    <li><b>Archivo CELER:</b> Excel con datos de pólizas vigentes (skiprows=3)</li>
    <li><b>Plantilla Maviso:</b> Archivo Excel destino con formato predefinido</li>
</ul>

<h3>🔄 Mapeo de Columnas</h3>
<table style="margin-left: 20px;">
    <tr><td>• Maviso A</td><td>← CELER U (Póliza)</td></tr>
    <tr><td>• Maviso C</td><td>← CELER R (Aseguradora)</td></tr>
    <tr><td>• Maviso E</td><td>← CELER S (Ramo)</td></tr>
    <tr><td>• Maviso K-L</td><td>← CELER W-X (Fechas)</td></tr>
    <tr><td>• Maviso O</td><td>← CELER AQ (Prima)</td></tr>
    <tr><td>• Maviso AD</td><td>← CELER B (Tomador)</td></tr>
</table>

<h3>⚡ Forma de Pago</h3>
<ul>
    <li>MENSUAL → Fraccionado</li>
    <li>ANUAL → Contado</li>
</ul>

<h3>📞 Soporte</h3>
<p>Para asistencia técnica, contacte al equipo de desarrollo.</p>
        """
        
        msg = QMessageBox(self)
        msg.setWindowTitle("Ayuda - Migración CELER → MAVISO")
        msg.setTextFormat(Qt.TextFormat.RichText)
        msg.setText(help_text)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()


# =============================================================================
# MAIN
# =============================================================================
def main():
    # Configurar DPI awareness para pantallas de alta resolución
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.ApplicationAttribute.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.ApplicationAttribute.AA_UseHighDpiPixmaps, True)
    
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    # Aplicar stylesheet
    app.setStyleSheet(DARK_STYLESHEET)
    
    # Crear y mostrar ventana
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
