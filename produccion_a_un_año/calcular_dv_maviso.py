#!/usr/bin/env python3
"""
Calculador de Dígitos de Verificación para Maviso
=================================================
Programa interactivo que calcula el DV de NITs usando la API DIAN.

REGLA DE NEGOCIO:
- Columna AC = "J" (Persona Jurídica) → SE CALCULA el DV
- Columna AC = "N" (Persona Natural) → NO se calcula

Solo modifica la columna AB (Identificación).
"""

import pandas as pd
import re
import os
import sys
import subprocess
import time
import requests
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# URL de la API
API_URL = "http://localhost:8000"


class CalculadorDVMaviso:
    """Calculador de dígitos de verificación para archivos Maviso"""

    def __init__(self):
        self.wb = None
        self.ws = None
        self.archivo_entrada = None
        self.api_proceso = None

    def limpiar_pantalla(self):
        """Limpia la pantalla de la consola"""
        os.system('cls' if os.name == 'nt' else 'clear')

    def mostrar_banner(self):
        """Muestra el banner del programa"""
        print()
        print("╔" + "═" * 58 + "╗")
        print("║" + "  CALCULADOR DV - MAVISO (API DIAN)  ".center(58) + "║")
        print("║" + "  Personas Jurídicas (J) únicamente  ".center(58) + "║")
        print("╚" + "═" * 58 + "╝")
        print()

    def verificar_api(self):
        """Verifica si la API está corriendo"""
        try:
            response = requests.get(f"{API_URL}/health", timeout=2)
            return response.status_code == 200
        except:
            return False

    def matar_proceso_puerto(self, puerto=8000):
        """Mata cualquier proceso usando el puerto especificado (Windows)"""
        try:
            resultado = subprocess.run(
                ["netstat", "-ano"],
                capture_output=True,
                text=True,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            for linea in resultado.stdout.split('\n'):
                if f":{puerto}" in linea and "LISTENING" in linea:
                    partes = linea.split()
                    if partes:
                        pid = partes[-1]
                        if pid.isdigit():
                            subprocess.run(
                                ["taskkill", "/F", "/PID", pid],
                                capture_output=True,
                                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
                            )
                            time.sleep(1)
                            return True
        except:
            pass
        return False

    def iniciar_api(self):
        """Inicia la API de DIAN en segundo plano"""
        print("\n[0] INICIANDO API DIAN")
        print("-" * 50)
        
        # Primero verificar si ya está corriendo
        if self.verificar_api():
            print("  ✅ API ya está corriendo")
            return True
        
        # Liberar el puerto 8000 si está ocupado
        print("  🔍 Verificando puerto 8000...")
        if self.matar_proceso_puerto(8000):
            print("  🧹 Puerto 8000 liberado (proceso anterior terminado)")
            time.sleep(1)  # Esperar a que se libere completamente
        else:
            print("  ✅ Puerto 8000 disponible")
        
        print("  🚀 Iniciando servidor API...")
        
        # Ruta al directorio backend
        script_dir = Path(__file__).parent.parent
        backend_dir = script_dir / 'backend'
        
        if not backend_dir.exists():
            print(f"  ❌ No se encuentra el directorio backend: {backend_dir}")
            return False
        
        try:
            # Iniciar uvicorn en segundo plano
            self.api_proceso = subprocess.Popen(
                [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8000"],
                cwd=str(backend_dir),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            # Esperar a que la API esté lista
            for i in range(10):
                time.sleep(1)
                if self.verificar_api():
                    print("  ✅ API iniciada correctamente")
                    return True
                print(f"     Esperando... ({i+1}/10)")
            
            print("  ❌ No se pudo iniciar la API")
            return False
            
        except Exception as e:
            print(f"  ❌ Error iniciando API: {e}")
            return False

    def detener_api(self):
        """Detiene la API si fue iniciada por este programa"""
        if self.api_proceso:
            self.api_proceso.terminate()
            print("\n  ℹ️  API detenida")

    def calcular_digito_verificacion(self, nit, reintentos=3):
        """
        Calcula el dígito de verificación usando la API de DIAN.
        
        Args:
            nit: NIT sin dígito de verificación (solo números)
            reintentos: Número de reintentos en caso de error
            
        Returns:
            str: Dígito de verificación calculado (0-9) o None si hay error
        """
        # Limpiar el NIT - solo números
        nit_str = re.sub(r'\D', '', str(nit))
        
        if not nit_str or len(nit_str) > 15:
            return None

        for intento in range(reintentos):
            try:
                # Llamar a la API
                response = requests.post(
                    f"{API_URL}/calcular",
                    json={"nit": nit_str},
                    timeout=10
                )
                
                if response.status_code == 200:
                    data = response.json()
                    return str(data["digito_verificacion"])
                else:
                    # Si falla con código de error, reintentar
                    if intento < reintentos - 1:
                        time.sleep(0.5)
                        continue
                    return None

            except requests.exceptions.ConnectionError:
                # La API se desconectó - intentar reiniciarla
                if intento < reintentos - 1:
                    print(f"\n     🔄 Reconectando API (intento {intento + 2}/{reintentos})...")
                    if not self.verificar_api():
                        # Reiniciar la API
                        self.matar_proceso_puerto(8000)
                        time.sleep(1)
                        self._reiniciar_api_silencioso()
                        time.sleep(2)
                    continue
                return None
            except requests.exceptions.Timeout:
                if intento < reintentos - 1:
                    time.sleep(1)
                    continue
                return None
            except Exception as e:
                if intento < reintentos - 1:
                    time.sleep(0.5)
                    continue
                return None
        
        return None

    def _reiniciar_api_silencioso(self):
        """Reinicia la API sin mensajes verbosos"""
        script_dir = Path(__file__).parent.parent
        backend_dir = script_dir / 'backend'
        
        if backend_dir.exists():
            try:
                self.api_proceso = subprocess.Popen(
                    [sys.executable, "-m", "uvicorn", "app:app", "--host", "127.0.0.1", "--port", "8000"],
                    cwd=str(backend_dir),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
                )
                # Esperar a que inicie
                for _ in range(5):
                    time.sleep(1)
                    if self.verificar_api():
                        print("     ✅ API reconectada")
                        return True
            except:
                pass
        return False

    def seleccionar_archivo(self):
        """Permite al usuario seleccionar el archivo Excel"""
        print("\n[1] SELECCIONAR ARCHIVO MAVISO")
        print("-" * 50)
        
        # Mostrar archivos disponibles en la carpeta output
        output_dir = Path(__file__).parent / 'output'
        if output_dir.exists():
            archivos = list(output_dir.glob('*.xlsx'))
            if archivos:
                print("\n  📁 Archivos en carpeta output:")
                for i, f in enumerate(archivos[-10:], 1):  # Mostrar últimos 10
                    print(f"     {i}. {f.name}")
                print()
        
        while True:
            archivo = input("  Ingrese la ruta del archivo Excel (o 'q' para salir): ").strip()
            
            if archivo.lower() == 'q':
                return False
            
            # Si ingresa un número, usar archivo de la lista
            if archivo.isdigit() and output_dir.exists():
                archivos = list(output_dir.glob('*.xlsx'))[-10:]
                idx = int(archivo) - 1
                if 0 <= idx < len(archivos):
                    archivo = str(archivos[idx])
            
            # Remover comillas si las tiene
            archivo = archivo.strip('"').strip("'")
            
            if not os.path.exists(archivo):
                print(f"  ❌ El archivo no existe: {archivo}")
                continue
            
            if not archivo.endswith(('.xlsx', '.xls', '.xlsm')):
                print("  ❌ El archivo debe ser Excel (.xlsx, .xls, .xlsm)")
                continue
            
            try:
                self.wb = load_workbook(archivo)
                self.ws = self.wb.active
                self.archivo_entrada = archivo
                
                print(f"\n  ✅ Archivo cargado: {os.path.basename(archivo)}")
                print(f"  ✅ Total de filas: {self.ws.max_row:,}")
                print(f"  ✅ Total de columnas: {self.ws.max_column}")
                return True
                
            except Exception as e:
                print(f"  ❌ No se pudo leer el archivo: {e}")
                continue

    def mostrar_preview(self):
        """Muestra una vista previa de las columnas relevantes"""
        print("\n[2] VISTA PREVIA DE DATOS")
        print("-" * 50)
        
        # Columnas relevantes
        COL_AB = 28  # Identificación
        COL_AC = 29  # Tipo persona
        
        print("\n  Columnas a procesar:")
        print(f"     • Columna AB (col {COL_AB}): Identificación (NIT)")
        print(f"     • Columna AC (col {COL_AC}): Tipo persona (J/N)")
        print()
        print("  Muestra de datos (primeras 10 filas con datos):")
        print("  " + "-" * 46)
        print(f"  {'Fila':<6} {'Identificación':<20} {'Tipo':<6}")
        print("  " + "-" * 46)
        
        mostradas = 0
        for fila in range(2, min(self.ws.max_row + 1, 100)):
            identificacion = self.ws.cell(row=fila, column=COL_AB).value
            tipo_persona = self.ws.cell(row=fila, column=COL_AC).value
            
            if identificacion:
                tipo_str = str(tipo_persona).strip().upper() if tipo_persona else "?"
                print(f"  {fila:<6} {str(identificacion)[:20]:<20} {tipo_str:<6}")
                mostradas += 1
                if mostradas >= 10:
                    break
        
        print("  " + "-" * 46)
        print()
        
        # Contar tipos
        juridicas = 0
        naturales = 0
        otros = 0
        
        for fila in range(2, self.ws.max_row + 1):
            tipo = self.ws.cell(row=fila, column=COL_AC).value
            if tipo:
                tipo_str = str(tipo).strip().upper()
                if tipo_str == 'J':
                    juridicas += 1
                elif tipo_str == 'N':
                    naturales += 1
                else:
                    otros += 1
        
        print("  📊 Resumen de tipos de persona:")
        print(f"     • Jurídicas (J): {juridicas:,} ← SE CALCULARÁN")
        print(f"     • Naturales (N): {naturales:,} ← Sin cambios")
        if otros > 0:
            print(f"     • Otros: {otros:,} ← Sin cambios")
        print()
        
        return True

    def confirmar_proceso(self):
        """Pide confirmación antes de procesar"""
        print("\n[3] CONFIRMACIÓN")
        print("-" * 50)
        print()
        print("  ⚠️  REGLAS DE NEGOCIO:")
        print("     • Solo se modificará la columna AB (Identificación)")
        print("     • Solo se procesarán las filas con tipo 'J' (Jurídica)")
        print("     • Formato resultante: NIT-DV (ej: 900123456-7)")
        print()
        
        respuesta = input("  ¿Desea continuar? (s/n): ").strip().lower()
        return respuesta == 's'

    def procesar_archivo(self):
        """Procesa el archivo aplicando DV a personas jurídicas"""
        print("\n[4] PROCESANDO NITs")
        print("-" * 50)
        
        # Columnas relevantes (1-indexed para openpyxl)
        COL_AB = 28  # Identificación
        COL_AC = 29  # Tipo persona
        
        total_filas = self.ws.max_row - 1  # Excluir encabezado
        juridicas = 0
        naturales = 0
        modificadas = 0
        ya_tiene_dv = 0
        errores = 0
        
        print(f"\n  Procesando {total_filas:,} filas...")
        print()
        
        for fila in range(2, self.ws.max_row + 1):
            tipo_persona = self.ws.cell(row=fila, column=COL_AC).value
            identificacion = self.ws.cell(row=fila, column=COL_AB).value
            
            # Solo procesar personas jurídicas (J)
            if tipo_persona and str(tipo_persona).strip().upper() == 'J':
                juridicas += 1
                
                if identificacion:
                    nit_str = str(identificacion).strip()
                    
                    # Limpiar posibles .0 de números flotantes
                    if nit_str.endswith('.0'):
                        nit_str = nit_str[:-2]
                    
                    # Verificar si ya tiene DV (formato NIT-DV)
                    if '-' in nit_str:
                        ya_tiene_dv += 1
                        print(f"     ⏭️  Fila {fila}: {nit_str} (ya tiene DV)")
                        continue
                    
                    # Solo procesar si es numérico válido
                    nit_limpio = re.sub(r'\D', '', nit_str)
                    if not nit_limpio or len(nit_limpio) < 5:
                        errores += 1
                        print(f"     ⚠️  Fila {fila}: {nit_str} (NIT inválido)")
                        continue
                    
                    try:
                        # Calcular DV usando API DIAN
                        dv = self.calcular_digito_verificacion(nit_limpio)
                        
                        if dv:
                            # Aplicar formato NIT-DV
                            nuevo_valor = f"{nit_limpio}-{dv}"
                            self.ws.cell(row=fila, column=COL_AB).value = nuevo_valor
                            modificadas += 1
                            
                            # Mostrar TODOS los cambios
                            print(f"     ✓ Fila {fila}: {nit_str} → {nuevo_valor}")
                        else:
                            errores += 1
                            print(f"     ⚠️ Fila {fila}: Error calculando DV para {nit_str}")
                    except Exception as e:
                        errores += 1
                        print(f"     ❌ Fila {fila}: Excepción - {e}")
            else:
                naturales += 1
            
            # Reportar progreso cada 500 filas
            if (fila - 1) % 500 == 0:
                print(f"     📝 Procesadas {fila - 1:,} / {total_filas:,} filas...")
        
        # Resumen
        print()
        print("  " + "=" * 46)
        print("  RESUMEN DEL PROCESO")
        print("  " + "=" * 46)
        print(f"     📊 Total filas:           {total_filas:,}")
        print(f"     👔 Personas Jurídicas:    {juridicas:,}")
        print(f"     👤 Personas Naturales:    {naturales:,}")
        print(f"     ✅ NITs modificados:      {modificadas:,}")
        print(f"     ⏭️  Ya tenían DV:          {ya_tiene_dv:,}")
        if errores > 0:
            print(f"     ⚠️  Errores:              {errores:,}")
        print("  " + "=" * 46)
        
        return {
            'total': total_filas,
            'juridicas': juridicas,
            'naturales': naturales,
            'modificadas': modificadas,
            'ya_tiene_dv': ya_tiene_dv,
            'errores': errores
        }

    def guardar_archivo(self, stats):
        """Guarda el archivo modificado"""
        print("\n[5] GUARDAR ARCHIVO")
        print("-" * 50)
        
        if stats['modificadas'] == 0:
            print("\n  ℹ️  No hubo modificaciones, no es necesario guardar.")
            return False
        
        # Generar nombre de archivo de salida
        base = Path(self.archivo_entrada)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_salida = base.parent / f"{base.stem}_con_DV_{timestamp}{base.suffix}"
        
        print(f"\n  📄 Archivo de salida: {archivo_salida.name}")
        print()
        
        respuesta = input("  ¿Guardar archivo? (s/n): ").strip().lower()
        
        if respuesta == 's':
            try:
                self.wb.save(archivo_salida)
                print(f"\n  ✅ Archivo guardado exitosamente!")
                print(f"  📁 Ruta: {archivo_salida}")
                return True
            except Exception as e:
                print(f"\n  ❌ No se pudo guardar: {e}")
                return False
        else:
            print("\n  ℹ️  Archivo no guardado")
            return False

    def ejecutar(self):
        """Ejecuta el programa interactivo"""
        self.limpiar_pantalla()
        self.mostrar_banner()
        
        # Iniciar la API
        if not self.iniciar_api():
            print("\n  ❌ No se puede continuar sin la API")
            print("     Por favor inicie la API manualmente:")
            print("     cd backend; python -m uvicorn app:app --reload")
            input("\n  Presione ENTER para salir...")
            return
        
        try:
            # Paso 1: Seleccionar archivo
            if not self.seleccionar_archivo():
                print("\n  ℹ️  Programa terminado por el usuario")
                return
            
            # Paso 2: Vista previa
            self.mostrar_preview()
            
            # Paso 3: Confirmar
            if not self.confirmar_proceso():
                print("\n  ℹ️  Proceso cancelado por el usuario")
                return
            
            # Paso 4: Procesar
            stats = self.procesar_archivo()
            
            # Paso 5: Guardar
            self.guardar_archivo(stats)
            
            print()
            print("╔" + "═" * 58 + "╗")
            print("║" + "  ✅ PROCESO COMPLETADO  ".center(58) + "║")
            print("╚" + "═" * 58 + "╝")
            print()
        
        finally:
            self.detener_api()
            input("  Presione ENTER para salir...")


def main():
    """Función principal"""
    calculador = CalculadorDVMaviso()
    calculador.ejecutar()


if __name__ == "__main__":
    main()
