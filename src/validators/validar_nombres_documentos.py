"""
Validación de Coincidencia Nombre-Documento
===========================================
Este script compara los nombres asociados a cada número de documento
entre SOFTSEGUROS y CELER, detectando inconsistencias incluso con errores
de escritura usando algoritmos de similitud de texto.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import logging
from difflib import SequenceMatcher
import re

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ValidadorNombresDocumentos:
    """Valida la coincidencia entre nombres y documentos en ambas bases"""
    
    def __init__(self, archivo_softseguros, archivo_celer):
        self.archivo_softseguros = archivo_softseguros
        self.archivo_celer = archivo_celer
        self.df_soft = None
        self.df_celer = None
        self.inconsistencias = []
        
    def cargar_datos(self):
        """Carga ambos archivos Excel"""
        logger.info("Cargando archivos...")
        try:
            self.df_soft = pd.read_excel(self.archivo_softseguros)
            self.df_celer = pd.read_excel(self.archivo_celer)
            logger.info(f"✅ SOFTSEGUROS: {len(self.df_soft)} registros")
            logger.info(f"✅ CELER: {len(self.df_celer)} registros")
        except Exception as e:
            logger.error(f"❌ Error al cargar archivos: {e}")
            raise
    
    def normalizar_texto(self, texto):
        """Normaliza texto para comparación: mayúsculas, sin tildes, sin espacios extras"""
        if pd.isna(texto):
            return ""
        
        texto = str(texto).upper().strip()
        
        # Remover tildes
        replacements = {
            'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
            'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U',
            'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
            'Ñ': 'N'
        }
        for orig, repl in replacements.items():
            texto = texto.replace(orig, repl)
        
        # Remover caracteres especiales y múltiples espacios
        texto = re.sub(r'[^\w\s]', ' ', texto)
        texto = re.sub(r'\s+', ' ', texto)
        
        return texto.strip()
    
    def limpiar_identificacion(self, identificacion):
        """Limpia identificación para comparación"""
        if pd.isna(identificacion):
            return None
        return re.sub(r'[^\d]', '', str(identificacion))
    
    def calcular_similitud(self, texto1, texto2):
        """Calcula similitud entre dos textos (0.0 a 1.0)"""
        texto1_norm = self.normalizar_texto(texto1)
        texto2_norm = self.normalizar_texto(texto2)
        
        if not texto1_norm or not texto2_norm:
            return 0.0
        
        return SequenceMatcher(None, texto1_norm, texto2_norm).ratio()
    
    def construir_nombre_completo_soft(self, row):
        """Construye nombre completo desde SOFTSEGUROS"""
        nombres = str(row['NOMBRES']) if not pd.isna(row['NOMBRES']) else ""
        apellidos = str(row['APELLIDOS']) if not pd.isna(row['APELLIDOS']) else ""
        
        nombre_completo = f"{nombres} {apellidos}".strip()
        return nombre_completo if nombre_completo else nombres
    
    def preparar_datos(self):
        """Prepara DataFrames con campos normalizados"""
        logger.info("\n=== PREPARANDO DATOS ===")
        
        # SOFTSEGUROS
        self.df_soft['ID_LIMPIO'] = self.df_soft['NÚMERO DE DOCUMENTO'].apply(
            self.limpiar_identificacion
        )
        self.df_soft['NOMBRE_COMPLETO'] = self.df_soft.apply(
            self.construir_nombre_completo_soft, axis=1
        )
        self.df_soft['NOMBRE_NORMALIZADO'] = self.df_soft['NOMBRE_COMPLETO'].apply(
            self.normalizar_texto
        )
        
        # CELER
        self.df_celer['ID_LIMPIO'] = self.df_celer['Identificacion'].apply(
            self.limpiar_identificacion
        )
        self.df_celer['NOMBRE_NORMALIZADO'] = self.df_celer['Tomador'].apply(
            self.normalizar_texto
        )
        
        # Filtrar registros con ID válido
        self.df_soft = self.df_soft[self.df_soft['ID_LIMPIO'].notna()].copy()
        self.df_celer = self.df_celer[self.df_celer['ID_LIMPIO'].notna()].copy()
        
        logger.info(f"Registros válidos SOFTSEGUROS: {len(self.df_soft)}")
        logger.info(f"Registros válidos CELER: {len(self.df_celer)}")
    
    def validar_coincidencias(self, umbral_similitud=0.85):
        """
        Valida coincidencias entre nombres y documentos
        umbral_similitud: 0.0 a 1.0, donde 1.0 es coincidencia exacta
        """
        logger.info(f"\n=== VALIDANDO COINCIDENCIAS (umbral: {umbral_similitud}) ===")
        
        # Agrupar SOFTSEGUROS por ID (puede haber duplicados)
        soft_agrupado = self.df_soft.groupby('ID_LIMPIO').agg({
            'NOMBRE_COMPLETO': 'first',
            'NOMBRE_NORMALIZADO': 'first',
            'TIPO DE DOCUMENTO': 'first'
        }).reset_index()
        
        # Agrupar CELER por ID
        celer_agrupado = self.df_celer.groupby('ID_LIMPIO').agg({
            'Tomador': 'first',
            'NOMBRE_NORMALIZADO': 'first',
            'Tipo_Doc': 'first'
        }).reset_index()
        
        # Encontrar IDs en común
        ids_comunes = set(soft_agrupado['ID_LIMPIO']) & set(celer_agrupado['ID_LIMPIO'])
        logger.info(f"IDs en común entre ambas bases: {len(ids_comunes)}")
        
        coincidencias_exactas = 0
        similitudes_altas = 0
        inconsistencias_detectadas = 0
        
        for id_doc in ids_comunes:
            # Obtener nombres de cada base
            soft_row = soft_agrupado[soft_agrupado['ID_LIMPIO'] == id_doc].iloc[0]
            celer_row = celer_agrupado[celer_agrupado['ID_LIMPIO'] == id_doc].iloc[0]
            
            nombre_soft = soft_row['NOMBRE_NORMALIZADO']
            nombre_celer = celer_row['NOMBRE_NORMALIZADO']
            
            # Calcular similitud
            similitud = self.calcular_similitud(nombre_soft, nombre_celer)
            
            if similitud == 1.0:
                coincidencias_exactas += 1
            elif similitud >= umbral_similitud:
                similitudes_altas += 1
            else:
                # Inconsistencia detectada
                inconsistencias_detectadas += 1
                self.inconsistencias.append({
                    'id_documento': id_doc,
                    'tipo_doc_soft': soft_row['TIPO DE DOCUMENTO'],
                    'tipo_doc_celer': celer_row['Tipo_Doc'],
                    'nombre_softseguros': soft_row['NOMBRE_COMPLETO'],
                    'nombre_celer': celer_row['Tomador'],
                    'similitud': round(similitud, 3),
                    'problema': self._clasificar_problema(similitud)
                })
        
        logger.info(f"\n📊 RESULTADOS:")
        logger.info(f"  ✅ Coincidencias exactas: {coincidencias_exactas}")
        logger.info(f"  ⚠️  Similitudes altas (>{umbral_similitud}): {similitudes_altas}")
        logger.info(f"  ❌ Inconsistencias detectadas: {inconsistencias_detectadas}")
        
        if inconsistencias_detectadas > 0:
            logger.info(f"\n🔍 MUESTRA DE INCONSISTENCIAS:")
            for inc in self.inconsistencias[:5]:
                logger.info(f"\n  ID: {inc['id_documento']} ({inc['tipo_doc_soft']})")
                logger.info(f"  SOFTSEGUROS: {inc['nombre_softseguros'][:60]}")
                logger.info(f"  CELER:       {inc['nombre_celer'][:60]}")
                logger.info(f"  Similitud: {inc['similitud']} - {inc['problema']}")
    
    def _clasificar_problema(self, similitud):
        """Clasifica el tipo de problema según la similitud"""
        if similitud >= 0.7:
            return "ERROR MENOR DE ESCRITURA"
        elif similitud >= 0.5:
            return "DIFERENCIA MODERADA"
        elif similitud >= 0.3:
            return "DIFERENCIA SIGNIFICATIVA"
        else:
            return "NOMBRES COMPLETAMENTE DIFERENTES"
    
    def validar_consistencia_interna(self):
        """Valida que un mismo ID no tenga múltiples nombres diferentes en cada base"""
        logger.info("\n=== VALIDANDO CONSISTENCIA INTERNA ===")
        
        # SOFTSEGUROS - buscar IDs con múltiples nombres
        soft_duplicados = self.df_soft.groupby('ID_LIMPIO').agg({
            'NOMBRE_NORMALIZADO': lambda x: x.nunique()
        }).reset_index()
        
        soft_problemas = soft_duplicados[soft_duplicados['NOMBRE_NORMALIZADO'] > 1]
        
        if len(soft_problemas) > 0:
            logger.warning(f"⚠️  SOFTSEGUROS: {len(soft_problemas)} IDs con múltiples nombres")
            for _, row in soft_problemas.head(3).iterrows():
                id_doc = row['ID_LIMPIO']
                nombres = self.df_soft[self.df_soft['ID_LIMPIO'] == id_doc]['NOMBRE_COMPLETO'].tolist()
                logger.warning(f"  ID {id_doc}:")
                for nombre in set(nombres):
                    logger.warning(f"    - {nombre}")
        else:
            logger.info("✅ SOFTSEGUROS: No hay IDs con múltiples nombres")
        
        # CELER - buscar IDs con múltiples nombres
        celer_duplicados = self.df_celer.groupby('ID_LIMPIO').agg({
            'NOMBRE_NORMALIZADO': lambda x: x.nunique()
        }).reset_index()
        
        celer_problemas = celer_duplicados[celer_duplicados['NOMBRE_NORMALIZADO'] > 1]
        
        if len(celer_problemas) > 0:
            logger.warning(f"⚠️  CELER: {len(celer_problemas)} IDs con múltiples nombres")
        else:
            logger.info("✅ CELER: No hay IDs con múltiples nombres (correcto)")
    
    def generar_reporte(self, ruta_salida='data/output'):
        """Genera reporte detallado de inconsistencias"""
        logger.info("\n=== GENERANDO REPORTE ===")
        
        Path(ruta_salida).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_reporte = f"{ruta_salida}/VALIDACION_NOMBRES_DOCUMENTOS_{timestamp}.xlsx"
        
        with pd.ExcelWriter(archivo_reporte, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            resumen = [
                ['VALIDACIÓN NOMBRE-DOCUMENTO', ''],
                ['Fecha', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['', ''],
                ['IDs analizados', len(set(self.df_soft['ID_LIMPIO']) & set(self.df_celer['ID_LIMPIO']))],
                ['Inconsistencias detectadas', len(self.inconsistencias)],
                ['', ''],
                ['DISTRIBUCIÓN POR TIPO DE PROBLEMA', ''],
            ]
            
            # Contar por tipo de problema
            if self.inconsistencias:
                df_inc = pd.DataFrame(self.inconsistencias)
                conteo_problemas = df_inc['problema'].value_counts()
                for problema, cantidad in conteo_problemas.items():
                    resumen.append([problema, cantidad])
            
            df_resumen = pd.DataFrame(resumen, columns=['Descripción', 'Valor'])
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja 2: Inconsistencias
            if self.inconsistencias:
                df_inconsistencias = pd.DataFrame(self.inconsistencias)
                # Ordenar por similitud ascendente (peores primero)
                df_inconsistencias = df_inconsistencias.sort_values('similitud')
                df_inconsistencias.to_excel(writer, sheet_name='Inconsistencias', index=False)
                
                # Aplicar formato condicional
                worksheet = writer.sheets['Inconsistencias']
                
                # Colorear según similitud
                from openpyxl.styles import PatternFill
                
                for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df_inconsistencias)+1), start=2):
                    similitud_cell = row[5]  # Columna de similitud
                    if similitud_cell.value < 0.3:
                        # Rojo para diferencias críticas
                        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif similitud_cell.value < 0.5:
                        # Naranja para diferencias significativas
                        fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    else:
                        # Amarillo para errores menores
                        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    
                    for cell in row:
                        cell.fill = fill
            
            # Aplicar formato a headers
            from openpyxl.styles import Font, Alignment
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-ajustar columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
        
        logger.info(f"✅ Reporte generado: {archivo_reporte}")
        return archivo_reporte


def main():
    """Función principal"""
    # Usar el archivo corregido de SOFTSEGUROS
    archivo_softseguros = 'data/output/CLIENTES_SOFTSEGUROS_CORREGIDO_20251104_130714.xlsx'
    archivo_celer = 'CLIENTES VIGENTES CELER.xlsx'
    
    logger.info("="*60)
    logger.info("VALIDACIÓN DE NOMBRES Y DOCUMENTOS")
    logger.info("="*60)
    
    # Crear validador
    validador = ValidadorNombresDocumentos(archivo_softseguros, archivo_celer)
    
    # Cargar datos
    validador.cargar_datos()
    
    # Preparar datos
    validador.preparar_datos()
    
    # Validar consistencia interna
    validador.validar_consistencia_interna()
    
    # Validar coincidencias entre bases
    validador.validar_coincidencias(umbral_similitud=0.85)
    
    # Generar reporte
    archivo_reporte = validador.generar_reporte()
    
    logger.info("\n" + "="*60)
    logger.info("✅ VALIDACIÓN COMPLETADA")
    logger.info("="*60)
    logger.info(f"📊 Reporte: {archivo_reporte}")


if __name__ == "__main__":
    main()
